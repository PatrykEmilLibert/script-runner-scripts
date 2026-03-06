import xml.etree.ElementTree as ET
import csv
import os
import re
import urllib.request
import customtkinter as ctk
from tkinter import filedialog, messagebox
from datetime import datetime
from urllib.parse import urlparse
import tempfile
import time
import concurrent.futures  # <--- DODANO: Do wielowątkowości
import openpyxl  # <--- DODANO: Do obsługi plików XLSX
from openpyxl.utils import get_column_letter # <--- DODANO: Do auto-dopasowania kolumn

# --- Ustawienia ---
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")
DOMYSLNA_SCIEZKA_ZAPISU = os.path.join(os.path.expanduser("~"), "Downloads")
MAX_WORKERS = 10  # <--- DODANO: Liczba jednoczesnych pobrań
ROZSZERZENIA = ['manoFR', 'manoDE', 'manoIT', 'manoES', 'manoFR_extra', 'manoDE_extra', 'manoIT_extra', 'manoES_extra']  # Rozszerzenia do przetwarzania
EXCEL_MAX_LEN = 32767
_ILLEGAL_CTRL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

KOLOR_TLA_GLOWNY = "#FFF9FC"
KOLOR_TLA_SEKCJI = "#FFFFFF"
KOLOR_AKCENT = "hotpink"
KOLOR_AKCENT_HOVER = "#FF1493"
KOLOR_TEKSTU = "#2b2b2b"
KOLOR_STATUSU = "#6b7280"
# ------------------

def sanitize_for_excel(value):
    """Usuwa nielegalne znaki i przycina tekst do limitu Excela (32 767)."""
    if value is None:
        return ""
    text = str(value)
    if _ILLEGAL_CTRL_CHARS.search(text):
        text = _ILLEGAL_CTRL_CHARS.sub(" ", text)
    if len(text) > EXCEL_MAX_LEN:
        text = text[:EXCEL_MAX_LEN]
    return text

def to_number(value):
    """Próbuje zrzutować wartość na float; zwraca None, gdy się nie da."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        txt = value.strip().replace(',', '.')
        if not txt:
            return None
        try:
            return float(txt)
        except ValueError:
            return None
    return None

def clean_text(text):
    """Zastępuje znaki nowej linii i inne białe znaki pojedynczą spacją."""
    if not text:
        return ""
    return " ".join(text.split())

def wczytaj_aps_weight(sciezka_csv):
    """Wczytuje plik CSV z kolumnami sku,aps_weight i zwraca słownik sku -> aps_weight."""
    aps_weight_dict = {}
    try:
        with open(sciezka_csv, 'r', encoding='utf-8-sig') as plik_csv:
            reader = csv.DictReader(plik_csv)
            for row in reader:
                sku = row.get('sku', '').strip()
                aps_weight = row.get('aps_weight', '').strip()
                if sku and aps_weight:
                    try:
                        aps_weight_dict[sku] = float(aps_weight)
                    except ValueError:
                        pass  # Ignoruj błędne wartości
        return aps_weight_dict, None
    except Exception as e:
        return {}, str(e)

def pobierz_xml(url, sciezka_docelowa):
    """
    Pobiera plik XML z podanego URL.
    Zwraca (True, None) przy sukcesie lub (False, str(e)) przy błędzie.
    """
    try:
        opener = urllib.request.build_opener()
        opener.addheaders = [('User-agent', 'Mozilla/5.0')]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(url, sciezka_docelowa)
        return True, None
    except Exception as e:
        return False, str(e) # Zwraca błąd zamiast messagebox

def parsuj_xml(sciezka_pliku, aps_weight_dict=None):
    """
    Parsuje plik XML i ekstrahuje dane.
    Zwraca (atrybuty, maks_obr, dane, None) przy sukcesie 
    lub ([], 0, [], str(e)) przy błędzie.
    """
    try:
        drzewo = ET.parse(sciezka_pliku)
        root = drzewo.getroot()
        atrybuty = set()
        maks_liczba_obrazow = 0
        dane = []

        for element in root.findall("o"):
            cat_elem = element.find("cat")
            name_elem = element.find("name")
            desc_elem = element.find("desc")

            wiersz = {
                "id": element.get("id"),
                "url": element.get("url"),
                "price": element.get("price"),
                "avail": element.get("avail"),
                "weight": element.get("weight"),
                "stock": element.get("stock"),
                "cat": clean_text(cat_elem.text) if cat_elem is not None else "",
                "name": clean_text(name_elem.text) if name_elem is not None else "",
                "desc": clean_text(desc_elem.text) if desc_elem is not None else ""
            }
            
            # Aktualizuj weight z aps_weight jeśli dostępne
            if aps_weight_dict and wiersz["id"] in aps_weight_dict:
                wiersz["weight"] = str(aps_weight_dict[wiersz["id"]])
            
            atrybuty_elem = element.find("attrs")
            if atrybuty_elem is not None:
                for atrybut in atrybuty_elem.findall("a"):
                    nazwa_atrybutu = atrybut.get("name")
                    if nazwa_atrybutu:
                        atrybuty.add(nazwa_atrybutu)
                        wiersz[nazwa_atrybutu] = clean_text(atrybut.text)

            liczba_obrazow_w_wierszu = 0
            obrazy_elem = element.find("imgs")
            if obrazy_elem is not None:
                main_image = obrazy_elem.find("main")
                if main_image is not None and main_image.get("url"):
                    wiersz["image0"] = main_image.get("url")
                    liczba_obrazow_w_wierszu = 1
                
                start_index = 1 if "image0" in wiersz else 0

                for i, img in enumerate(obrazy_elem.findall("i"), start=start_index):
                    if img.get("url"):
                        wiersz[f"image{i}"] = img.get("url")
                        liczba_obrazow_w_wierszu = max(liczba_obrazow_w_wierszu, i + 1)
            
            maks_liczba_obrazow = max(maks_liczba_obrazow, liczba_obrazow_w_wierszu)
            dane.append(wiersz)
            
        # ZMODYFIKOWANO: Zwraca wynik zamiast pokazywać błąd
        return sorted(list(atrybuty)), maks_liczba_obrazow, dane, None
        
    except FileNotFoundError as e:
        return [], 0, [], f"Nie znaleziono pliku: {sciezka_pliku} ({e})"
    except ET.ParseError as e:
        return [], 0, [], f"Błąd parsowania XML w {os.path.basename(sciezka_pliku)}: {e}"
    except Exception as e:
        return [], 0, [], f"Nieoczekiwany błąd parsowania: {e}"

def zapisz_do_excel_jeden_arkusz(dane, atrybuty_lista, maks_liczba_obrazow, sciezka_pliku):
    """Zapisuje dane do jednego pliku Excel z jednym arkuszem."""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dane"
        
        pola_podstawowe = ['id', 'avail', 'weight', 'stock', 'cat', 'name', 'desc']
        pola_price = [f"price_{roz}" for roz in ROZSZERZENIA]
        excluded_attrs = {'producent', 'kod producenta', 'producer', 'producer code', 'producent_code', 'producer_code'}
        pola_atrybutow = [attr for attr in atrybuty_lista if attr.lower() not in excluded_attrs] 
        pola_obrazow = [f"image{i}" for i in range(min(maks_liczba_obrazow, 5))]
        pola = pola_podstawowe + pola_price + pola_atrybutow + pola_obrazow
        
        # Nagłówek
        for col_num, pole in enumerate(pola, 1):
            ws.cell(row=1, column=col_num, value=pole)
        
        # Dane
        for row_num, wiersz in enumerate(dane, 2):
            for col_num, pole in enumerate(pola, 1):
                raw_val = wiersz.get(pole, '')
                if pole == 'weight' or pole.startswith('price_'):
                    num_val = to_number(raw_val)
                    ws.cell(row=row_num, column=col_num, value=num_val if num_val is not None else sanitize_for_excel(raw_val))
                else:
                    ws.cell(row=row_num, column=col_num, value=sanitize_for_excel(raw_val))
        
        # Auto-dopasowanie kolumn
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 80 else 80
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(sciezka_pliku)
        return True, None
    except Exception as e:
        return False, str(e)

# --- FUNKCJA ZAPISU BŁĘDÓW (z poprzedniej prośby) ---
def zapisz_bledy_do_xlsx(bledne_linki, sciezka_zapisu):
    if not bledne_linki:
        return None
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Błędy Pobierania"
        ws.append(["Nieudany URL", "Powód błędu"])
        for url, powod in bledne_linki:
            ws.append([sanitize_for_excel(url), sanitize_for_excel(powod)])
        
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 80 else 80
            ws.column_dimensions[column_letter].width = adjusted_width

        teraz_format_czasu = datetime.now().strftime("%Y%m%d-%H%M%S")
        nazwa_pliku_xlsx = f"RAPORT_BLEDOW_POBIERANIA_{teraz_format_czasu}.xlsx"
        sciezka_pliku_xlsx = os.path.join(sciezka_zapisu, nazwa_pliku_xlsx)
        
        wb.save(sciezka_pliku_xlsx)
        return sciezka_pliku_xlsx
    except Exception as e:
        messagebox.showerror("Błąd zapisu raportu błędów", 
                             f"Nie udało się zapisać pliku XLSX z błędami: {e}")
        return None
# ----------------------------------------

# --- NOWA FUNKCJA ROBOCZA (WORKER) ---
def pobierz_i_parsuj_url(url, aps_weight_dict=None):
    """
    Pobiera i parsuje jeden URL. Przeznaczona do uruchamiania w osobnym wątku.
    Zwraca status i dane.
    """
    katalog_tymczasowy = tempfile.gettempdir()
    nazwa_pliku_url = os.path.basename(urlparse(url).path) or f"feed_{hash(url)}.xml"
    nazwa_bazowa_xml = os.path.splitext(nazwa_pliku_url)[0]
    teraz_timestamp_temp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    sciezka_lokalna_xml = os.path.join(katalog_tymczasowy, f"temp_{nazwa_bazowa_xml}_{teraz_timestamp_temp}.xml")

    # Wyznaczenie rozszerzenia
    if '_mano' in nazwa_bazowa_xml:
        # Przywracamy pełny sufiks (np. manoFR, manoFR_extra), aby klucze cen odpowiadały ROZSZERZENIA
        rozszerzenie_part = nazwa_bazowa_xml.split('_mano', 1)[1]
        rozszerzenie = f"mano{rozszerzenie_part}" if rozszerzenie_part else 'UNKNOWN'
    else:
        rozszerzenie = 'UNKNOWN'

    # 1. Pobieranie
    sukces_pobierania, powod_bledu_pob = pobierz_xml(url, sciezka_lokalna_xml)
    if not sukces_pobierania:
        return ('blad_pobierania', (url, powod_bledu_pob, nazwa_pliku_url))

    # 2. Parsowanie
    atrybuty, maks_obr, dane, powod_bledu_pars = parsuj_xml(sciezka_lokalna_xml, aps_weight_dict)
    
    # 3. Czyszczenie pliku tymczasowego
    try:
        os.remove(sciezka_lokalna_xml)
    except Exception as e:
        print(f"Ostrzeżenie: Nie udało się usunąć pliku tymczasowego {sciezka_lokalna_xml}: {e}")

    # 4. Zwracanie wyników
    if powod_bledu_pars:
        return ('blad_parsowania', (url, powod_bledu_pars, nazwa_pliku_url))
    
    if not dane:
         return ('blad_parsowania', (url, "Brak elementów <o> po parsowaniu", nazwa_pliku_url))

    return ('sukces', (dane, atrybuty, maks_obr, nazwa_bazowa_xml, nazwa_pliku_url, rozszerzenie))
# ----------------------------------------


# --- PRZEBUDOWANA FUNKCJA PRZETWARZANIA ---
def przetworz_na_excel_po_rozszerzeniach(nazwy, sciezka_zapisu, sciezka_csv, app_instance):
    """Przetwarza URL-e dla podanych nazw i zapisuje do jednego pliku Excel z kolumnami price_rozszerzenie."""
    # Wczytaj aps_weight
    aps_weight_dict = {}
    if sciezka_csv:
        aps_weight_dict, blad_csv = wczytaj_aps_weight(sciezka_csv)
        if blad_csv:
            messagebox.showwarning("Błąd wczytywania CSV", f"Nie udało się wczytać pliku CSV: {blad_csv}")
    
    urls = []
    for nazwa in nazwy:
        for roz in ROZSZERZENIA:
            urls.append(f"https://sm-prods.com/feeds/{nazwa}_{roz}.xml")
    
    dane_po_id = {}  # id -> {dane podstawowe, prices: {roz: price}}
    
    liczba_url = len(urls)
    sukcesy_przetwarzania = 0
    bledy_pobierania = 0
    bledy_parsowania = 0
    bledy_zapisu = 0
    bledne_linki = []

    if liczba_url == 0:
        app_instance.update_status("Nie podano żadnych URL-i.", 0)
        return

    if not os.path.exists(sciezka_zapisu):
        try:
            os.makedirs(sciezka_zapisu)
        except Exception as e:
            messagebox.showerror("Błąd ścieżki zapisu", f"Nie można utworzyć katalogu: {sciezka_zapisu}\nBłąd: {e}\nPliki będą zapisywane w katalogu roboczym.")
            sciezka_zapisu = os.getcwd()
            app_instance.pole_sciezki_zapisu.delete(0, ctk.END)
            app_instance.pole_sciezki_zapisu.insert(0, sciezka_zapisu)

    app_instance.update_status(f"Rozpoczynam przetwarzanie {liczba_url} linków (max {MAX_WORKERS} wątków)...", 0)

    # Użycie ThreadPoolExecutor do jednoczesnego pobierania i parsowania
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Tworzenie listy "przyszłych" zadań
        futures = [executor.submit(pobierz_i_parsuj_url, url, aps_weight_dict) for url in urls]

        # Przetwarzanie wyników w kolejności ich kończenia się
        for i, future in enumerate(concurrent.futures.as_completed(futures)):
            postep = (i + 1) / liczba_url
            
            try:
                status, data = future.result()
                
                if status == 'sukces':
                    dane, atrybuty, maks_obr, nazwa_bazowa_xml, nazwa_pliku_url, rozszerzenie = data
                    for wiersz in dane:
                        prod_id = wiersz['id']
                        if prod_id not in dane_po_id:
                            dane_po_id[prod_id] = {
                                'podstawowe': {k: v for k, v in wiersz.items() if k != 'price'},
                                'prices': {},
                                'atrybuty': set(atrybuty),
                                'maks_obr': maks_obr
                            }
                        else:
                            # Uzupełnij brakujące obrazki/atrybuty z kolejnych feedów tego samego produktu
                            podstawowe = dane_po_id[prod_id]['podstawowe']
                            for klucz, wartosc in wiersz.items():
                                if klucz == 'price':
                                    continue
                                if klucz.startswith('image') and wartosc and not podstawowe.get(klucz):
                                    podstawowe[klucz] = wartosc
                                # Atrybuty: jeśli nowy atrybut lub brakująca wartość, uzupełnij
                                if klucz not in podstawowe and wartosc:
                                    podstawowe[klucz] = wartosc
                        dane_po_id[prod_id]['prices'][rozszerzenie] = wiersz['price']
                        dane_po_id[prod_id]['atrybuty'].update(atrybuty)
                        dane_po_id[prod_id]['maks_obr'] = max(dane_po_id[prod_id]['maks_obr'], maks_obr)
                    sukcesy_przetwarzania += 1
                    app_instance.update_status(f"Pobrano {i+1}/{liczba_url}: {nazwa_pliku_url}", postep)

                elif status == 'blad_pobierania':
                    url, powod, nazwa_pliku_url = data
                    bledy_pobierania += 1
                    bledne_linki.append((url, powod))
                    app_instance.update_status(f"Błąd pobierania {i+1}/{liczba_url}: {nazwa_pliku_url}", postep)
                    time.sleep(0.1) # Krótka pauza na odczytanie statusu

                elif status == 'blad_parsowania':
                    url, powod, nazwa_pliku_url = data
                    bledy_parsowania += 1
                    # Błędów parsowania nie dodajemy do raportu XLSX, ale liczymy je
                    app_instance.update_status(f"Błąd parsowania {i+1}/{liczba_url}: {nazwa_pliku_url}", postep)
                    time.sleep(0.1)

            except Exception as e:
                # To jest nieoczekiwany błąd w samej logice wątku
                bledy_parsowania += 1 # Traktujemy to jako błąd przetwarzania
                app_instance.update_status(f"Błąd krytyczny wątku {i+1}/{liczba_url}: {e}", postep)
                print(f"Błąd krytyczny w wątku: {e}")


    # --- Zapisywanie wyników ---
    
    sciezka_pliku_bledow = zapisz_bledy_do_xlsx(bledne_linki, sciezka_zapisu)
    liczba_bledow_ogolem = bledy_pobierania + bledy_parsowania + bledy_zapisu
    
    podsumowanie_bledow_tekst = (
        f"Błędy pobierania: {bledy_pobierania}\n"
        f"Błędy parsowania: {bledy_parsowania}\n"
        f"Błędy zapisu: {bledy_zapisu}"
    )

    if not dane_po_id:
        app_instance.update_status(f"Nie udało się przetworzyć żadnych danych. Błędy: {liczba_bledow_ogolem}", 0)
        wiadomosc_ostrzezenia = (
            f"Nie udało się pobrać ani sparsować danych z żadnego podanego URL.\n\n"
            f"{podsumowanie_bledow_tekst}"
        )
        if sciezka_pliku_bledow:
            wiadomosc_ostrzezenia += f"\n\nZapisano raport błędów pobierania:\n{os.path.basename(sciezka_pliku_bledow)}"
        messagebox.showwarning("Brak danych", wiadomosc_ostrzezenia)
        app_instance.reset_gui_after_delay()
        return

    # Przygotowanie danych do zapisu
    all_atrybuty = set()
    global_maks_obr = 0
    for info in dane_po_id.values():
        all_atrybuty.update(info['atrybuty'])
        global_maks_obr = max(global_maks_obr, info['maks_obr'])
    
    dane_do_zapisu = []
    for prod_id, info in dane_po_id.items():
        wiersz = info['podstawowe'].copy()
        for roz in ROZSZERZENIA:
            wiersz[f'price_{roz}'] = info['prices'].get(roz, '')
        dane_do_zapisu.append(wiersz)
    
    # Tworzenie nazwy pliku
    teraz_format_czasu = datetime.now().strftime("%d%m%y-%H%M%S")
    if len(nazwy) == 1:
        nazwa_pliku = f"{nazwy[0]}_merged_{teraz_format_czasu}.xlsx"
    else:
        nazwa_pliku = f"{nazwy[0]}_and_{len(nazwy)-1}_more_merged_{teraz_format_czasu}.xlsx"
    nazwa_pliku_excel = os.path.join(sciezka_zapisu, nazwa_pliku)
    
    app_instance.update_status("Zapisywanie połączonych danych do Excel...", 0.98)
    
    sukces_zapisu, powod_bledu_zapisu = zapisz_do_excel_jeden_arkusz(dane_do_zapisu, sorted(list(all_atrybuty)), global_maks_obr, nazwa_pliku_excel)
    
    if not sukces_zapisu:
        bledy_zapisu += 1

    app_instance.update_status(f"Zakończono. Przetworzono: {sukcesy_przetwarzania}, Błędy: {liczba_bledow_ogolem}", 1)

    # --- Końcowe podsumowanie ---
    if dane_do_zapisu and sukces_zapisu:
        wiadomosc_sukcesu = (
            f"Pomyślnie przetworzono {sukcesy_przetwarzania} z {liczba_url} plików.\n"
            f"Zapisano połączone dane do:\n{os.path.abspath(nazwa_pliku_excel)}\n\n"
            f"{podsumowanie_bledow_tekst}"
        )
        if sciezka_pliku_bledow:
            wiadomosc_sukcesu += f"\n\nZapisano raport błędów pobierania:\n{os.path.basename(sciezka_pliku_bledow)}"
        messagebox.showinfo("Sukces", wiadomosc_sukcesu)
    else:
        wiadomosc_ostrzezenia = (
            f"Zakończono przetwarzanie, ale wystąpiły błędy.\n"
            f"Pomyślnie przetworzono: {sukcesy_przetwarzania} z {liczba_url}\n\n"
            f"{podsumowanie_bledow_tekst}"
        )
        if sciezka_pliku_bledow:
            wiadomosc_ostrzezenia += f"\n\nZapisano raport błędów pobierania:\n{os.path.basename(sciezka_pliku_bledow)}"
        messagebox.showwarning("Zakończono z błędami", wiadomosc_ostrzezenia)

    app_instance.reset_gui_after_delay()


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Konwerter XML do Excel")
        self.configure(fg_color=KOLOR_TLA_GLOWNY)
        
        window_width = 600
        window_height = 550
        self.minsize(500, 450)

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        center_x = int(screen_width/2 - window_width / 2)
        center_y = int(screen_height/2 - window_height / 2)
        self.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)  
        self.grid_rowconfigure(1, weight=0)  
        self.grid_rowconfigure(2, weight=0)  
        self.grid_rowconfigure(3, weight=0)  
        self.grid_rowconfigure(4, weight=0)  
        self.grid_rowconfigure(5, weight=0)  

        input_frame_urls = ctk.CTkFrame(self, fg_color=KOLOR_TLA_SEKCJI, border_width=1, border_color=KOLOR_AKCENT)
        input_frame_urls.grid(row=0, column=0, padx=20, pady=(20,10), sticky="nsew")
        input_frame_urls.grid_columnconfigure(0, weight=1)
        input_frame_urls.grid_rowconfigure(1, weight=1) 

        etykieta_nazwa = ctk.CTkLabel(input_frame_urls, text="Wprowadź nazwy dla generowania linków XML (każda w nowej linii):", text_color=KOLOR_TEKSTU)
        etykieta_nazwa.grid(row=0, column=0, padx=10, pady=(10, 5), sticky="w")
        self.pole_nazwa = ctk.CTkTextbox(input_frame_urls, height=100, fg_color="white", border_width=1, border_color=KOLOR_AKCENT, text_color=KOLOR_TEKSTU)
        self.pole_nazwa.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")

        input_frame_path = ctk.CTkFrame(self, fg_color=KOLOR_TLA_SEKCJI, border_width=1, border_color=KOLOR_AKCENT)
        input_frame_path.grid(row=1, column=0, padx=20, pady=(5,10), sticky="ew")
        input_frame_path.grid_columnconfigure(1, weight=1)

        etykieta_sciezki = ctk.CTkLabel(input_frame_path, text="Katalog zapisu pliku Excel:", text_color=KOLOR_TEKSTU)
        etykieta_sciezki.grid(row=0, column=0, padx=(10,5), pady=5, sticky="w")
        
        self.pole_sciezki_zapisu = ctk.CTkEntry(input_frame_path, fg_color="white", border_color=KOLOR_AKCENT, text_color=KOLOR_TEKSTU)
        self.pole_sciezki_zapisu.grid(row=0, column=1, padx=(0,5), pady=5, sticky="ew")
        self.pole_sciezki_zapisu.insert(0, DOMYSLNA_SCIEZKA_ZAPISU)

        przycisk_wybierz_sciezke = ctk.CTkButton(input_frame_path, text="Wybierz folder", width=120, command=self.wybierz_katalog_zapisu, fg_color=KOLOR_AKCENT, hover_color=KOLOR_AKCENT_HOVER, text_color="white")
        przycisk_wybierz_sciezke.grid(row=0, column=2, padx=(0,10), pady=5, sticky="e")

        input_frame_csv = ctk.CTkFrame(self, fg_color=KOLOR_TLA_SEKCJI, border_width=1, border_color=KOLOR_AKCENT)
        input_frame_csv.grid(row=2, column=0, padx=20, pady=(5,10), sticky="ew")
        input_frame_csv.grid_columnconfigure(1, weight=1)

        etykieta_csv = ctk.CTkLabel(input_frame_csv, text="Plik CSV z aps_weight:", text_color=KOLOR_TEKSTU)
        etykieta_csv.grid(row=0, column=0, padx=(10,5), pady=5, sticky="w")
        
        self.pole_csv = ctk.CTkEntry(input_frame_csv, fg_color="white", border_color=KOLOR_AKCENT, text_color=KOLOR_TEKSTU)
        self.pole_csv.grid(row=0, column=1, padx=(0,5), pady=5, sticky="ew")

        przycisk_wybierz_csv = ctk.CTkButton(input_frame_csv, text="Wybierz plik", width=120, command=self.wybierz_plik_csv, fg_color=KOLOR_AKCENT, hover_color=KOLOR_AKCENT_HOVER, text_color="white")
        przycisk_wybierz_csv.grid(row=0, column=2, padx=(0,10), pady=5, sticky="e")

        self.przycisk_przetworz = ctk.CTkButton(self, text="Przetwórz na Excel", command=self.rozpocznij_przetwarzanie_action, height=40, fg_color=KOLOR_AKCENT, hover_color=KOLOR_AKCENT_HOVER, text_color="white")
        self.przycisk_przetworz.grid(row=3, column=0, padx=20, pady=10, sticky="ew")

        self.progress_bar = ctk.CTkProgressBar(self, height=10, progress_color=KOLOR_AKCENT, fg_color="#F7D4E8")
        self.progress_bar.grid(row=4, column=0, padx=20, pady=(0, 5), sticky="ew")
        self.progress_bar.set(0)

        self.status_label = ctk.CTkLabel(self, text="Gotowy.", text_color=KOLOR_STATUSU)
        self.status_label.grid(row=5, column=0, padx=20, pady=(0,10), sticky="ew")

    def wybierz_katalog_zapisu(self):
        sciezka_katalogu = filedialog.askdirectory(initialdir=self.pole_sciezki_zapisu.get() or DOMYSLNA_SCIEZKA_ZAPISU)
        if sciezka_katalogu:
            self.pole_sciezki_zapisu.delete(0, ctk.END)
            self.pole_sciezki_zapisu.insert(0, sciezka_katalogu)

    def wybierz_plik_csv(self):
        sciezka_pliku = filedialog.askopenfilename(initialdir=os.getcwd(), title="Wybierz plik CSV z aps_weight", filetypes=[("Pliki CSV", "*.csv")])
        if sciezka_pliku:
            self.pole_csv.delete(0, ctk.END)
            self.pole_csv.insert(0, sciezka_pliku)

    def update_status(self, message, progress_value=None):
        """Aktualizuje etykietę statusu i pasek postępu."""
        self.status_label.configure(text=message)
        if progress_value is not None:
            self.progress_bar.set(progress_value)
        self.update_idletasks() # Wymusza odświeżenie GUI

    def reset_gui_after_delay(self, delay_ms=4000):
        """Resetuje pasek postępu i status po opóźnieniu."""
        self.after(delay_ms, lambda: self.update_status("Gotowy.", 0))
        self.przycisk_przetworz.configure(state="normal", text="Przetwórz na JEDEN plik CSV")

    def rozpocznij_przetwarzanie_action(self):
        """Rozpoczyna proces przetwarzania dla podanych nazw."""
        nazwy_text = self.pole_nazwa.get("1.0", ctk.END) 
        nazwy = [nazwa.strip() for nazwa in nazwy_text.splitlines() if nazwa.strip()]

        sciezka_zapisu_gui = self.pole_sciezki_zapisu.get().strip()
        if not sciezka_zapisu_gui:
            sciezka_zapisu_gui = DOMYSLNA_SCIEZKA_ZAPISU

        sciezka_csv_gui = self.pole_csv.get().strip()

        if not nazwy:
            messagebox.showerror("Błąd", "Musisz wprowadzić co najmniej jedną nazwę.")
            return

        self.przycisk_przetworz.configure(state="disabled", text="Przetwarzanie...")
        self.update_idletasks()
        
        # Uruchomienie funkcji przetwarzającej
        przetworz_na_excel_po_rozszerzeniach(nazwy, sciezka_zapisu_gui, sciezka_csv_gui, self)

if __name__ == "__main__":
    app = App()
    app.mainloop()