import concurrent.futures
import csv
import ctypes
import os
import sys
import tempfile
import urllib.request
import xml.etree.ElementTree as ET
from ctypes import wintypes
from datetime import datetime
from urllib.parse import urlparse

import openpyxl
from openpyxl.utils import get_column_letter

try:
    from PySide6.QtCore import QThread, Qt, Signal
    from PySide6.QtGui import QColor, QFont, QLinearGradient, QPainter
    from PySide6.QtWidgets import (
        QApplication,
        QFileDialog,
        QFrame,
        QGridLayout,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPlainTextEdit,
        QProgressBar,
        QPushButton,
        QScrollArea,
        QSizePolicy,
        QVBoxLayout,
        QWidget,
    )
except ImportError as error:
    raise ImportError(
        "Brak PySide6. Zainstaluj: pip install PySide6 openpyxl"
    ) from error


MAX_WORKERS = 10
DEFAULT_OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
BASE_FIELDS = ["id", "url", "price", "avail", "weight", "stock", "cat", "name", "desc"]
CSV_DELIMITER = "|"


def clean_text(text):
    """Zastępuje znaki nowej linii i inne białe znaki pojedynczą spacją."""
    if not text:
        return ""
    return " ".join(text.split())


def download_xml(url, target_path):
    """Pobiera plik XML z podanego URL. Zwraca (True, None) lub (False, powód)."""
    try:
        opener = urllib.request.build_opener()
        opener.addheaders = [("User-agent", "Mozilla/5.0")]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(url, target_path)
        return True, None
    except Exception as error:
        return False, str(error)


def parse_xml(file_path):
    """
    Parsuje plik XML i ekstrahuje dane produktowe.
    Zwraca (atrybuty, maks_obrazow, dane, None) lub ([], 0, [], powod_bledu).
    """
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        attributes = set()
        max_images = 0
        rows = []

        for element in root.findall("o"):
            cat_elem = element.find("cat")
            name_elem = element.find("name")
            desc_elem = element.find("desc")

            row = {
                "id": element.get("id"),
                "url": element.get("url"),
                "price": element.get("price"),
                "avail": element.get("avail"),
                "weight": element.get("weight"),
                "stock": element.get("stock"),
                "cat": clean_text(cat_elem.text) if cat_elem is not None else "",
                "name": clean_text(name_elem.text) if name_elem is not None else "",
                "desc": clean_text(desc_elem.text) if desc_elem is not None else "",
            }

            attrs_elem = element.find("attrs")
            if attrs_elem is not None:
                for attr in attrs_elem.findall("a"):
                    attr_name = attr.get("name")
                    if attr_name:
                        attributes.add(attr_name)
                        row[attr_name] = clean_text(attr.text)

            images_in_row = 0
            imgs_elem = element.find("imgs")
            if imgs_elem is not None:
                main_image = imgs_elem.find("main")
                if main_image is not None and main_image.get("url"):
                    row["image0"] = main_image.get("url")
                    images_in_row = 1

                start_index = 1 if "image0" in row else 0
                for i, img in enumerate(imgs_elem.findall("i"), start=start_index):
                    if img.get("url"):
                        row[f"image{i}"] = img.get("url")
                        images_in_row = max(images_in_row, i + 1)

            max_images = max(max_images, images_in_row)
            rows.append(row)

        return sorted(attributes), max_images, rows, None

    except FileNotFoundError as error:
        return [], 0, [], f"Nie znaleziono pliku: {file_path} ({error})"
    except ET.ParseError as error:
        return [], 0, [], f"Błąd parsowania XML w {os.path.basename(file_path)}: {error}"
    except Exception as error:
        return [], 0, [], f"Nieoczekiwany błąd parsowania: {error}"


def write_csv(rows, attributes, max_images, file_path):
    """Zapisuje połączone dane do jednego pliku CSV."""
    fields = BASE_FIELDS + list(attributes) + [f"image{i}" for i in range(max_images)]
    try:
        with open(file_path, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=fields, delimiter=CSV_DELIMITER, extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(rows)
        return True, None
    except Exception as error:
        return False, str(error)


def save_error_report(download_errors, parse_errors, output_dir):
    """Zapisuje raport błędów pobierania/parsowania do pliku XLSX. Zwraca ścieżkę lub None."""
    if not download_errors and not parse_errors:
        return None

    workbook = openpyxl.Workbook()
    sheets = []

    if download_errors:
        sheet = workbook.active
        sheet.title = "Bledy pobierania"
        sheet.append(["Nieudany URL", "Powód błędu", "Plik"])
        for url, reason, file_name in download_errors:
            sheet.append([url, reason, file_name])
        sheets.append(sheet)
    else:
        workbook.remove(workbook.active)

    if parse_errors:
        sheet = workbook.create_sheet("Bledy parsowania")
        sheet.append(["URL", "Powód błędu", "Plik"])
        for url, reason, file_name in parse_errors:
            sheet.append([url, reason, file_name])
        sheets.append(sheet)

    for sheet in sheets:
        for column in sheet.columns:
            max_length = 0
            letter = get_column_letter(column[0].column)
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[letter].width = min(max_length + 2, 80)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    report_path = os.path.join(output_dir, f"RAPORT_BLEDOW_XMLCSV_{timestamp}.xlsx")
    workbook.save(report_path)
    return report_path


def download_and_parse_url(url):
    """Pobiera i parsuje jeden URL. Przeznaczone do uruchamiania w osobnym wątku."""
    temp_dir = tempfile.gettempdir()
    file_name = os.path.basename(urlparse(url).path) or f"feed_{abs(hash(url))}.xml"
    base_name = os.path.splitext(file_name)[0]
    temp_stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    local_xml_path = os.path.join(temp_dir, f"temp_{base_name}_{temp_stamp}.xml")

    success, error_message = download_xml(url, local_xml_path)
    if not success:
        return "download_error", (url, error_message, file_name)

    attributes, max_images, rows, parse_error = parse_xml(local_xml_path)

    try:
        os.remove(local_xml_path)
    except OSError:
        pass

    if parse_error:
        return "parse_error", (url, parse_error, file_name)

    if not rows:
        return "parse_error", (url, "Brak elementów <o> po parsowaniu", file_name)

    return "success", (rows, attributes, max_images, base_name, file_name)


def enable_windows_acrylic(win_id):
    if os.name != "nt":
        return False

    class ACCENT_POLICY(ctypes.Structure):
        _fields_ = [
            ("AccentState", ctypes.c_int),
            ("AccentFlags", ctypes.c_int),
            ("GradientColor", ctypes.c_uint32),
            ("AnimationId", ctypes.c_int),
        ]

    class WINDOWCOMPOSITIONATTRIBDATA(ctypes.Structure):
        _fields_ = [
            ("Attribute", ctypes.c_int),
            ("Data", ctypes.c_void_p),
            ("SizeOfData", ctypes.c_size_t),
        ]

    ACCENT_ENABLE_ACRYLICBLURBEHIND = 4
    WCA_ACCENT_POLICY = 19

    try:
        user32 = ctypes.windll.user32
        set_window_composition_attribute = user32.SetWindowCompositionAttribute
    except Exception:
        return False

    hwnd = wintypes.HWND(int(win_id))
    accent = ACCENT_POLICY(
        AccentState=ACCENT_ENABLE_ACRYLICBLURBEHIND,
        AccentFlags=2,
        GradientColor=0xEEF5E8FF,
        AnimationId=0,
    )
    data = WINDOWCOMPOSITIONATTRIBDATA(
        Attribute=WCA_ACCENT_POLICY,
        Data=ctypes.cast(ctypes.pointer(accent), ctypes.c_void_p),
        SizeOfData=ctypes.sizeof(accent),
    )

    try:
        return bool(set_window_composition_attribute(hwnd, ctypes.byref(data)))
    except Exception:
        return False


class ProcessorThread(QThread):
    progress_signal = Signal(str, float, str)
    done_signal = Signal(dict)
    error_signal = Signal(str)

    def __init__(self, urls, output_dir):
        super().__init__()
        self.urls = urls
        self.output_dir = output_dir

    def _emit_progress(self, message, value, tone="normal"):
        self.progress_signal.emit(message, value, tone)

    def run(self):
        try:
            if not self.urls:
                self.error_signal.emit("Podaj co najmniej jeden URL pliku XML.")
                return

            if not os.path.exists(self.output_dir):
                try:
                    os.makedirs(self.output_dir, exist_ok=True)
                except Exception as error:
                    self.error_signal.emit(
                        f"Nie można utworzyć katalogu zapisu:\n{self.output_dir}\n{error}"
                    )
                    return

            all_rows = []
            all_attributes = set()
            global_max_images = 0
            base_names = []
            download_errors = []
            parse_errors = []
            total = len(self.urls)

            self._emit_progress(
                f"Rozpoczynam przetwarzanie {total} linków (max {MAX_WORKERS} wątków)...", 0.02
            )

            with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                futures = [executor.submit(download_and_parse_url, url) for url in self.urls]
                for index, future in enumerate(concurrent.futures.as_completed(futures), start=1):
                    progress = 0.02 + (0.93 * index / total)
                    try:
                        status, data = future.result()
                        if status == "success":
                            rows, attributes, max_images, base_name, file_name = data
                            all_rows.extend(rows)
                            all_attributes.update(attributes)
                            global_max_images = max(global_max_images, max_images)
                            base_names.append(base_name)
                            self._emit_progress(f"Pobrano {index}/{total}: {file_name}", progress)
                        elif status == "download_error":
                            download_errors.append(data)
                            self._emit_progress(
                                f"Błąd pobierania {index}/{total}: {data[2]}", progress, "warn"
                            )
                        else:
                            parse_errors.append(data)
                            self._emit_progress(
                                f"Błąd parsowania {index}/{total}: {data[2]}", progress, "warn"
                            )
                    except Exception as error:
                        parse_errors.append(("?", f"Błąd krytyczny wątku: {error}", "?"))
                        self._emit_progress(f"Błąd krytyczny wątku {index}/{total}", progress, "warn")

            error_count = len(download_errors) + len(parse_errors)

            if not all_rows:
                report_path = save_error_report(download_errors, parse_errors, self.output_dir)
                message = (
                    "Nie udało się pobrać ani sparsować danych z żadnego podanego URL.\n\n"
                    f"Błędy pobierania: {len(download_errors)}\n"
                    f"Błędy parsowania: {len(parse_errors)}"
                )
                if report_path:
                    message += f"\n\nRaport błędów:\n{os.path.basename(report_path)}"
                self.done_signal.emit(
                    {
                        "ok": False,
                        "title": "Brak danych",
                        "message": message,
                        "tone": "warn",
                        "progress": 0.0,
                    }
                )
                return

            combined_name = "_".join(base_names)
            if len(combined_name) > 100:
                combined_name = f"{base_names[0]}_and_{len(base_names) - 1}_more"

            timestamp = datetime.now().strftime("%d%m%y-%H%M%S")
            csv_path = os.path.join(self.output_dir, f"{combined_name}_{timestamp}.csv")

            self._emit_progress("Zapisywanie połączonych danych...", 0.98)
            save_ok, save_error = write_csv(
                all_rows, sorted(all_attributes), global_max_images, csv_path
            )

            report_path = save_error_report(download_errors, parse_errors, self.output_dir)

            summary_lines = [
                f"Przetworzone pliki XML: {len(base_names)}/{total}",
                f"Wiersze produktów: {len(all_rows)}",
                f"Kolumny atrybutów: {len(all_attributes)}",
                f"Maks. liczba obrazów: {global_max_images}",
                f"Błędy pobierania: {len(download_errors)}",
                f"Błędy parsowania: {len(parse_errors)}",
            ]

            if save_ok:
                summary_lines.append(f"\nZapisano CSV:\n{os.path.abspath(csv_path)}")
            else:
                summary_lines.append(f"\nBłąd zapisu CSV: {save_error}")

            if report_path:
                summary_lines.append(f"Raport błędów: {os.path.basename(report_path)}")

            ok = save_ok and error_count == 0
            self.done_signal.emit(
                {
                    "ok": ok,
                    "title": "Sukces" if ok else "Zakończono z błędami",
                    "message": "\n".join(summary_lines),
                    "tone": "ok" if ok else "warn",
                    "progress": 1.0,
                }
            )
        except Exception as error:
            self.error_signal.emit(str(error))


class GradientBackgroundWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.setObjectName("Root")

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        gradient = QLinearGradient(0, 0, self.width(), self.height())
        gradient.setColorAt(0.0, QColor("#e8daf5"))
        gradient.setColorAt(0.5, QColor("#e0d0f0"))
        gradient.setColorAt(1.0, QColor("#d8c8eb"))
        painter.fillRect(self.rect(), gradient)

        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor("#f5d5e8"))
        painter.drawEllipse(-150, -100, 450, 350)

        painter.setBrush(QColor("#dfc8e8"))
        painter.drawEllipse(self.width() - 300, -80, 400, 300)

        painter.setBrush(QColor("#f8c2df"))
        painter.drawEllipse(self.width() - 240, self.height() - 190, 380, 260)


class GlassCard(QFrame):
    def __init__(self, object_name="GlassCard"):
        super().__init__()
        self.setObjectName(object_name)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Konwerter XML → CSV")
        self.resize(760, 640)
        self.setMinimumSize(620, 560)

        self.worker = None

        self.root = GradientBackgroundWidget()
        self.setCentralWidget(self.root)

        self._build_ui()
        self._apply_styles()

    def showEvent(self, event):
        super().showEvent(event)
        enable_windows_acrylic(self.winId())

    def _build_ui(self):
        root_layout = QVBoxLayout(self.root)
        root_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("QScrollArea { background: transparent; }")

        scroll_host = QWidget()
        scroll_host.setStyleSheet("background: transparent;")
        outer_layout = QVBoxLayout(scroll_host)
        outer_layout.setContentsMargins(22, 22, 22, 22)
        outer_layout.setSpacing(14)

        header = GlassCard("HeaderCard")
        header.setMinimumHeight(120)
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(28, 20, 28, 20)
        header_layout.setSpacing(6)

        title = QLabel("Konwerter XML → CSV")
        title.setObjectName("Title")
        title.setWordWrap(True)

        subtitle = QLabel(
            "Pobierz wiele feedów XML jednocześnie i połącz je w jeden plik CSV."
        )
        subtitle.setObjectName("Subtitle")
        subtitle.setWordWrap(True)

        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        outer_layout.addWidget(header)

        form_card = GlassCard()
        form_layout = QVBoxLayout(form_card)
        form_layout.setContentsMargins(18, 16, 18, 16)
        form_layout.setSpacing(12)

        form_layout.addWidget(
            self._section_label("Linki XML", "Wklej URL-e plików XML, każdy w nowej linii.")
        )
        self.url_input = QPlainTextEdit()
        self.url_input.setPlaceholderText(
            "https://example.com/feed.xml\nhttps://example.com/feed2.xml"
        )
        self.url_input.setMinimumHeight(180)
        self.url_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        form_layout.addWidget(self.url_input, 1)

        outer_layout.addWidget(form_card, 1)

        settings_card = GlassCard()
        settings_layout = QGridLayout(settings_card)
        settings_layout.setContentsMargins(18, 16, 18, 16)
        settings_layout.setHorizontalSpacing(10)
        settings_layout.setVerticalSpacing(10)

        output_label = QLabel("Folder zapisu CSV:")
        output_label.setObjectName("FieldLabel")
        self.output_input = QLineEdit(DEFAULT_OUTPUT_DIR)
        self.output_btn = QPushButton("Wybierz folder")
        self.output_btn.clicked.connect(self.pick_output_dir)

        settings_layout.addWidget(output_label, 0, 0)
        settings_layout.addWidget(self.output_input, 0, 1)
        settings_layout.addWidget(self.output_btn, 0, 2)
        settings_layout.setColumnStretch(1, 1)
        outer_layout.addWidget(settings_card)

        self.run_btn = QPushButton("Przetwórz na JEDEN plik CSV")
        self.run_btn.setObjectName("RunButton")
        self.run_btn.clicked.connect(self.run_processing)
        outer_layout.addWidget(self.run_btn)

        self.progress = QProgressBar()
        self.progress.setRange(0, 1000)
        self.progress.setValue(0)
        outer_layout.addWidget(self.progress)

        self.status = QLabel("Gotowy.")
        self.status.setObjectName("Status")
        self.status.setWordWrap(True)
        outer_layout.addWidget(self.status)

        scroll.setWidget(scroll_host)
        root_layout.addWidget(scroll)

    def _section_label(self, title, subtitle):
        wrapper = QWidget()
        layout = QVBoxLayout(wrapper)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        t = QLabel(title)
        t.setObjectName("SectionTitle")
        s = QLabel(subtitle)
        s.setObjectName("SectionSub")
        layout.addWidget(t)
        layout.addWidget(s)
        return wrapper

    def _apply_styles(self):
        self.setStyleSheet(
            """
            QWidget#Root {
                background: transparent;
                color: #4c1636;
                font-family: 'Segoe UI', 'Tahoma', sans-serif;
                font-size: 13px;
            }
            QFrame#HeaderCard {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 150, 200, 100),
                    stop:1 rgba(180, 120, 200, 110));
                border: 1px solid rgba(255,255,255,100);
                border-radius: 22px;
            }
            QFrame#GlassCard {
                background: rgba(255,255,255,50);
                border: 1px solid rgba(255,255,255,80);
                border-radius: 20px;
            }
            QLabel#Title {
                color: #4c1636;
                font-size: 26px;
                font-weight: 800;
                margin-top: 2px;
                line-height: 1.2;
            }
            QLabel#Subtitle {
                color: #6a2a52;
                font-size: 13px;
                line-height: 1.25;
            }
            QLabel#SectionTitle {
                color: #4c1636;
                font-size: 16px;
                font-weight: 700;
            }
            QLabel#SectionSub {
                color: #7d4165;
                font-size: 12px;
            }
            QLabel#FieldLabel {
                color: #5b2142;
                font-weight: 700;
            }
            QPlainTextEdit, QLineEdit {
                background: rgba(255,255,255,70);
                border: 1px solid rgba(255,255,255,120);
                border-radius: 12px;
                padding: 8px;
                color: #4c1636;
                selection-background-color: rgba(255,150,200,150);
            }
            QPlainTextEdit:focus, QLineEdit:focus {
                border: 1px solid rgba(255,200,230,180);
                background: rgba(255,255,255,90);
            }
            QPushButton {
                background: rgba(255,150,200,120);
                border: 1px solid rgba(255,255,255,150);
                border-radius: 12px;
                color: #4c1636;
                padding: 9px 14px;
                font-weight: 700;
            }
            QPushButton:hover {
                background: rgba(255,150,200,160);
            }
            QPushButton:disabled {
                background: rgba(200,150,180,80);
                color: rgba(76,22,54,120);
            }
            QPushButton#RunButton {
                min-height: 44px;
                border-radius: 14px;
                font-size: 15px;
                background: rgba(255,120,180,140);
            }
            QPushButton#RunButton:hover {
                background: rgba(255,120,180,180);
            }
            QProgressBar {
                background: rgba(255,255,255,60);
                border-radius: 7px;
                border: 1px solid rgba(255,255,255,100);
                min-height: 12px;
                text-align: center;
                color: transparent;
            }
            QProgressBar::chunk {
                border-radius: 7px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255,120,180,180),
                    stop:1 rgba(180,100,160,180));
            }
            QLabel#Status {
                color: #5b2142;
                font-size: 13px;
                font-weight: 600;
                padding-bottom: 6px;
            }
            """
        )

    def pick_output_dir(self):
        selected = QFileDialog.getExistingDirectory(
            self, "Wybierz folder zapisu", self.output_input.text().strip() or DEFAULT_OUTPUT_DIR
        )
        if selected:
            self.output_input.setText(selected)

    def run_processing(self):
        urls = [line.strip() for line in self.url_input.toPlainText().splitlines() if line.strip()]
        output_dir = self.output_input.text().strip() or DEFAULT_OUTPUT_DIR

        if not urls:
            QMessageBox.warning(self, "Błąd", "Musisz podać co najmniej jeden URL pliku XML.")
            return

        self.output_input.setText(output_dir)
        self.run_btn.setEnabled(False)
        self.run_btn.setText("Przetwarzanie...")
        self.status.setText("Start przetwarzania...")
        self.progress.setValue(10)

        self.worker = ProcessorThread(urls, output_dir)
        self.worker.progress_signal.connect(self.on_progress)
        self.worker.done_signal.connect(self.on_done)
        self.worker.error_signal.connect(self.on_error)
        self.worker.start()

    def on_progress(self, message, value, tone):
        color = "#a35300" if tone == "warn" else "#8d3b68"
        self.status.setStyleSheet(f"color: {color}; font-weight: 600;")
        self.status.setText(message)
        self.progress.setValue(max(0, min(1000, int(value * 1000))))

    def on_done(self, payload):
        tone = payload.get("tone", "ok")
        color = "#1f7a4c" if tone == "ok" else "#a35300"
        self.status.setStyleSheet(f"color: {color}; font-weight: 700;")
        self.status.setText(payload.get("title", "Zakończono"))
        self.progress.setValue(int(payload.get("progress", 1.0) * 1000))

        if payload.get("ok", False):
            QMessageBox.information(self, payload.get("title", "Sukces"), payload.get("message", ""))
        else:
            QMessageBox.warning(self, payload.get("title", "Uwaga"), payload.get("message", ""))

        self._reset_run_button()

    def on_error(self, message):
        self.status.setStyleSheet("color: #a35300; font-weight: 700;")
        self.status.setText("Błąd krytyczny")
        self.progress.setValue(0)
        self._reset_run_button()
        QMessageBox.critical(self, "Błąd", message)

    def _reset_run_button(self):
        self.run_btn.setEnabled(True)
        self.run_btn.setText("Przetwórz na JEDEN plik CSV")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
