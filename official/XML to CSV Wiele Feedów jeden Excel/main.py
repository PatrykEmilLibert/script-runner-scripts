import xml.etree.ElementTree as ET
import csv
import os
import urllib.request
import customtkinter as ctk
from tkinter import filedialog, messagebox
from datetime import datetime
from urllib.parse import urlparse
import tempfile
import time
import concurrent.futures  # <--- DODANO: Do wielowątkowości
import openpyxl  # <--- DODANO: Do obsługi plików XLSX
from openpyxl.utils import get_column_letter # <--- DODANO: Do auto-dopasowania kolumn

# --- Ustawienia ---
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")
DOMYSLNA_SCIEZKA_ZAPISU = os.path.join(os.path.expanduser("~"), "Downloads")
MAX_WORKERS = 10  # <--- DODANO: Liczba jednoczesnych pobrań
# ------------------

def clean_text(text):
    """Zastępuje znaki nowej linii i inne białe znaki pojedynczą spacją."""
    if not text:
        return ""
    return " ".join(text.split())

def pobierz_xml(url, sciezka_docelowa):
    """
    Pobiera plik XML z podanego URL.
    Zwraca (True, None) przy sukcesie lub (False, str(e)) przy błędzie.
    """
    try:
        opener = urllib.request.build_opener()
        opener.addheaders = [('User-agent', 'Mozilla/5.0')]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(url, sciezka_docelowa)
        return True, None
    except Exception as e:
        return False, str(e) # Zwraca błąd zamiast messagebox

def parsuj_xml(sciezka_pliku):
    """
    Parsuje plik XML i ekstrahuje dane.
    Zwraca (atrybuty, maks_obr, dane, None) przy sukcesie 
    lub ([], 0, [], str(e)) przy błędzie.
    """
    try:
        drzewo = ET.parse(sciezka_pliku)
        root = drzewo.getroot()
        atrybuty = set()
        maks_liczba_obrazow = 0
        dane = []

        for element in root.findall("o"):
            cat_elem = element.find("cat")
            name_elem = element.find("name")
            desc_elem = element.find("desc")

            wiersz = {
                "id": element.get("id"),
                "url": element.get("url"),
                "price": element.get("price"),
                "avail": element.get("avail"),
                "weight": element.get("weight"),
                "stock": element.get("stock"),
                "cat": clean_text(cat_elem.text) if cat_elem is not None else "",
                "name": clean_text(name_elem.text) if name_elem is not None else "",
                "desc": clean_text(desc_elem.text) if desc_elem is not None else ""
            }
            
            atrybuty_elem = element.find("attrs")
            if atrybuty_elem is not None:
                for atrybut in atrybuty_elem.findall("a"):
                    nazwa_atrybutu = atrybut.get("name")
                    if nazwa_atrybutu:
                        atrybuty.add(nazwa_atrybutu)
                        wiersz[nazwa_atrybutu] = clean_text(atrybut.text)

            liczba_obrazow_w_wierszu = 0
            obrazy_elem = element.find("imgs")
            if obrazy_elem is not None:
                main_image = obrazy_elem.find("main")
                if main_image is not None and main_image.get("url"):
                    wiersz["image0"] = main_image.get("url")
                    liczba_obrazow_w_wierszu = 1
                
                start_index = 1 if "image0" in wiersz else 0

                for i, img in enumerate(obrazy_elem.findall("i"), start=start_index):
                    if img.get("url"):
                        wiersz[f"image{i}"] = img.get("url")
                        liczba_obrazow_w_wierszu = max(liczba_obrazow_w_wierszu, i + 1)
            
            maks_liczba_obrazow = max(maks_liczba_obrazow, liczba_obrazow_w_wierszu)
            dane.append(wiersz)
            
        # ZMODYFIKOWANO: Zwraca wynik zamiast pokazywać błąd
        return sorted(list(atrybuty)), maks_liczba_obrazow, dane, None
        
    except FileNotFoundError as e:
        return [], 0, [], f"Nie znaleziono pliku: {sciezka_pliku} ({e})"
    except ET.ParseError as e:
        return [], 0, [], f"Błąd parsowania XML w {os.path.basename(sciezka_pliku)}: {e}"
    except Exception as e:
        return [], 0, [], f"Nieoczekiwany błąd parsowania: {e}"

def zapisz_do_csv(dane, atrybuty_lista, maks_liczba_obrazow, sciezka_pliku):
    """Zapisuje połączone dane do jednego pliku CSV."""
    pola_podstawowe = ['id', 'url', 'price', 'avail', 'weight', 'stock', 'cat', 'name', 'desc']
    pola_atrybutow = atrybuty_lista 
    pola_obrazow = [f"image{i}" for i in range(maks_liczba_obrazow)]
    pola = pola_podstawowe + pola_atrybutow + pola_obrazow

    try:
        with open(sciezka_pliku, 'w', encoding='utf-8-sig', newline='') as plik_csv:
            writer = csv.DictWriter(plik_csv, fieldnames=pola, delimiter='|', extrasaction='ignore')
            writer.writeheader()
            writer.writerows(dane)
        # ZMODYFIKOWANO: Usunięto messagebox, podsumowanie będzie w funkcji nadrzędnej
        return True, None
    except Exception as e:
        return False, str(e)

# --- FUNKCJA ZAPISU BŁĘDÓW (z poprzedniej prośby) ---
def zapisz_bledy_do_xlsx(bledne_linki, sciezka_zapisu):
    if not bledne_linki:
        return None
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Błędy Pobierania"
        ws.append(["Nieudany URL", "Powód błędu"])
        for url, powod in bledne_linki:
            ws.append([url, powod])
        
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 80 else 80
            ws.column_dimensions[column_letter].width = adjusted_width

        teraz_format_czasu = datetime.now().strftime("%Y%m%d-%H%M%S")
        nazwa_pliku_xlsx = f"RAPORT_BLEDOW_POBIERANIA_{teraz_format_czasu}.xlsx"
        sciezka_pliku_xlsx = os.path.join(sciezka_zapisu, nazwa_pliku_xlsx)
        
        wb.save(sciezka_pliku_xlsx)
        return sciezka_pliku_xlsx
    except Exception as e:
        messagebox.showerror("Błąd zapisu raportu błędów", 
                             f"Nie udało się zapisać pliku XLSX z błędami: {e}")
        return None
# ----------------------------------------

# --- NOWA FUNKCJA ROBOCZA (WORKER) ---
def pobierz_i_parsuj_url(url):
    """
    Pobiera i parsuje jeden URL. Przeznaczona do uruchamiania w osobnym wątku.
    Zwraca status i dane.
    """
    katalog_tymczasowy = tempfile.gettempdir()
    nazwa_pliku_url = os.path.basename(urlparse(url).path) or f"feed_{hash(url)}.xml"
    nazwa_bazowa_xml = os.path.splitext(nazwa_pliku_url)[0]
    teraz_timestamp_temp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    sciezka_lokalna_xml = os.path.join(katalog_tymczasowy, f"temp_{nazwa_bazowa_xml}_{teraz_timestamp_temp}.xml")

    # 1. Pobieranie
    sukces_pobierania, powod_bledu_pob = pobierz_xml(url, sciezka_lokalna_xml)
    if not sukces_pobierania:
        return ('blad_pobierania', (url, powod_bledu_pob, nazwa_pliku_url))

    # 2. Parsowanie
    atrybuty, maks_obr, dane, powod_bledu_pars = parsuj_xml(sciezka_lokalna_xml)
    
    # 3. Czyszczenie pliku tymczasowego
    try:
        os.remove(sciezka_lokalna_xml)
    except Exception as e:
        print(f"Ostrzeżenie: Nie udało się usunąć pliku tymczasowego {sciezka_lokalna_xml}: {e}")

    # 4. Zwracanie wyników
    if powod_bledu_pars:
        return ('blad_parsowania', (url, powod_bledu_pars, nazwa_pliku_url))
    
    if not dane:
         return ('blad_parsowania', (url, "Brak elementów <o> po parsowaniu", nazwa_pliku_url))

    return ('sukces', (dane, atrybuty, maks_obr, nazwa_bazowa_xml, nazwa_pliku_url))
# ----------------------------------------


# --- PRZEBUDOWANA FUNKCJA PRZETWARZANIA ---
def przetworz_wiele_url_jeden_plik(urls, sciezka_zapisu_csv, app_instance):
    """Przetwarza wiele URL-i (wielowątkowo) i zapisuje do jednego pliku CSV."""
    all_dane = []
    all_atrybuty = set()
    global_maks_liczba_obrazow = 0
    all_nazwy_bazowe = []
    
    liczba_url = len(urls)
    sukcesy_przetwarzania = 0
    bledy_pobierania = 0
    bledy_parsowania = 0
    bledy_zapisu = 0
    bledne_linki = []

    if liczba_url == 0:
        app_instance.update_status("Nie podano żadnych URL-i.", 0)
        return

    if not os.path.exists(sciezka_zapisu_csv):
        try:
            os.makedirs(sciezka_zapisu_csv)
        except Exception as e:
            messagebox.showerror("Błąd ścieżki zapisu", f"Nie można utworzyć katalogu: {sciezka_zapisu_csv}\nBłąd: {e}\nPliki będą zapisywane w katalogu roboczym.")
            sciezka_zapisu_csv = os.getcwd()
            app_instance.pole_sciezki_zapisu.delete(0, ctk.END)
            app_instance.pole_sciezki_zapisu.insert(0, sciezka_zapisu_csv)

    app_instance.update_status(f"Rozpoczynam przetwarzanie {liczba_url} linków (max {MAX_WORKERS} wątków)...", 0)

    # Użycie ThreadPoolExecutor do jednoczesnego pobierania i parsowania
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        # Tworzenie listy "przyszłych" zadań
        futures = [executor.submit(pobierz_i_parsuj_url, url) for url in urls]

        # Przetwarzanie wyników w kolejności ich kończenia się
        for i, future in enumerate(concurrent.futures.as_completed(futures)):
            postep = (i + 1) / liczba_url
            
            try:
                status, data = future.result()
                
                if status == 'sukces':
                    dane, atrybuty, maks_obr, nazwa_bazowa_xml, nazwa_pliku_url = data
                    all_dane.extend(dane)
                    all_atrybuty.update(atrybuty)
                    global_maks_liczba_obrazow = max(global_maks_liczba_obrazow, maks_obr)
                    all_nazwy_bazowe.append(nazwa_bazowa_xml)
                    sukcesy_przetwarzania += 1
                    app_instance.update_status(f"Pobrano {i+1}/{liczba_url}: {nazwa_pliku_url}", postep)

                elif status == 'blad_pobierania':
                    url, powod, nazwa_pliku_url = data
                    bledy_pobierania += 1
                    bledne_linki.append((url, powod))
                    app_instance.update_status(f"Błąd pobierania {i+1}/{liczba_url}: {nazwa_pliku_url}", postep)
                    time.sleep(0.1) # Krótka pauza na odczytanie statusu

                elif status == 'blad_parsowania':
                    url, powod, nazwa_pliku_url = data
                    bledy_parsowania += 1
                    # Błędów parsowania nie dodajemy do raportu XLSX, ale liczymy je
                    app_instance.update_status(f"Błąd parsowania {i+1}/{liczba_url}: {nazwa_pliku_url}", postep)
                    time.sleep(0.1)

            except Exception as e:
                # To jest nieoczekiwany błąd w samej logice wątku
                bledy_parsowania += 1 # Traktujemy to jako błąd przetwarzania
                app_instance.update_status(f"Błąd krytyczny wątku {i+1}/{liczba_url}: {e}", postep)
                print(f"Błąd krytyczny w wątku: {e}")


    # --- Zapisywanie wyników (po zakończeniu wszystkich wątków) ---
    
    sciezka_pliku_bledow = zapisz_bledy_do_xlsx(bledne_linki, sciezka_zapisu_csv)
    liczba_bledow_ogolem = bledy_pobierania + bledy_parsowania + bledy_zapisu
    
    podsumowanie_bledow_tekst = (
        f"Błędy pobierania: {bledy_pobierania}\n"
        f"Błędy parsowania: {bledy_parsowania}\n"
        f"Błędy zapisu: {bledy_zapisu}"
    )

    if not all_dane:
        app_instance.update_status(f"Nie udało się przetworzyć żadnych danych. Błędy: {liczba_bledow_ogolem}", 0)
        wiadomosc_ostrzezenia = (
            f"Nie udało się pobrać ani sparsować danych z żadnego podanego URL.\n\n"
            f"{podsumowanie_bledow_tekst}"
        )
        if sciezka_pliku_bledow:
            wiadomosc_ostrzezenia += f"\n\nZapisano raport błędów pobierania:\n{os.path.basename(sciezka_pliku_bledow)}"
        messagebox.showwarning("Brak danych", wiadomosc_ostrzezenia)
        app_instance.reset_gui_after_delay()
        return

    # Tworzenie nazwy pliku
    nazwa_laczona = "_".join(all_nazwy_bazowe)
    if len(nazwa_laczona) > 100:
        nazwa_laczona = f"{all_nazwy_bazowe[0]}_and_{len(all_nazwy_bazowe)-1}_more"
        
    teraz_format_czasu_csv = datetime.now().strftime("%d%m%y-%H%M%S")
    nazwa_pliku_csv = os.path.join(sciezka_zapisu_csv, f"{nazwa_laczona}_{teraz_format_czasu_csv}.csv")

    app_instance.update_status("Zapisywanie połączonych danych...", 0.98)
    
    sukces_zapisu, powod_bledu_zapisu = zapisz_do_csv(all_dane, sorted(list(all_atrybuty)), global_maks_liczba_obrazow, nazwa_pliku_csv)
    
    if not sukces_zapisu:
        bledy_zapisu += 1
        # Aktualizujemy tekst podsumowania błędów, jeśli zapis się nie udał
        podsumowanie_bledow_tekst = (
            f"Błędy pobierania: {bledy_pobierania}\n"
            f"Błędy parsowania: {bledy_parsowania}\n"
            f"Błędy zapisu: {bledy_zapisu} ({powod_bledu_zapisu})"
        )

    app_instance.update_status(f"Zakończono. Przetworzono: {sukcesy_przetwarzania}, Błędy: {liczba_bledow_ogolem}", 1)

    # --- Końcowe podsumowanie ---
    if sukcesy_przetwarzania > 0 and sukces_zapisu:
        wiadomosc_sukcesu = (
            f"Pomyślnie przetworzono {sukcesy_przetwarzania} z {liczba_url} plików.\n"
            f"Zapisano połączone dane do:\n{os.path.abspath(nazwa_pliku_csv)}\n\n"
            f"{podsumowanie_bledow_tekst}"
        )
        if sciezka_pliku_bledow:
            wiadomosc_sukcesu += f"\n\nZapisano raport błędów pobierania:\n{os.path.basename(sciezka_pliku_bledow)}"
        messagebox.showinfo("Sukces", wiadomosc_sukcesu)
    else:
        wiadomosc_ostrzezenia = (
            f"Zakończono przetwarzanie, ale wystąpiły błędy.\n"
            f"Pomyślnie przetworzono: {sukcesy_przetwarzania} z {liczba_url}\n\n"
            f"{podsumowanie_bledow_tekst}"
        )
        if sciezka_pliku_bledow:
            wiadomosc_ostrzezenia += f"\n\nZapisano raport błędów pobierania:\n{os.path.basename(sciezka_pliku_bledow)}"
        messagebox.showwarning("Zakończono z błędami", wiadomosc_ostrzezenia)

    app_instance.reset_gui_after_delay()


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Konwerter XML do CSV")
        
        window_width = 600
        window_height = 550
        self.minsize(500, 450)

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        center_x = int(screen_width/2 - window_width / 2)
        center_y = int(screen_height/2 - window_height / 2)
        self.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)  
        self.grid_rowconfigure(1, weight=0)  
        self.grid_rowconfigure(2, weight=0)  
        self.grid_rowconfigure(3, weight=0)  
        self.grid_rowconfigure(4, weight=0)  

        input_frame_urls = ctk.CTkFrame(self)
        input_frame_urls.grid(row=0, column=0, padx=20, pady=(20,10), sticky="nsew")
        input_frame_urls.grid_columnconfigure(0, weight=1)
        input_frame_urls.grid_rowconfigure(1, weight=1) 

        etykieta_url = ctk.CTkLabel(input_frame_urls, text="Wklej URL-e plików XML (każdy w nowej linii):")
        etykieta_url.grid(row=0, column=0, padx=10, pady=(10, 5), sticky="w")
        self.pole_url = ctk.CTkTextbox(input_frame_urls, height=150) 
        self.pole_url.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")

        input_frame_path = ctk.CTkFrame(self)
        input_frame_path.grid(row=1, column=0, padx=20, pady=(5,10), sticky="ew")
        input_frame_path.grid_columnconfigure(1, weight=1)

        etykieta_sciezki = ctk.CTkLabel(input_frame_path, text="Katalog zapisu pliku CSV:")
        etykieta_sciezki.grid(row=0, column=0, padx=(10,5), pady=5, sticky="w")
        
        self.pole_sciezki_zapisu = ctk.CTkEntry(input_frame_path)
        self.pole_sciezki_zapisu.grid(row=0, column=1, padx=(0,5), pady=5, sticky="ew")
        self.pole_sciezki_zapisu.insert(0, DOMYSLNA_SCIEZKA_ZAPISU)

        przycisk_wybierz_sciezke = ctk.CTkButton(input_frame_path, text="Wybierz folder", width=120, command=self.wybierz_katalog_zapisu)
        przycisk_wybierz_sciezke.grid(row=0, column=2, padx=(0,10), pady=5, sticky="e")

        self.przycisk_przetworz = ctk.CTkButton(self, text="Przetwórz na JEDEN plik CSV", command=self.rozpocznij_przetwarzanie_action, height=40)
        self.przycisk_przetworz.grid(row=2, column=0, padx=20, pady=10, sticky="ew")

        self.progress_bar = ctk.CTkProgressBar(self, height=10)
        self.progress_bar.grid(row=3, column=0, padx=20, pady=(0, 5), sticky="ew")
        self.progress_bar.set(0)

        self.status_label = ctk.CTkLabel(self, text="Gotowy.", text_color="gray")
        self.status_label.grid(row=4, column=0, padx=20, pady=(0,10), sticky="ew")

    def wybierz_katalog_zapisu(self):
        sciezka_katalogu = filedialog.askdirectory(initialdir=self.pole_sciezki_zapisu.get() or DOMYSLNA_SCIEZKA_ZAPISU)
        if sciezka_katalogu:
            self.pole_sciezki_zapisu.delete(0, ctk.END)
            self.pole_sciezki_zapisu.insert(0, sciezka_katalogu)

    def update_status(self, message, progress_value=None):
        """Aktualizuje etykietę statusu i pasek postępu."""
        self.status_label.configure(text=message)
        if progress_value is not None:
            self.progress_bar.set(progress_value)
        self.update_idletasks() # Wymusza odświeżenie GUI

    def reset_gui_after_delay(self, delay_ms=4000):
        """Resetuje pasek postępu i status po opóźnieniu."""
        self.after(delay_ms, lambda: self.update_status("Gotowy.", 0))
        self.przycisk_przetworz.configure(state="normal", text="Przetwórz na JEDEN plik CSV")

    def rozpocznij_przetwarzanie_action(self):
        """Rozpoczyna proces przetwarzania wielu URL-i na jeden plik."""
        urls_text = self.pole_url.get("1.0", ctk.END) 
        urls = [url.strip() for url in urls_text.splitlines() if url.strip()] 

        sciezka_zapisu_csv_gui = self.pole_sciezki_zapisu.get().strip()
        if not sciezka_zapisu_csv_gui:
            sciezka_zapisu_csv_gui = DOMYSLNA_SCIEZKA_ZAPISU
            self.pole_sciezki_zapisu.insert(0, sciezka_zapisu_csv_gui)

        if not urls:
            messagebox.showerror("Błąd", "Musisz podać co najmniej jeden URL pliku XML.")
            return

        self.przycisk_przetworz.configure(state="disabled", text="Przetwarzanie...")
        self.update_idletasks()
        
        # Uruchomienie funkcji przetwarzającej (będzie blokować GUI, 
        # ale 'update_idletasks' w 'update_status' zapewni odświeżanie paska)
        przetworz_wiele_url_jeden_plik(urls, sciezka_zapisu_csv_gui, self)

if __name__ == "__main__":
    app = App()
    app.mainloop()