import xml.etree.ElementTree as ET
import os
import urllib.request
import customtkinter as ctk
from tkinter import filedialog, messagebox
from datetime import datetime
from urllib.parse import urlparse
import tempfile
import time
import openpyxl
import re

# Ustawienia CustomTkinter
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

# --- Standardowa ścieżka zapisu ---
DOMYSLNA_SCIEZKA_ZAPISU = os.path.join(os.path.expanduser("~"), "Downloads")
# ----------------------------------------------------

# --- LOGIKA KOREKTY ZNAKÓW ---

# Wzorzec Regex do usuwania większości znaków emoji
EMOJI_PATTERN = re.compile(
    "["
    "\U0001F600-\U0001F64F"  # emotikony
    "\U0001F300-\U0001F5FF"  # symbole i piktogramy
    "\U0001F680-\U0001F6FF"  # transport i symbole map
    "\U0001F700-\U0001F77F"  # symbole alchemiczne
    "\U0001F780-\U0001F7FF"  # Rozszerzone kształty geometryczne
    "\U0001F800-\U0001F8FF"  # Dodatkowe strzałki-C
    "\U0001F900-\U0001F9FF"  # Dodatkowe symbole i piktogramy
    "\U0001FA00-\U0001FA6F"  # Symbole szachowe
    "\U0001FA70-\U0001FAFF"  # Symbole i piktogramy rozszerzone-A
    "\U00002702-\U000027B0"  # Dingbaty
    "\U000024C2-\U0001F251"
    "\U0001f926-\U0001f937"
    "\U00010000-\U0010ffff"
    "\u2640-\u2642"
    "\u2600-\u2B55"
    "\u200d"
    "\u23cf"
    "\u23e9"
    "\u231a"
    "\ufe0f"  # selektor wariacji
    "\u3030"
    "]+",
    flags=re.UNICODE,
)

# Słownik mapujący nieprawidłowe znaki/encje na poprawne polskie litery i symbole.
POLISH_CHAR_MAP = {
    # Złożone/długie encje jako pierwsze
    '&#378;ó&#322;ty': 'żółty',

    # Nazwane encje HTML (Wielkie litery)
    '&Aacute;': 'Ą', '&Cacute;': 'Ć', '&Eacute;': 'Ę', '&Lacute;': 'Ł',
    '&Nacute;': 'Ń', '&Oacute;': 'Ó', '&Sacute;': 'Ś', '&Zacute;': 'Ź',
    '&Zdot;': 'Ż',

    # Nazwane encje HTML (Małe litery)
    '&aacute;': 'ą', '&cacute;': 'ć', '&eacute;': 'ę', '&lacute;': 'ł',
    '&nacute;': 'ń', '&oacute;': 'ó', '&sacute;': 'ś', '&zacute;': 'ź',
    '&zdot;': 'ż',

    # Numeryczne encje HTML (Wielkie litery)
    '&#260;': 'Ą', '&#262;': 'Ć', '&#280;': 'Ę', '&#321;': 'Ł',
    '&#323;': 'Ń', '&#211;': 'Ó', '&#346;': 'Ś', '&#377;': 'Ź',
    '&#379;': 'Ż',

    # Numeryczne encje HTML (Małe litery)
    '&#261;': 'ą', '&#263;': 'ć', '&#281;': 'ę', '&#322;': 'ł',
    '&#324;': 'ń', '&#243;': 'ó', '&#347;': 'ś', '&#378;': 'ź',
    '&#380;': 'ż',
    
    # Inne popularne symbole i encje
    '&deg;': '°', '&bull;': '•', '&ndash;': '–', '&rsquo;': '’',
    '&bdquo;': '„', '&rdquo;': '”', '&#10036;&#65039;': '', '&#10035;&#65039;': '',
    '&#9851;&#65039;': '', '&#128209;': '', '&#8222;': '„', '&#8221;': '”',
    '&#8216;': '‘', '&#8217;': '’', '&#8211;': '–', '&#8203;': '',
    '&#9989;': '', '&#9749;': '', '&#11088;': '', '&#10003;': '',
    '&#34;': '"', '&#39;': "'", '&#x2013;': '–', '&#2013;': '–',
    '&#2019;': '’', '&nbsp;': ' ', '&amp;': '&', '&lt;': '<',
    '&gt;': '>', '&quot;': '"', '&apos;': "'", '&#178;': '²',
    '&#8220;': '“', '&#8230;': '…', '&#9679;': '•',

    # Znaki do usunięcia
    '✔': '', '✅': '', '❓': '', '▶️': '', '⭐': '', '⚡': '', '➡': '',
}
# Sortowanie słownika, aby najpierw przetwarzać dłuższe klucze
POLISH_CHAR_MAP_SORTED = dict(sorted(POLISH_CHAR_MAP.items(), key=lambda item: len(item[0]), reverse=True))

def correct_text(text):
    """Stosuje wszystkie zdefiniowane poprawki do pojedynczego ciągu tekstowego."""
    if not isinstance(text, str) or not text:
        return text
    
    corrected_text = text
    
    # 1. Zamień wszystkie znane nieprawidłowe sekwencje znaków
    for wrong_str, correct_char in POLISH_CHAR_MAP_SORTED.items():
        if wrong_str in corrected_text:
            corrected_text = corrected_text.replace(wrong_str, correct_char)
            
    # 2. Usuń wszystkie znaki emoji za pomocą wzorca regex
    corrected_text = EMOJI_PATTERN.sub(r'', corrected_text)
    
    # 3. Usuń wiodące białe znaki
    corrected_text = corrected_text.lstrip()
    
    return corrected_text

# --- FUNKCJE PODSTAWOWE ---

def clean_text(text):
    """Zastępuje znaki nowej linii i inne białe znaki pojedynczą spacją."""
    if not text:
        return ""
    return " ".join(text.split())

def pobierz_xml(url, sciezka_docelowa):
    """Pobiera plik XML z podanego URL."""
    try:
        opener = urllib.request.build_opener()
        opener.addheaders = [('User-agent', 'Mozilla/5.0')]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(url, sciezka_docelowa)
        return True
    except Exception as e:
        messagebox.showerror(f"Błąd pobierania ({url})", f"Wystąpił błąd: {e}")
        return False

def parsuj_xml(sciezka_pliku):
    """Parsuje plik XML i ekstrahuje id, name, desc, EAN oraz id_bl."""
    try:
        drzewo = ET.parse(sciezka_pliku)
        root = drzewo.getroot()
        dane = []

        for element in root.findall("o"):
            name_elem = element.find("name")
            desc_elem = element.find("desc")

            wiersz = {
                "id": element.get("id"),
                "EAN": "",
                "name": clean_text(name_elem.text) if name_elem is not None else "",
                "desc": clean_text(desc_elem.text) if desc_elem is not None else "",
                "id_bl": ""
            }

            ean_val = element.get("EAN")
            if ean_val:
                wiersz["EAN"] = ean_val
            else:
                atrybuty_elem = element.find("attrs")
                if atrybuty_elem is not None:
                    for atrybut in atrybuty_elem.findall("a"):
                        if atrybut.get("name") == "EAN":
                            wiersz["EAN"] = clean_text(atrybut.text)
                            break

            atrybuty_idbl = element.find("attrs")
            if atrybuty_idbl is not None:
                for atrybut_idbl in atrybuty_idbl.findall("a"):
                    if atrybut_idbl.get("name") == "id_bl":
                        wiersz["id_bl"] = clean_text(atrybut_idbl.text)
                        break

            dane.append(wiersz)
            
        return dane
        
    except FileNotFoundError:
        messagebox.showerror("Błąd pliku", f"Nie znaleziono pliku: {sciezka_pliku}")
        return []
    except ET.ParseError as e:
        messagebox.showerror("Błąd parsowania XML", f"Błąd w pliku {os.path.basename(sciezka_pliku)}: {e}")
        return []
    except Exception as e:
        messagebox.showerror("Nieoczekiwany błąd parsowania", f"Wystąpił błąd: {e}")
        return []

def zapisz_do_excel(dane, sciezka_pliku):
    """Zapisuje dane do pliku Excel, jednocześnie korygując znaki."""
    pola = ['id', 'EAN', 'id_bl', 'name', 'desc']
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Produkty"
    sheet.append(pola) # Zapis nagłówków

    try:
        for wiersz_danych in dane:
            wiersz_do_zapisu = []
            for pole in pola:
                wartosc = wiersz_danych.get(pole, "")
                # Stosuj korektę tylko do ciągów tekstowych
                poprawiona_wartosc = correct_text(wartosc) if isinstance(wartosc, str) else wartosc
                wiersz_do_zapisu.append(poprawiona_wartosc)
            sheet.append(wiersz_do_zapisu)
        
        workbook.save(sciezka_pliku)
        return True
    except Exception as e:
        messagebox.showerror(f"Błąd zapisu Excel ({os.path.basename(sciezka_pliku)})", f"Wystąpił błąd: {e}")
        return False
    finally:
        if workbook:
            workbook.close()

def stworz_plik_podsumowania(sciezki_plikow, sciezka_zapisu_folder):
    """Tworzy plik Excel z podsumowaniem liczby wierszy w podanych plikach."""
    if not sciezki_plikow:
        return None

    data_dzis = datetime.now().strftime("%Y-%m-%d")
    nazwa_pliku_podsumowania = os.path.join(sciezka_zapisu_folder, f"Podsumowanie_Ilości_{data_dzis}.xlsx")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Podsumowanie"
    sheet.append(["Nazwa Pliku", "Ilość Wierszy (bez nagłówka)"])

    try:
        for sciezka_pliku in sciezki_plikow:
            nazwa_pliku_z_rozszerzeniem = os.path.basename(sciezka_pliku)
            nazwa_pliku, _ = os.path.splitext(nazwa_pliku_z_rozszerzeniem)
            try:
                wb_danych = openpyxl.load_workbook(sciezka_pliku, read_only=True)
                ws_danych = wb_danych.active
                # sheet.max_row liczy wszystkie wiersze, w tym nagłówek
                ilosc_wierszy = ws_danych.max_row - 1
                sheet.append([nazwa_pliku, ilosc_wierszy])
                wb_danych.close()
            except Exception as e:
                # Jeśli nie uda się odczytać pliku, zapisz informację o błędzie
                sheet.append([nazwa_pliku, f"Błąd odczytu: {e}"])

        workbook.save(nazwa_pliku_podsumowania)
        # Zwróć ścieżkę do pliku podsumowania, aby można było ją wyświetlić w komunikacie
        return nazwa_pliku_podsumowania
    except Exception as e:
        messagebox.showerror("Błąd tworzenia podsumowania", f"Nie udało się utworzyć pliku podsumowania.\nBłąd: {e}")
        return None
    finally:
        if workbook:
            workbook.close()

def przetworz_wiele_url_osobne_pliki(urls, sciezka_zapisu, app_instance):
    """Przetwarza wiele URL-i i zapisuje każdy do osobnego, poprawionego pliku Excel."""
    katalog_tymczasowy = tempfile.gettempdir()
    liczba_url = len(urls)
    sukcesy = 0
    bledy_pobierania, bledy_parsowania, bledy_zapisu = 0, 0, 0
    pomyslnie_zapisane_pliki = []

    if liczba_url == 0:
        app_instance.update_status("Nie podano żadnych URL-i.", 0)
        return

    if not os.path.exists(sciezka_zapisu):
        try:
            os.makedirs(sciezka_zapisu)
        except Exception as e:
            messagebox.showerror("Błąd ścieżki zapisu", f"Nie można utworzyć katalogu: {sciezka_zapisu}\nBłąd: {e}")
            return

    for i, url in enumerate(urls):
        url = url.strip()
        if not url:
            continue

        postep = (i + 1) / liczba_url
        nazwa_pliku_url = os.path.basename(urlparse(url).path) or f"feed_{i+1}.xml"
        app_instance.update_status(f"Przetwarzanie {i+1}/{liczba_url}: {nazwa_pliku_url}...", postep)

        nazwa_bazowa_xml = os.path.splitext(nazwa_pliku_url)[0]
        
        # Wyodrębnij nazwę hurtowni
        nazwa_hurtowni = nazwa_bazowa_xml.split('_')[0]
        data_dzis = datetime.now().strftime("%Y-%m-%d")

        teraz_timestamp_temp = datetime.now().strftime("%Y%m%d%H%M%S%f")
        sciezka_lokalna_xml = os.path.join(katalog_tymczasowy, f"temp_{nazwa_bazowa_xml}_{teraz_timestamp_temp}.xml")

        if pobierz_xml(url, sciezka_lokalna_xml):
            dane = parsuj_xml(sciezka_lokalna_xml)
            if dane:
                nazwa_pliku_excel = os.path.join(sciezka_zapisu, f"{nazwa_hurtowni}_{data_dzis}.xlsx")
                app_instance.update_status(f"Zapisywanie: {os.path.basename(nazwa_pliku_excel)}...", postep)
                if zapisz_do_excel(dane, nazwa_pliku_excel):
                    sukcesy += 1
                    pomyslnie_zapisane_pliki.append(nazwa_pliku_excel)
                else:
                    bledy_zapisu += 1
            else:
                bledy_parsowania += 1
                app_instance.update_status(f"Błąd parsowania {nazwa_pliku_url}. Pomijanie.", postep)
                time.sleep(0.5)
            
            try:
                os.remove(sciezka_lokalna_xml)
            except Exception as e:
                print(f"Ostrzeżenie: Nie udało się usunąć pliku tymczasowego {sciezka_lokalna_xml}: {e}")
        else:
            bledy_pobierania += 1
            app_instance.update_status(f"Błąd pobierania {nazwa_pliku_url}. Pomijanie.", postep)
            time.sleep(0.5)

    plik_podsumowania = None
    if pomyslnie_zapisane_pliki:
        plik_podsumowania = stworz_plik_podsumowania(pomyslnie_zapisane_pliki, sciezka_zapisu)

    if sukcesy > 0:
        komunikat_podsumowania = ""
        if plik_podsumowania:
            komunikat_podsumowania = f"\n\nUtworzono również plik podsumowania:\n{os.path.basename(plik_podsumowania)}"
            
        messagebox.showinfo("Zakończono przetwarzanie", f"Pomyślnie przetworzono i zapisano {sukcesy} z {liczba_url} plików w:\n{os.path.abspath(sciezka_zapisu)}\n\n"
                                                      f"Błędy pobierania: {bledy_pobierania}\n"
                                                      f"Błędy parsowania: {bledy_parsowania}\n"
                                                      f"Błędy zapisu: {bledy_zapisu}"
                                                      f"{komunikat_podsumowania}")
    else:
        messagebox.showwarning("Zakończono przetwarzanie", f"Nie udało się pomyślnie przetworzyć żadnego pliku.\n"
                                                        f"Błędy pobierania: {bledy_pobierania}\n"
                                                        f"Błędy parsowania: {bledy_parsowania}\n"
                                                        f"Błędy zapisu: {bledy_zapisu}")

    app_instance.update_status(f"Zakończono. Zapisano: {sukcesy}. Błędy: {bledy_pobierania+bledy_parsowania+bledy_zapisu}", 1)
    app_instance.reset_gui_after_delay()

# --- INTERFEJS UŻYTKOWNIKA (GUI) ---
class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Konwerter XML do Excel z Korektą Znaków")
        window_width, window_height = 600, 550
        self.minsize(500, 450)

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)
        self.geometry(f"{window_width}x{window_height}+{center_x}+{center_y}")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        input_frame_urls = ctk.CTkFrame(self)
        input_frame_urls.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="nsew")
        input_frame_urls.grid_columnconfigure(0, weight=1)
        input_frame_urls.grid_rowconfigure(1, weight=1)

        etykieta_url = ctk.CTkLabel(input_frame_urls, text="Wklej URL-e plików XML (każdy w nowej linii):")
        etykieta_url.grid(row=0, column=0, padx=10, pady=(10, 5), sticky="w")
        self.pole_url = ctk.CTkTextbox(input_frame_urls, height=150)
        self.pole_url.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")

        input_frame_path = ctk.CTkFrame(self)
        input_frame_path.grid(row=1, column=0, padx=20, pady=(5, 10), sticky="ew")
        input_frame_path.grid_columnconfigure(1, weight=1)

        etykieta_sciezki = ctk.CTkLabel(input_frame_path, text="Katalog zapisu plików Excel:")
        etykieta_sciezki.grid(row=0, column=0, padx=(10, 5), pady=5, sticky="w")
        
        self.pole_sciezki_zapisu = ctk.CTkEntry(input_frame_path)
        self.pole_sciezki_zapisu.grid(row=0, column=1, padx=(0, 5), pady=5, sticky="ew")
        self.pole_sciezki_zapisu.insert(0, DOMYSLNA_SCIEZKA_ZAPISU)

        przycisk_wybierz_sciezke = ctk.CTkButton(input_frame_path, text="Wybierz folder", width=120, command=self.wybierz_katalog_zapisu)
        przycisk_wybierz_sciezke.grid(row=0, column=2, padx=(0, 10), pady=5, sticky="e")

        self.przycisk_przetworz = ctk.CTkButton(self, text="Przetwórz na OSOBNE pliki Excel", command=self.rozpocznij_przetwarzanie_action, height=40)
        self.przycisk_przetworz.grid(row=2, column=0, padx=20, pady=10, sticky="ew")

        self.progress_bar = ctk.CTkProgressBar(self, height=10)
        self.progress_bar.grid(row=3, column=0, padx=20, pady=(0, 5), sticky="ew")
        self.progress_bar.set(0)

        self.status_label = ctk.CTkLabel(self, text="Gotowy.", text_color="gray")
        self.status_label.grid(row=4, column=0, padx=20, pady=(0, 10), sticky="ew")

    def wybierz_katalog_zapisu(self):
        """Otwiera okno dialogowe do wyboru katalogu zapisu."""
        sciezka_katalogu = filedialog.askdirectory(initialdir=self.pole_sciezki_zapisu.get() or DOMYSLNA_SCIEZKA_ZAPISU)
        if sciezka_katalogu:
            self.pole_sciezki_zapisu.delete(0, ctk.END)
            self.pole_sciezki_zapisu.insert(0, sciezka_katalogu)

    def update_status(self, message, progress_value=None):
        """Aktualizuje etykietę statusu i pasek postępu."""
        self.status_label.configure(text=message)
        if progress_value is not None:
            self.progress_bar.set(progress_value)
        self.update_idletasks()

    def reset_gui_after_delay(self, delay_ms=4000):
        """Resetuje pasek postępu i status po opóźnieniu."""
        self.after(delay_ms, lambda: self.update_status("Gotowy.", 0))
        self.przycisk_przetworz.configure(state="normal", text="Przetwórz na OSOBNE pliki Excel")

    def rozpocznij_przetwarzanie_action(self):
        """Rozpoczyna proces przetwarzania URL-i na osobne pliki Excel."""
        urls_text = self.pole_url.get("1.0", ctk.END)
        urls = [url.strip() for url in urls_text.splitlines() if url.strip()]

        sciezka_zapisu_gui = self.pole_sciezki_zapisu.get().strip()
        if not sciezka_zapisu_gui:
            sciezka_zapisu_gui = DOMYSLNA_SCIEZKA_ZAPISU
            self.pole_sciezki_zapisu.insert(0, sciezka_zapisu_gui)

        if not urls:
            messagebox.showerror("Błąd", "Musisz podać co najmniej jeden URL pliku XML.")
            return

        self.przycisk_przetworz.configure(state="disabled", text="Przetwarzanie...")
        self.update_idletasks()
        
        przetworz_wiele_url_osobne_pliki(urls, sciezka_zapisu_gui, self)

if __name__ == "__main__":
    app = App()
    app.mainloop()

