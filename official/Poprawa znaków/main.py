import customtkinter
from tkinter import filedialog, messagebox
import openpyxl
import os
import sys
import re

# This script requires the following libraries:
# pip install customtkinter openpyxl
# For the "Excel Engine" on Windows, it also requires:
# pip install pywin32

# --- Engine Detection ---
PYWIN32_AVAILABLE = False
if sys.platform == "win32":
    try:
        import win32com.client
        PYWIN32_AVAILABLE = True
    except ImportError:
        print("Biblioteka pywin32 nie jest zainstalowana. Silnik Excel będzie niedostępny.")
        print("Aby go zainstalować, uruchom: pip install pywin32")

# --- Emoji Removal Pattern ---
# A comprehensive regex to find and remove most emoji characters.
EMOJI_PATTERN = re.compile(
    "["
    "\U0001F600-\U0001F64F"  # emoticons
    "\U0001F300-\U0001F5FF"  # symbols & pictographs
    "\U0001F680-\U0001F6FF"  # transport & map symbols
    "\U0001F700-\U0001F77F"  # alchemical symbols
    "\U0001F780-\U0001F7FF"  # Geometric Shapes Extended
    "\U0001F800-\U0001F8FF"  # Supplemental Arrows-C
    "\U0001F900-\U0001F9FF"  # Supplemental Symbols and Pictographs
    "\U0001FA00-\U0001FA6F"  # Chess Symbols
    "\U0001FA70-\U0001FAFF"  # Symbols and Pictographs Extended-A
    "\U00002702-\U000027B0"  # Dingbats
    "\U000024C2-\U0001F251"
    "\U0001f926-\U0001f937"
    "\U00010000-\U0010ffff"
    "\u2640-\u2642"
    "\u2600-\u2B55"
    "\u200d"
    "\u23cf"
    "\u23e9"
    "\u231a"
    "\ufe0f"  # variation selector
    "\u3030"
    "]+",
    flags=re.UNICODE,
)

# --- Character Mapping ---
# Comprehensive dictionary mapping incorrect characters/entities to correct Polish letters.
POLISH_CHAR_MAP = {
    # Complex/long entities first
    '&#378;ó&#322;ty': 'żółty',

    # HTML Named Entities (Uppercase)
    '&Aacute;': 'Ą', '&Cacute;': 'Ć', '&Eacute;': 'Ę', '&Lacute;': 'Ł',
    '&Nacute;': 'Ń', '&Oacute;': 'Ó', '&Sacute;': 'Ś', '&Zacute;': 'Ź',
    '&Zdot;': 'Ż',

    # HTML Named Entities (Lowercase)
    '&aacute;': 'ą', '&cacute;': 'ć', '&eacute;': 'ę', '&lacute;': 'ł',
    '&nacute;': 'ń', '&oacute;': 'ó', '&sacute;': 'ś', '&zacute;': 'ź',
    '&zdot;': 'ż',

    # HTML Numeric Entities (Uppercase)
    '&#260;': 'Ą', '&#262;': 'Ć', '&#280;': 'Ę', '&#321;': 'Ł',
    '&#323;': 'Ń', '&#211;': 'Ó', '&#346;': 'Ś', '&#377;': 'Ź',
    '&#379;': 'Ż',

    # HTML Numeric Entities (Lowercase)
    '&#261;': 'ą', '&#263;': 'ć', '&#281;': 'ę', '&#322;': 'ł',
    '&#324;': 'ń', '&#243;': 'ó', '&#347;': 'ś', '&#378;': 'ź',
    '&#380;': 'ż',

    # --- NEWLY ADDED AND OTHER COMMON SYMBOLS ---
    # Replacements for common HTML entities
    '&deg;': '°',
    '&bull;': '•',
    '&ndash;': '–',
    '&rsquo;': '’',
    '&bdquo;': '„',
    '&rdquo;': '”',
    
    # Characters to be removed
    '✔': '', '✅': '', '❓': '', '▶️': '', '⭐': '', '⚡': '', '➡': '',

    # Other common symbols & entities to remove or replace
    '&#10036;&#65039;': '', '&#10035;&#65039;': '', '&#9851;&#65039;': '',
    '&#128209;': '',      '&#8222;': '„',          '&#8221;': '”',
    '&#8216;': '‘',       '&#8217;': '’',          '&#8211;': '–',
    '&#8203;': '',        '&#9989;': '',           '&#9749;': '',
    '&#11088;': '',       '&#10003;': '',          '&#34;': '"',
    '&#39;': "'",         '&#x2013;': '–',        '&#2013;': '–',
    '&#2019;': '’',       '&nbsp;': ' ',          '&amp;': '&',
    '&lt;': '<',         '&gt;': '>',            '&quot;': '"',
    '&apos;': "'",        '&#178;': '²',          '&#8220;': '“',
    '&#8230;': '…',       '&#9679;': '•',
}
# Sort the dictionary to process longer keys first, ensuring correct replacement.
POLISH_CHAR_MAP_SORTED = dict(sorted(POLISH_CHAR_MAP.items(), key=lambda item: len(item[0]), reverse=True))

def correct_text(text):
    """Applies all defined text corrections to a single string."""
    if not isinstance(text, str) or not text:
        return text
    
    corrected_text = text
    
    # 1. Replace all known incorrect character sequences
    for wrong_str, correct_char in POLISH_CHAR_MAP_SORTED.items():
        if wrong_str in corrected_text:
            corrected_text = corrected_text.replace(wrong_str, correct_char)
            
    # 2. Remove all emoji characters using the pre-compiled regex pattern
    corrected_text = EMOJI_PATTERN.sub(r'', corrected_text)
    
    # 3. Remove any leading whitespace characters (spaces, tabs, newlines)
    corrected_text = corrected_text.lstrip()
    
    return corrected_text

# --- Processing Engines ---

def correct_excel_chars_openpyxl(filepath, progress_callback):
    """
    Standard Engine: Creates a new, clean Excel file using openpyxl.
    This process is fast and cross-platform but removes all original formatting.
    """
    source_workbook = None
    new_workbook = None
    try:
        source_workbook = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        new_workbook = openpyxl.Workbook()
        new_workbook.remove(new_workbook.active)

        total_cells = sum(sheet.max_row * sheet.max_column for sheet in source_workbook.worksheets)
        processed_cells = 0
        cells_with_changes = 0

        progress_callback(f"Silnik Standardowy: Przetwarzanie {total_cells} komórek... (0%)")
        
        for source_sheet in source_workbook.worksheets:
            new_sheet = new_workbook.create_sheet(title=source_sheet.title)
            
            for row_idx, row in enumerate(source_sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    original_value = cell.value
                    new_value = correct_text(original_value)
                    
                    new_sheet.cell(row=row_idx, column=col_idx, value=new_value)
                    
                    if new_value != original_value:
                        cells_with_changes += 1

                    processed_cells += 1
                    if total_cells > 0 and (processed_cells % (total_cells // 100 + 1) == 0):
                        progress_percent = int((processed_cells / total_cells) * 100)
                        progress_callback(f"Silnik Standardowy: Przetwarzanie... ({min(progress_percent, 100)}%)")

        directory, filename = os.path.split(filepath)
        name, ext = os.path.splitext(filename)
        new_filename = f"{name}_corrected_standard{ext}"
        new_filepath = os.path.join(directory, new_filename)

        progress_callback("Silnik Standardowy: Zapisywanie pliku...")
        new_workbook.save(new_filepath)
        return new_filepath, cells_with_changes

    finally:
        if source_workbook: source_workbook.close()
        if new_workbook: new_workbook.close()

def correct_excel_chars_pywin32(filepath, progress_callback):
    """
    Excel Engine: Uses pywin32 to automate the actual Excel application.
    This process is slower but preserves all formatting. Requires Excel on Windows.
    """
    excel = None
    workbook = None
    try:
        abs_filepath = os.path.abspath(filepath)

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(abs_filepath)
        cells_with_changes = 0

        progress_callback("Silnik Excel: Otwieranie pliku...")
        
        total_sheets = workbook.Worksheets.Count
        current_sheet_num = 0

        for sheet in workbook.Worksheets:
            current_sheet_num += 1
            progress_callback(f"Silnik Excel: Przetwarzanie arkusza {current_sheet_num}/{total_sheets} ('{sheet.Name}')...")

            used_range = sheet.UsedRange
            if not used_range.Rows.Count: continue

            cell_values = used_range.Value
            
            if used_range.Rows.Count == 1 and used_range.Columns.Count == 1:
                original_value = cell_values
                new_value = correct_text(original_value)
                if new_value != original_value:
                    cells_with_changes += 1
                used_range.Value = new_value
            else:
                corrected_values = []
                for row in cell_values:
                    new_row = []
                    # Handle cases where a row might not be a tuple (e.g., single column)
                    row_iterable = row if hasattr(row, '__iter__') else (row,)
                    for val in row_iterable:
                        original_value = val
                        new_value = correct_text(original_value)
                        if new_value != original_value:
                            cells_with_changes += 1
                        new_row.append(new_value)
                    corrected_values.append(new_row)
                used_range.Value = corrected_values

        directory, filename = os.path.split(abs_filepath)
        name, ext = os.path.splitext(filename)
        new_filename = f"{name}_corrected_excel{ext}"
        new_filepath_abs = os.path.join(directory, new_filename)

        progress_callback("Silnik Excel: Zapisywanie pliku...")
        workbook.SaveAs(new_filepath_abs)
        return new_filepath_abs, cells_with_changes

    except Exception as e:
        raise IOError(f"Błąd silnika Excel: {e}. Upewnij się, że MS Excel jest zainstalowany i plik nie jest uszkodzony lub chroniony hasłem.")
    finally:
        if workbook:
            workbook.Close(SaveChanges=False)
        if excel:
            excel.Quit()

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.title("Korektor Polskich Znaków w Excelu")
        self.geometry("600x450")
        self.resizable(False, False)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(5, weight=1)

        self.selected_file_path = None

        # --- Widgets ---
        self.file_frame = customtkinter.CTkFrame(self)
        self.file_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nsew")
        self.file_frame.grid_columnconfigure(0, weight=1)

        self.file_path_label = customtkinter.CTkLabel(self.file_frame, text="Nie wybrano pliku Excel")
        self.file_path_label.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        self.select_file_button = customtkinter.CTkButton(self.file_frame, text="Wybierz plik", command=self.select_file)
        self.select_file_button.grid(row=0, column=1, padx=10, pady=10)

        self.engine_frame = customtkinter.CTkFrame(self)
        self.engine_frame.grid(row=1, column=0, padx=20, pady=5, sticky="ew")
        self.engine_frame.grid_columnconfigure(1, weight=1)
        
        customtkinter.CTkLabel(self.engine_frame, text="Silnik przetwarzania:").grid(row=0, column=0, padx=10, pady=10)
        
        engine_options = ["Standardowy (Szybki)", "Excel (Zachowuje formatowanie)"]
        self.engine_selector = customtkinter.CTkSegmentedButton(self.engine_frame, values=engine_options, command=self.on_engine_change)
        self.engine_selector.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
        
        if not PYWIN32_AVAILABLE:
            self.engine_selector.configure(state="disabled")
        self.engine_selector.set(engine_options[0])
        
        self.action_frame = customtkinter.CTkFrame(self)
        self.action_frame.grid(row=2, column=0, padx=20, pady=5, sticky="nsew")
        self.action_frame.grid_columnconfigure(0, weight=1)

        self.start_button = customtkinter.CTkButton(self.action_frame, text="Rozpocznij korekcję", command=self.start_correction, state="disabled")
        self.start_button.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        self.excel_warning_label = customtkinter.CTkLabel(self, text="", text_color="#E53935", wraplength=550, font=customtkinter.CTkFont(weight="bold"))
        self.excel_warning_label.grid(row=3, column=0, padx=20, pady=(5, 0), sticky="ew")
        
        self.info_label = customtkinter.CTkLabel(self, text="", text_color="gray", wraplength=550)
        self.info_label.grid(row=4, column=0, padx=20, pady=5, sticky="ew")

        self.status_label = customtkinter.CTkLabel(self, text="Wybierz plik i silnik, aby rozpocząć.", text_color="gray")
        self.status_label.grid(row=5, column=0, padx=20, pady=10, sticky="ew")
        
        self.on_engine_change(engine_options[0])

    def on_engine_change(self, value):
        if value == "Standardowy (Szybki)":
            self.info_label.configure(text="Silnik Standardowy: Szybki, nie wymaga Excela, ale usuwa formatowanie.", text_color="gray")
            self.excel_warning_label.configure(text="")
        else:
            if PYWIN32_AVAILABLE:
                self.info_label.configure(text="Silnik Excel: Zachowuje formatowanie. Wymaga MS Excel i działa tylko na Windows.", text_color="orange")
                self.excel_warning_label.configure(text="UWAGA: Ta operacja może zamknąć wszystkie otwarte pliki Excel. Zapisz swoją pracę przed kontynuacją!")
            else:
                self.info_label.configure(text="Silnik Excel jest niedostępny. Zainstaluj pywin32 lub upewnij się, że jesteś na Windows.", text_color="red")
                self.excel_warning_label.configure(text="")

    def select_file(self):
        filepath = filedialog.askopenfilename(title="Wybierz plik Excel", filetypes=[("Pliki Excel", "*.xlsx;*.xlsm;*.xlsb;*.xls")])
        if filepath:
            self.selected_file_path = filepath
            self.file_path_label.configure(text=f"Plik: {os.path.basename(filepath)}")
            self.start_button.configure(state="normal")
            self.update_status("Plik wybrany. Możesz rozpocząć korekcję.", "blue")
        else:
            self.selected_file_path = None
            self.file_path_label.configure(text="Nie wybrano pliku Excel")
            self.start_button.configure(state="disabled")
            self.update_status("Anulowano wybór. Wybierz plik.", "gray")

    def start_correction(self):
        if not self.selected_file_path:
            messagebox.showwarning("Brak pliku", "Najpierw wybierz plik Excel.")
            return

        selected_engine = self.engine_selector.get()
        if selected_engine == "Excel (Zachowuje formatowanie)" and PYWIN32_AVAILABLE:
            proceed = messagebox.askokcancel(
                "Potwierdzenie operacji",
                "Używasz silnika Excel. Ta operacja może spowodować zamknięcie wszystkich niezapisanych plików Excel.\n\n"
                "Upewnij się, że zapisałeś swoją pracę. Czy chcesz kontynuować?"
            )
            if not proceed:
                return

        self.set_ui_state("disabled")
        self.update_status("Rozpoczynam przetwarzanie...", "orange")
        self.after(100, self.run_correction_task)

    def run_correction_task(self):
        selected_engine = self.engine_selector.get()
        try:
            if selected_engine == "Standardowy (Szybki)":
                new_filepath, corrected_count = correct_excel_chars_openpyxl(self.selected_file_path, self.update_progress)
            elif PYWIN32_AVAILABLE:
                new_filepath, corrected_count = correct_excel_chars_pywin32(self.selected_file_path, self.update_progress)
            else:
                raise RuntimeError("Próba użycia niedostępnego silnika Excel.")
            
            success_message = (f"Korekta zakończona! Wprowadzono zmiany w {corrected_count} komórkach.\n\n"
                               f"Zapisano jako: {os.path.basename(new_filepath)}")
            self.update_status(f"Sukces! Wprowadzono zmiany w {corrected_count} komórkach.", "green")
            messagebox.showinfo("Sukces", success_message)
        except Exception as e:
            self.handle_error(e)
        finally:
            self.set_ui_state("normal")

    def handle_error(self, error):
        self.update_status(f"Wystąpił błąd: {error}", "red")
        messagebox.showerror("Błąd krytyczny", f"Wystąpił nieoczekiwany błąd:\n\n{error}")

    def set_ui_state(self, state):
        self.start_button.configure(state=state)
        self.select_file_button.configure(state=state)
        if PYWIN32_AVAILABLE:
            self.engine_selector.configure(state=state)

    def update_status(self, text, color):
        self.status_label.configure(text=text, text_color=color)
        self.update_idletasks()

    def update_progress(self, progress_text):
        self.status_label.configure(text=progress_text)
        self.update_idletasks()

if __name__ == "__main__":
    customtkinter.set_appearance_mode("System")
    customtkinter.set_default_color_theme("blue")
    app = App()
    app.mainloop()
