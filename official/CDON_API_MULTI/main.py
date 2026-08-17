"""
CDON Product Importer (POST /v2/articles/bulk) - PySide6

Wsad: plik Excel (.xlsx/.xlsm) generowany przez cdon_for_Magda.py albo CSV.
Excel jest czytany bezpośrednio - nie trzeba go konwertować do CSV.

UWAGA: ten skrypt, CDON_API_update_MULTI.py oraz cdon_for_Magda.py tworzą jedną
rodzinę - każdą zmianę logiki wsadu trzeba nanieść na wszystkie trzy.
"""

import sys
import os
import io
import csv
import json
import time
import base64
import threading
import traceback
import concurrent.futures
import re
from datetime import datetime, date, time as dtime

import requests
import openpyxl

from PySide6.QtCore import Qt, QObject, Signal
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox,
    QLabel, QPushButton, QComboBox, QCheckBox, QSlider, QProgressBar,
    QPlainTextEdit, QFileDialog, QMessageBox
)

# --- STAŁE WSPÓLNE DLA RODZINY SKRYPTÓW CDON ---

# Kraje UE (GPSR): jeśli producent jest spoza tej listy, CDON wymaga responsible_person.
EU_COUNTRY_CODES = {
    "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR", "DE", "GR",
    "HU", "IE", "IT", "LV", "LT", "LU", "MT", "NL", "PL", "PT", "RO", "SK",
    "SI", "ES", "SE"
}

INPUT_FILE_FILTER = (
    "Wsad CDON (*.xlsx *.xlsm *.csv);;"
    "Excel (*.xlsx *.xlsm);;"
    "CSV (*.csv);;"
    "Wszystkie pliki (*)"
)

ACCENT_COLOR = "#ff69b4"
ACCENT_HOVER = "#e754a8"
STOP_COLOR = "#c2188b"
STOP_HOVER = "#a31574"
APP_BG = "#fff8fc"
PANEL_BG = "#ffeaf5"
INPUT_BG = "#fff1f8"
TEXT_COLOR = "#4a2a3a"

STYLESHEET = """
QWidget { background-color: #fff8fc; color: #4a2a3a; font-family: 'Segoe UI', Arial; font-size: 13px; }
QGroupBox { background-color: #ffeaf5; border: 1px solid #ff69b4; border-radius: 8px;
            margin-top: 14px; padding: 14px 10px 10px 10px; font-weight: bold; }
QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; color: #c2188b; }
QPushButton { background-color: #ff69b4; color: white; border: none; border-radius: 6px;
              padding: 8px 14px; font-weight: bold; }
QPushButton:hover { background-color: #e754a8; }
QPushButton:disabled { background-color: #f2cde1; color: #ffffff; }
QPushButton#stopButton { background-color: #c2188b; }
QPushButton#stopButton:hover { background-color: #a31574; }
QComboBox, QPlainTextEdit { background-color: #fff1f8; border: 1px solid #ff69b4;
                            border-radius: 6px; padding: 5px; color: #4a2a3a; }
QComboBox:disabled { color: #b08a9c; }
QComboBox QAbstractItemView { background-color: #fff1f8; selection-background-color: #ffd6ea;
                              color: #4a2a3a; border: 1px solid #ff69b4; }
QProgressBar { border: 1px solid #ff69b4; border-radius: 6px; text-align: center;
               background-color: #fff1f8; height: 18px; }
QProgressBar::chunk { background-color: #ff69b4; border-radius: 5px; }
QSlider::groove:horizontal { height: 6px; background: #fff1f8; border: 1px solid #ffd6ea; border-radius: 3px; }
QSlider::handle:horizontal { background: #ff69b4; width: 16px; margin: -6px 0; border-radius: 8px; }
QSlider::sub-page:horizontal { background: #ff69b4; border-radius: 3px; }
QCheckBox::indicator { width: 16px; height: 16px; border: 1px solid #ff69b4;
                       border-radius: 4px; background: #fff1f8; }
QCheckBox::indicator:checked { background: #ff69b4; }
"""


# --- WCZYTYWANIE WSADU (EXCEL / CSV) ---

def cell_to_text(value):
    """Zamienia komórkę Excela na tekst bez psucia liczb (gtin, ceny, VAT)."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if value != value or value in (float('inf'), float('-inf')):
            return ""
        if value.is_integer():
            return str(int(value))
        return repr(value)  # najkrótsza reprezentacja round-trip, bez 214.78000000000003
    if isinstance(value, int):
        return str(value)
    if isinstance(value, (datetime, date, dtime)):
        return value.isoformat()
    return str(value).strip()


def list_sheet_names(path):
    """Zwraca listę arkuszy pliku Excel (pusta lista dla CSV)."""
    if not path.lower().endswith(('.xlsx', '.xlsm')):
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def load_excel_rows(path, sheet_name=None):
    """Czyta arkusz Excela do listy słowników {nagłówek: tekst}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], ws.title

        headers = [cell_to_text(h).strip().replace('"', '').replace("'", "") for h in header_row]

        data = []
        for raw in rows_iter:
            row = {}
            has_value = False
            for idx, name in enumerate(headers):
                if not name:
                    continue
                value = cell_to_text(raw[idx]) if idx < len(raw) else ""
                if value:
                    has_value = True
                row[name] = value
            if has_value:
                data.append(row)
        return data, ws.title
    finally:
        wb.close()


def load_csv_rows(path, log=print):
    """Czyta CSV z autodetekcją kodowania i separatora."""
    content = None
    encoding_used = None
    for enc in ['utf-8-sig', 'utf-8', 'cp1250', 'latin-1']:
        try:
            with open(path, mode='r', encoding=enc) as f:
                content = f.read()
            encoding_used = enc
            break
        except UnicodeDecodeError:
            continue

    if content is None:
        raise ValueError("Nie rozpoznano kodowania pliku CSV.")

    f_io = io.StringIO(content)
    sample = f_io.read(4096)
    f_io.seek(0)

    delimiter = ';'
    try:
        if len(sample) > 5:
            delimiter = csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t', '|']).delimiter
    except Exception:
        pass

    log(f"Separator: '{delimiter}' | Kodowanie: {encoding_used}")

    reader = csv.DictReader(f_io, delimiter=delimiter)
    if reader.fieldnames:
        reader.fieldnames = [n.strip().replace('"', '').replace("'", "") for n in reader.fieldnames]

    rows = []
    for raw in reader:
        row = {}
        for k, v in raw.items():
            if not k:
                continue
            if isinstance(v, str):
                row[k.strip()] = v.strip()
            elif v is None:
                row[k.strip()] = ""
            else:
                row[k.strip()] = str(v)
        if any(row.values()):
            rows.append(row)
    return rows


def load_table(path, sheet_name=None, log=print):
    """Wspólne wejście: Excel albo CSV -> lista słowników."""
    ext = os.path.splitext(path)[1].lower()
    if ext in ('.xlsx', '.xlsm'):
        rows, used_sheet = load_excel_rows(path, sheet_name)
        log(f"Źródło: Excel | Arkusz: '{used_sheet}'")
        return rows
    if ext == '.xls':
        raise ValueError("Stary format .xls nie jest obsługiwany - zapisz plik jako .xlsx.")
    return load_csv_rows(path, log)


def resolve_config_path(filename="accounts.csv"):
    """accounts.csv obok skryptu, z zachowaniem zgodności ze starym uruchamianiem z CWD."""
    cwd_path = os.path.abspath(filename)
    if os.path.exists(cwd_path):
        return cwd_path
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)


# --- LOGIKA BIZNESOWA (API - IMPORTER POST) ---

class CDONClient:
    def __init__(self, merchant_id, api_token, use_sandbox=True, log_callback=None):
        self.merchant_id = merchant_id
        self.api_token = api_token
        self.base_url = "https://merchants-api.sandbox.cdon.com/api" if use_sandbox else "https://merchants-api.cdon.com/api"
        self.log = log_callback if log_callback else print
        self.request_count = 0
        self._count_lock = threading.Lock()

    def _wait_with_stop(self, wait_seconds, stop_event=None):
        """Czeka podaną liczbę sekund, ale przerywa natychmiast po STOP."""
        if stop_event is None:
            time.sleep(wait_seconds)
            return True

        end_time = time.time() + wait_seconds
        while time.time() < end_time:
            if stop_event.is_set():
                return False
            remaining = end_time - time.time()
            stop_event.wait(timeout=min(1.0, remaining))
        return not stop_event.is_set()

    def _get_headers(self):
        auth_string = f"{self.merchant_id}:{self.api_token}"
        auth_bytes = auth_string.encode('ascii')
        base64_bytes = base64.b64encode(auth_bytes)
        base64_auth = base64_bytes.decode('ascii')
        return {
            "Authorization": f"Basic {base64_auth}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": "CDON-Python-Importer/4.0-Qt"
        }

    def _parse_specifications(self, data, sku):
        market_languages = {
            "SE": "sv-SE",
            "DK": "da-DK",
            "FI": "fi-FI"
        }

        csv_specs = []
        for market_code, language in market_languages.items():
            group_key = f"specification_{market_code}_group"
            group_name = str(data.get(group_key, "")).strip()

            name_prefix = f"specification_{market_code}_name_"
            value_prefix = f"specification_{market_code}_value_"
            typo_value_prefix = f"specification_{market_code}_vaule_"

            indexes = set()
            for key in data.keys():
                if key.startswith(name_prefix):
                    suffix = key[len(name_prefix):]
                    if re.fullmatch(r"\d+", suffix):
                        indexes.add(int(suffix))

            attributes = []
            for idx in sorted(indexes):
                name_key = f"{name_prefix}{idx}"
                value_key = f"{value_prefix}{idx}"
                typo_value_key = f"{typo_value_prefix}{idx}"

                spec_name = str(data.get(name_key, "")).strip()
                spec_value = str(data.get(value_key, "")).strip()
                if not spec_value:
                    spec_value = str(data.get(typo_value_key, "")).strip()

                if not spec_name and not spec_value:
                    continue

                if not spec_name or not spec_value:
                    self.log(f"[WARN] {sku}: Niepełna specyfikacja {market_code} dla indeksu {idx} (name/value).")
                    continue

                attributes.append({
                    "name": spec_name,
                    "value": spec_value
                })

            if attributes:
                if not group_name:
                    group_name = "Specifications"
                    self.log(f"[WARN] {sku}: Brak {group_key}. Użyto domyślnej nazwy grupy '{group_name}' dla rynku {market_code}.")

                csv_specs.append({
                    "language": language,
                    "value": [
                        {
                            "name": group_name,
                            "value": attributes
                        }
                    ]
                })

        if csv_specs:
            return csv_specs

        raw_specs = data.get('specifications_json') or data.get('specifications')
        if not raw_specs or not str(raw_specs).strip():
            return []

        try:
            parsed = json.loads(str(raw_specs).strip())
        except json.JSONDecodeError:
            preview = str(raw_specs).strip().replace("\n", " ")[:160]
            self.log(f"[WARN] {sku}: Niepoprawny JSON w specifications: {preview}")
            return []

        if isinstance(parsed, dict):
            parsed = [parsed]

        if not isinstance(parsed, list):
            self.log(f"[WARN] {sku}: specifications musi być listą lub obiektem JSON.")
            return []

        valid_specs = []
        for spec in parsed:
            if not isinstance(spec, dict):
                continue

            language = spec.get("language")
            sections = spec.get("value")
            if not language or not isinstance(sections, list):
                continue

            clean_sections = []
            for section in sections:
                if not isinstance(section, dict):
                    continue

                section_name = section.get("name")
                attributes = section.get("value")
                if not section_name or not isinstance(attributes, list):
                    continue

                clean_attributes = []
                for attr in attributes:
                    if not isinstance(attr, dict):
                        continue

                    attr_name = attr.get("name")
                    attr_value = attr.get("value")
                    if attr_name is None or attr_value is None:
                        continue

                    attr_entry = {
                        "name": str(attr_name),
                        "value": str(attr_value)
                    }

                    description = attr.get("description")
                    if description is not None and str(description).strip():
                        attr_entry["description"] = str(description)

                    clean_attributes.append(attr_entry)

                if clean_attributes:
                    clean_sections.append({
                        "name": str(section_name),
                        "value": clean_attributes
                    })

            if clean_sections:
                valid_specs.append({
                    "language": str(language),
                    "value": clean_sections
                })

        if not valid_specs:
            self.log(f"[WARN] {sku}: specifications ma niepoprawną strukturę i zostało pominięte.")

        return valid_specs

    def _parse_manufacturer(self, data, sku):
        """
        Buduje obiekt 'manufacturer' (GPSR) razem z 'responsible_person'.
        Kolumny wsadu:
          manufacturer_name, manufacturer_street_address, manufacturer_city,
          manufacturer_postal_code, manufacturer_country, manufacturer_website,
          manufacturer_email,
          responsible_person_name, responsible_person_phone, responsible_person_email
        Zwraca dict albo None (gdy brak danych lub dane niekompletne).
        """
        def pick(*keys):
            for key in keys:
                raw = data.get(key)
                if raw is not None and str(raw).strip():
                    return str(raw).strip()
            return ""

        name = pick('manufacturer_name', 'manufacturerName')
        street = pick('manufacturer_street_address', 'manufacturer_street', 'manufacturerStreetAddress')
        city = pick('manufacturer_city', 'manufacturerCity')
        postal = pick('manufacturer_postal_code', 'manufacturer_zip', 'manufacturerPostalCode')
        country = pick('manufacturer_country', 'manufacturerCountry').upper()
        website = pick('manufacturer_website', 'manufacturerWebsite')
        email = pick('manufacturer_email', 'manufacturerEmail')

        rp_name = pick('responsible_person_name', 'responsiblePersonName')
        rp_phone = pick('responsible_person_phone', 'responsiblePersonPhone')
        rp_email = pick('responsible_person_email', 'responsiblePersonEmail')

        any_data = any([name, street, city, postal, country, website, email, rp_name, rp_phone, rp_email])
        if not any_data:
            return None

        if not name:
            self.log(f"[WARN] {sku}: Pominięto manufacturer - brak 'manufacturer_name'.")
            return None

        # address jest wymagany przez API razem z name
        missing = [label for label, value in (
            ('manufacturer_street_address', street),
            ('manufacturer_city', city),
            ('manufacturer_postal_code', postal),
            ('manufacturer_country', country)
        ) if not value]
        if missing:
            self.log(f"[WARN] {sku}: Pominięto manufacturer - brak wymaganych pól adresu: {', '.join(missing)}.")
            return None

        if len(country) != 2:
            self.log(f"[WARN] {sku}: 'manufacturer_country' = '{country}' musi być 2-literowym kodem ISO. Pominięto manufacturer.")
            return None

        manufacturer = {
            "name": name,
            "address": {
                "street_address": street,
                "city": city,
                "postal_code": postal,
                "country": country
            }
        }

        if website:
            manufacturer["website"] = website
        if email:
            manufacturer["email"] = email

        if rp_name:
            responsible_person = {"name": rp_name}
            if rp_phone:
                responsible_person["phone"] = rp_phone
            if rp_email:
                responsible_person["email"] = rp_email
            manufacturer["responsible_person"] = responsible_person
        else:
            if rp_phone or rp_email:
                self.log(f"[WARN] {sku}: Pominięto responsible_person - brak 'responsible_person_name'.")
            if country not in EU_COUNTRY_CODES:
                self.log(f"[WARN] {sku}: Producent z kraju '{country}' (spoza UE) - CDON wymaga 'responsible_person'.")

        return manufacturer

    def create_product_from_flat_data(self, data, stop_event=None):
        """Wysyła jeden produkt metodą POST (tworzenie/nadpisywanie)"""
        sku = data.get('sku')
        if not sku:
            if any(data.values()):
                self.log("[SKIP] Pominięto wiersz: Brak SKU")
            return False

        market_config = {
            'Se': {'code': 'SE', 'lang': 'sv-SE', 'currency': 'SEK'},
            'Dk': {'code': 'DK', 'lang': 'da-DK', 'currency': 'DKK'},
            'Fi': {'code': 'FI', 'lang': 'fi-FI', 'currency': 'EUR'}
        }

        api_markets = []
        api_titles = []
        api_descriptions = []
        api_prices = []
        api_shipping = []
        api_delivery = []

        # --- Przetwarzanie rynków ---
        for suffix, config in market_config.items():
            price_key = f'originalPrice{suffix}'

            # Jeśli jest cena, dodajemy rynek
            if data.get(price_key) and str(data[price_key]).strip():
                code = config['code']
                lang = config['lang']
                api_markets.append(code)

                # Tytuł
                if data.get(f'title{suffix}'):
                    api_titles.append({"language": lang, "value": str(data[f'title{suffix}'])})

                # Opis
                if data.get(f'description{suffix}'):
                    api_descriptions.append({"language": lang, "value": str(data[f'description{suffix}'])})

                # Ceny i VAT
                try:
                    price_str = str(data[f'originalPrice{suffix}']).replace(',', '.').strip()
                    price_val = float(price_str)

                    vat_str = str(data.get(f'vat{suffix}', '25')).replace(',', '.').strip()
                    if not vat_str: vat_str = '25'

                    vat_rate = float(vat_str)
                    # Poprawka VAT: 25 -> 0.25
                    if vat_rate > 1: vat_rate = vat_rate / 100.0

                    api_prices.append({
                        "market": code,
                        "value": {
                            "amount_including_vat": price_val,
                            "currency": config['currency'],
                            "vat_rate": vat_rate
                        }
                    })
                except ValueError:
                    self.log(f"[WARN] {sku}: Błąd ceny dla {suffix}.")

                # Czas dostawy
                min_time = data.get(f'deliveryTimeMin{suffix}')
                max_time = data.get(f'deliveryTimeMax{suffix}')
                if min_time and max_time:
                    try:
                        api_shipping.append({"market": code, "min": int(float(min_time)), "max": int(float(max_time))})
                    except: pass

                # Typ dostawy (mapowanie)
                del_type = data.get(f'delivery{suffix}')
                if del_type:
                    raw_val = del_type.strip()
                    lookup_key = raw_val.lower().replace(" ", "").replace("_", "")
                    d_map = {'homedelivery': 'home_delivery', 'servicepoint': 'service_point', 'mailbox': 'mailbox', 'digital': 'digital'}
                    val = d_map.get(lookup_key, raw_val.lower().replace(" ", "_"))
                    api_delivery.append({"market": code, "value": val})

        # Właściwości
        properties = []
        if data.get('weight') and str(data['weight']).strip():
            properties.append({"name": "weight_kg", "value": str(data['weight']).replace(',', '.')})

        specifications = self._parse_specifications(data, sku)
        manufacturer = self._parse_manufacturer(data, sku)

        # Zdjęcia (rozdzielane średnikiem)
        extra_images = []
        if data.get('extraImages') and str(data['extraImages']).strip():
            extra_images = [img.strip() for img in data['extraImages'].split(';') if img.strip()]

        # Stan magazynowy
        stock_int = 0
        if data.get('stock') and str(data['stock']).strip():
            try:
                stock_int = int(float(data['stock']))
            except: pass

        # shipped_from (EU / NON_EU), domyślnie EU
        shipped_from = "EU"
        if data.get('shipped_from') and str(data['shipped_from']).strip():
            shipped_from_val = str(data['shipped_from']).strip().upper()
            if shipped_from_val in ['EU', 'NON_EU']:
                shipped_from = shipped_from_val
            else:
                self.log(f"[WARN] {sku}: Nieprawidłowy shipped_from '{shipped_from_val}'. Dozwolone: EU, NON_EU. Użyto EU.")

        # --- BUDOWA BODY (POST) ---
        article_payload = {
            "sku": sku,
            "status": "for sale",
            "shipped_from": shipped_from,
            "quantity": stock_int,
            "main_image": data.get('mainImage'),
            "markets": api_markets,
            "price": api_prices,
            "shipping_time": api_shipping,
            "delivery_type": api_delivery,
            "title": api_titles,
            "description": api_descriptions,
            "category": data.get('category'),
            "brand": data.get('brand'),
            "gtin": data.get('gtin')
        }

        # Dodaj zdjęcia tylko jeśli istnieją
        if extra_images:
            article_payload["images"] = extra_images
        if properties:
            article_payload["properties"] = properties
        if specifications:
            article_payload["specifications"] = specifications
        if manufacturer:
            article_payload["manufacturer"] = manufacturer

        max_attempts = 6
        retry_wait_seconds = 600

        for attempt in range(1, max_attempts + 1):
            if stop_event is not None and stop_event.is_set():
                return None

            try:
                request_time = time.time()

                # POST /v2/articles/bulk (Tworzenie/Nadpisywanie)
                response = requests.post(
                    f"{self.base_url}/v2/articles/bulk",
                    headers=self._get_headers(),
                    data=json.dumps({"articles": [article_payload]}),
                    timeout=30
                )

                response_time = time.time() - request_time
                with self._count_lock:
                    self.request_count += 1
                    request_no = self.request_count

                # Parsuj response body
                try:
                    response_body = response.json()
                except:
                    response_body = response.text

                # ZAWSZE loguj szczegóły do TXT logu
                self.log(f"\n[REQUEST #{request_no}] SKU: {sku}")
                self.log(f"  Response Time: {response_time:.3f}s")
                self.log(f"  Status Code: {response.status_code}")
                self.log(f"  Response Body: {json.dumps(response_body) if isinstance(response_body, dict) else str(response_body)[:500]}")

                if response.status_code == 429:
                    if attempt < max_attempts:
                        self.log(f"[RATE LIMIT] {sku}: 429 Too Many Requests. Próba {attempt}/{max_attempts}. Czekam 10 minut przed ponowieniem...")
                        wait_completed = self._wait_with_stop(retry_wait_seconds, stop_event)
                        if not wait_completed:
                            self.log(f"[INFO] {sku}: Przerwano oczekiwanie retry przez STOP.")
                            return None
                        continue
                    self.log(f"[RATE LIMIT] {sku}: 429 Too Many Requests po {max_attempts} próbach. Przerywam.")
                    return False

                if response.status_code in [200, 201, 202]:
                    try:
                        response_data = response.json()

                        if isinstance(response_data, dict):
                            # Odpowiedź w formacie success/failed
                            failed = response_data.get('failed')
                            if isinstance(failed, list) and failed:
                                for item in failed:
                                    errors = item.get('errors') if isinstance(item, dict) else None
                                    if isinstance(errors, list):
                                        msg = "; ".join([
                                            f"{e.get('field', e.get('location', '?'))}: {e.get('message', '?')}"
                                            for e in errors if isinstance(e, dict)
                                        ])
                                    else:
                                        msg = json.dumps(item)
                                    self.log(f"[BŁĄD API] {sku}: {msg}")
                                return False

                            # Sprawdzenie błędów
                            if response_data.get('errors') or response_data.get('message') or response_data.get('description'):
                                msg = response_data.get('message') or response_data.get('description') or ""
                                if response_data.get('errors'):
                                    msg += " " + json.dumps(response_data.get('errors'))
                                self.log(f"[BŁĄD API] {sku}: {msg}")
                                return False

                            # Sukces (Batch ID lub Receipt)
                            if response_data.get('receipt') or response_data.get('batch_id') or response_data.get('success'):
                                self.log(f"[OK] {sku}: Wysłano")
                                return True

                            # OSTRZEŻENIE: OK status ale brak batch_id/receipt
                            self.log(f"[WARN] {sku}: Status OK ({response.status_code}) ale brak potwierdzenia (receipt/batch_id)")
                            return True

                        self.log(f"[OK] {sku}: Wysłano pomyślnie.")
                        return True

                    except json.JSONDecodeError:
                        # OSTRZEŻENIE: Status OK ale non-JSON response
                        self.log(f"[WARN] {sku}: Status {response.status_code}, brak JSON")
                        return True
                else:
                    self.log(f"[API ERROR] {sku}: {response.status_code} - {response.text}")
                    return False
            except Exception as e:
                if stop_event is not None and stop_event.is_set():
                    return None
                self.log(f"[EXCEPTION] {sku}: {str(e)}")
                return False


# --- GUI (PySide6) ---

class WorkerSignals(QObject):
    """Most między wątkiem roboczym a GUI - sygnały Qt są bezpieczne między wątkami."""
    log = Signal(str)
    progress = Signal(int, int)
    status = Signal(str)
    finished = Signal(str)


class CDONImporterWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("CDON Product Importer v4.0 - Qt / Excel")
        self.resize(900, 850)
        self.setStyleSheet(STYLESHEET)

        self.file_path = ""
        self.has_sheets = False
        self.is_running = False
        self.stop_event = threading.Event()
        self.active_executor = None

        self.processed_count = 0
        self.success_count = 0

        self.accounts_data = {}
        self.config_file = resolve_config_path("accounts.csv")
        self.log_file = None
        self._log_lock = threading.Lock()

        self.signals = WorkerSignals()
        self.signals.log.connect(self._append_log)
        self.signals.progress.connect(self._set_progress)
        self.signals.status.connect(self.status_label_set)
        self.signals.finished.connect(self._on_finished)

        self._load_accounts_config()
        self._create_widgets()
        self._refresh_accounts(initial=True)

    # --- Konfiguracja kont ---

    def _load_accounts_config(self):
        if not os.path.exists(self.config_file):
            try:
                with open(self.config_file, "w", encoding="utf-8") as f:
                    f.write("Nazwa Konta;MerchantID;APIToken\n")
                    f.write("SandboxTest;12345;abc-token-przyklad\n")
            except Exception as e:
                QMessageBox.critical(self, "Błąd", f"Nie można utworzyć pliku konfiguracyjnego: {e}")

        self.accounts_data = {}
        try:
            with open(self.config_file, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    parts = line.split(';')
                    if len(parts) >= 3:
                        name = parts[0].strip()
                        m_id = parts[1].strip()
                        token = parts[2].strip()
                        if name.lower() == "nazwa konta":
                            continue
                        if name:
                            self.accounts_data[name] = {"id": m_id, "token": token}
        except Exception as e:
            QMessageBox.critical(self, "Błąd konfiguracji", f"Błąd odczytu {self.config_file}:\n{e}")

    # --- Budowa interfejsu ---

    def _create_widgets(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # --- Konto API ---
        account_box = QGroupBox("Wybór konta API")
        account_grid = QGridLayout(account_box)
        account_grid.setColumnStretch(1, 1)

        account_grid.addWidget(QLabel("Konto:"), 0, 0)
        self.account_combo = QComboBox()
        account_grid.addWidget(self.account_combo, 0, 1)

        self.refresh_btn = QPushButton("Odśwież")
        self.refresh_btn.clicked.connect(self._refresh_accounts)
        account_grid.addWidget(self.refresh_btn, 0, 2)

        self.sandbox_check = QCheckBox("Tryb Sandbox (testowy)")
        account_grid.addWidget(self.sandbox_check, 1, 0, 1, 2)

        thread_row = QHBoxLayout()
        thread_row.addStretch(1)
        thread_row.addWidget(QLabel("Liczba wątków:"))
        self.thread_slider = QSlider(Qt.Horizontal)
        self.thread_slider.setMinimum(1)
        self.thread_slider.setMaximum(20)
        self.thread_slider.setValue(5)
        self.thread_slider.setFixedWidth(180)
        thread_row.addWidget(self.thread_slider)
        self.thread_value_label = QLabel("5")
        self.thread_value_label.setFixedWidth(28)
        self.thread_slider.valueChanged.connect(lambda v: self.thread_value_label.setText(str(v)))
        thread_row.addWidget(self.thread_value_label)
        account_grid.addLayout(thread_row, 1, 2)

        layout.addWidget(account_box)

        # --- Plik wsadu ---
        file_box = QGroupBox("Dane produktowe (Excel lub CSV)")
        file_grid = QGridLayout(file_box)
        file_grid.setColumnStretch(1, 1)

        self.select_file_btn = QPushButton("Wybierz plik...")
        self.select_file_btn.clicked.connect(self.select_file)
        file_grid.addWidget(self.select_file_btn, 0, 0)

        self.file_label = QLabel("Nie wybrano pliku")
        self.file_label.setStyleSheet("color: #a3849a;")
        file_grid.addWidget(self.file_label, 0, 1, 1, 2)

        file_grid.addWidget(QLabel("Arkusz:"), 1, 0)
        self.sheet_combo = QComboBox()
        self.sheet_combo.setEnabled(False)
        self.sheet_combo.addItem("(plik nie został wybrany)")
        file_grid.addWidget(self.sheet_combo, 1, 1, 1, 2)

        layout.addWidget(file_box)

        # --- Akcje ---
        action_box = QGroupBox("Import")
        action_layout = QVBoxLayout(action_box)

        self.start_btn = QPushButton("ROZPOCZNIJ IMPORT")
        self.start_btn.setMinimumHeight(42)
        self.start_btn.clicked.connect(self.start_process)
        action_layout.addWidget(self.start_btn)

        self.stop_btn = QPushButton("ZATRZYMAJ")
        self.stop_btn.setObjectName("stopButton")
        self.stop_btn.setMinimumHeight(42)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self.stop_process)
        action_layout.addWidget(self.stop_btn)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        action_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("Gotowy")
        self.status_label.setAlignment(Qt.AlignCenter)
        action_layout.addWidget(self.status_label)

        layout.addWidget(action_box)

        # --- Logi ---
        log_box = QGroupBox("Logi operacji")
        log_layout = QVBoxLayout(log_box)
        self.log_view = QPlainTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setMaximumBlockCount(20000)
        log_layout.addWidget(self.log_view)
        layout.addWidget(log_box, 1)

    # --- Sloty GUI ---

    def _refresh_accounts(self, initial=False):
        if not initial:
            self._load_accounts_config()
        names = list(self.accounts_data.keys())
        self.account_combo.clear()
        if names:
            self.account_combo.addItems(names)
            self.account_combo.setEnabled(True)
        else:
            self.account_combo.addItem("Brak kont w accounts.csv")
            self.account_combo.setEnabled(False)

    def select_file(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik wsadu", "", INPUT_FILE_FILTER
        )
        if not filename:
            return

        self.file_path = filename
        self.file_label.setText(os.path.basename(filename))
        self.file_label.setToolTip(filename)

        self.sheet_combo.clear()
        try:
            sheets = list_sheet_names(filename)
        except Exception as e:
            sheets = []
            QMessageBox.warning(self, "Błąd odczytu", f"Nie udało się odczytać listy arkuszy:\n{e}")

        self.has_sheets = bool(sheets)
        if sheets:
            self.sheet_combo.addItems(sheets)
        else:
            self.sheet_combo.addItem("(nie dotyczy - plik CSV)")
        self.sheet_combo.setEnabled(self.has_sheets)

    def selected_sheet(self):
        if self.has_sheets:
            return self.sheet_combo.currentText()
        return None

    def log(self, message):
        """Wywoływane z wątków roboczych - zapis do pliku + sygnał do GUI."""
        if self.log_file:
            try:
                with self._log_lock:
                    with open(self.log_file, 'a', encoding='utf-8') as f:
                        f.write(message + "\n")
            except Exception:
                pass
        self.signals.log.emit(message)

    def _append_log(self, message):
        self.log_view.appendPlainText(message)

    def _set_progress(self, current, total):
        self.progress_bar.setRange(0, max(total, 1))
        self.progress_bar.setValue(current)
        self.status_label.setText(f"Postęp: {current}/{total} (OK: {self.success_count})")

    def status_label_set(self, text):
        self.status_label.setText(text)

    def start_process(self):
        name = self.account_combo.currentText()
        if not name or name not in self.accounts_data:
            QMessageBox.warning(self, "Błąd", "Wybierz poprawne konto!")
            return

        if not self.file_path:
            QMessageBox.warning(self, "Brak pliku", "Wybierz plik wsadu (Excel lub CSV)!")
            return

        if not os.path.exists(self.file_path):
            QMessageBox.warning(self, "Brak pliku", f"Plik nie istnieje:\n{self.file_path}")
            return

        creds = self.accounts_data[name]

        self.is_running = True
        self.stop_event.clear()
        self.start_btn.setEnabled(False)
        self.start_btn.setText("PRZETWARZANIE W TLE...")
        self.stop_btn.setEnabled(True)
        self.thread_slider.setEnabled(False)
        self.select_file_btn.setEnabled(False)
        self.sheet_combo.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_view.clear()

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"cdon_import_{timestamp}.log")

        threading.Thread(
            target=self.run_import,
            args=(creds["id"], creds["token"], self.thread_slider.value(),
                  self.sandbox_check.isChecked(), self.file_path, self.selected_sheet()),
            daemon=True
        ).start()

    def stop_process(self):
        if self.is_running:
            self.is_running = False
            self.stop_event.set()
            self.stop_btn.setEnabled(False)
            self.stop_btn.setText("ZATRZYMYWANIE...")
            self.status_label.setText("Zatrzymywanie...")
            if self.active_executor is not None:
                try:
                    self.active_executor.shutdown(wait=False, cancel_futures=True)
                except Exception:
                    pass
            self.log("[INFO] Wysłano STOP. Anulowano zadania oczekujące i przerwano retry-wait.")

    # --- Wątek roboczy ---

    def run_import(self, merchant_id, api_token, num_workers, use_sandbox, file_path, sheet_name):
        client = CDONClient(merchant_id, api_token, use_sandbox, self.log)

        self.log("=" * 80)
        self.log(f"--- START IMPORTU (Wątków: {num_workers}) ---")
        self.log(f"Plik: {file_path}")
        self.log(f"Log file: {self.log_file}")
        self.log("Całe response'y z API są zapisywane do TXT logu")
        self.log("=" * 80)

        try:
            prepared_data = load_table(file_path, sheet_name, self.log)
        except Exception as e:
            self.log(f"[BŁĄD PLIKU] {str(e)}")
            self.signals.finished.emit(f"Nie udało się wczytać pliku:\n{e}")
            return

        try:
            if not prepared_data:
                self.log("[BŁĄD] Plik nie zawiera żadnych wierszy z danymi.")
                self.signals.finished.emit("Plik nie zawiera żadnych wierszy z danymi.")
                return

            headers_lower = [h.lower() for h in prepared_data[0].keys()]
            if 'sku' not in headers_lower:
                self.log("[BŁĄD] Brak kolumny 'sku'.")
                self.signals.finished.emit("Wsad nie zawiera kolumny 'sku'.")
                return

            total_rows = len(prepared_data)
            self.log(f"Produktów do przetworzenia: {total_rows}")
            self.log(f"Uruchamianie puli wątków (Max: {num_workers})...")

            self.processed_count = 0
            self.success_count = 0
            failed_count = 0
            lock = threading.Lock()

            def process_item(item_data):
                if self.stop_event.is_set() or not self.is_running:
                    return None
                time.sleep(0.01)  # Mały delay dla stabilności
                return client.create_product_from_flat_data(item_data, stop_event=self.stop_event)

            executor = concurrent.futures.ThreadPoolExecutor(max_workers=num_workers)
            self.active_executor = executor
            futures = [executor.submit(process_item, item) for item in prepared_data]

            try:
                for future in concurrent.futures.as_completed(futures):
                    if self.stop_event.is_set() or not self.is_running:
                        self.log("[INFO] Anulowanie pozostałych zadań...")
                        for pending_future in futures:
                            pending_future.cancel()
                        executor.shutdown(wait=False, cancel_futures=True)
                        break

                    try:
                        result = future.result()
                    except concurrent.futures.CancelledError:
                        continue
                    except Exception as exc:
                        self.log(f"[WĄTEK ERROR] {str(exc)}")
                        result = False

                    if result is None:
                        continue

                    with lock:
                        self.processed_count += 1
                        if result:
                            self.success_count += 1
                        else:
                            failed_count += 1
                        current = self.processed_count

                    self.signals.progress.emit(current, total_rows)
            finally:
                self.active_executor = None
                try:
                    executor.shutdown(wait=False, cancel_futures=True)
                except Exception:
                    pass

            self.log("\n" + "=" * 80)
            self.log("--- KONIEC IMPORTU ---")
            self.log(f"Przetworzono: {self.processed_count}/{total_rows} | Sukcesy: {self.success_count} | Błędy: {failed_count}")
            self.log(f"Log zapisany do: {self.log_file}")
            self.log("=" * 80)

            msg = f"Przetworzono: {self.processed_count}/{total_rows}\nSukcesy: {self.success_count}\nBłędy: {failed_count}"
            msg += f"\n\nPełne response'y z API w logu:\n{self.log_file}"
            if not self.is_running:
                msg += "\n\n(Zatrzymano przez użytkownika)"
            self.signals.finished.emit(msg)

        except Exception as e:
            self.log(f"[CRITICAL] {str(e)}")
            traceback.print_exc()
            self.signals.finished.emit(f"Błąd krytyczny:\n{e}")

    def _on_finished(self, message):
        self.is_running = False
        self.start_btn.setEnabled(True)
        self.start_btn.setText("ROZPOCZNIJ IMPORT")
        self.stop_btn.setEnabled(False)
        self.stop_btn.setText("ZATRZYMAJ")
        self.thread_slider.setEnabled(True)
        self.select_file_btn.setEnabled(True)
        self.sheet_combo.setEnabled(self.has_sheets)
        self.status_label.setText("Gotowy")
        QMessageBox.information(self, "Info", message)

    def closeEvent(self, event):
        if self.is_running:
            answer = QMessageBox.question(
                self, "Import w toku",
                "Import wciąż trwa. Zatrzymać i zamknąć?",
                QMessageBox.Yes | QMessageBox.No
            )
            if answer != QMessageBox.Yes:
                event.ignore()
                return
            self.stop_process()
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = CDONImporterWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
