"""
Opisy produktow: HTML -> czysty tekst z enterami (GUI, PyQt6)

Wejscie: XLSX albo CSV z opisami produktow.
Wyjscie: ten sam uklad kolumn, ale HTML zamieniony na czysty tekst, gdzie
formatowanie niosa entery - dla platform, ktore nie przyjmuja HTML.

Autor: Patryk Libert
"""

from __future__ import annotations

import csv
import html
import os
import re
import sys
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List, Optional

from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import (
    QAbstractItemView, QApplication, QCheckBox, QComboBox, QFileDialog, QFrame,
    QGridLayout, QGroupBox, QHBoxLayout, QHeaderView, QLabel, QLineEdit,
    QMainWindow, QMessageBox, QProgressBar, QPushButton, QScrollArea, QSizePolicy,
    QSpinBox, QSplitter, QTableWidget, QTableWidgetItem, QTextEdit, QVBoxLayout,
    QWidget,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAS_OPENPYXL = False


csv.field_size_limit(10_000_000)

# ----------------------------------------------------------------------------
# Konwerter HTML -> tekst
# ----------------------------------------------------------------------------

TAG_RE = re.compile(r"<\s*/?\s*[a-zA-Z][^>]*>")
ENCODED_TAG_RE = re.compile(r"&lt;\s*/?\s*[a-zA-Z]")

SKIP_TAGS = {"script", "style", "head", "noscript", "template", "svg", "iframe"}

# Bloki konczone pusta linia (akapit)
PARAGRAPH_TAGS = {
    "p", "div", "h1", "h2", "h3", "h4", "h5", "h6", "blockquote", "table",
    "ul", "ol", "dl", "section", "article", "header", "footer", "main",
    "aside", "nav", "figure", "figcaption", "address", "fieldset", "form",
    "pre", "center", "hr", "caption",
}
# Bloki konczone pojedynczym enterem
LINE_TAGS = {"li", "tr", "dt", "dd", "thead", "tbody", "tfoot", "legend", "option"}

CELL_TAGS = {"td", "th"}

# Znane nazwy tagow - do wylapania resztek, ktore po odkodowaniu encji
# (&lt;/strong&gt; wpisane omylkowo w tresci) wyladowalyby w tekscie jako tekst
KNOWN_TAGS = (
    "a|abbr|address|article|aside|b|blockquote|body|br|button|caption|center|cite|code|"
    "col|colgroup|dd|del|details|dfn|div|dl|dt|em|embed|fieldset|figcaption|figure|font|"
    "footer|form|h1|h2|h3|h4|h5|h6|head|header|hgroup|hr|html|i|iframe|img|input|ins|kbd|"
    "label|legend|li|main|mark|meta|nav|noscript|object|ol|optgroup|option|p|param|picture|"
    "pre|q|s|samp|script|section|select|small|source|span|strike|strong|style|sub|summary|"
    "sup|table|tbody|td|textarea|tfoot|th|thead|time|title|tr|track|tt|u|ul|var|video|wbr"
)
LEFTOVER_TAG_RE = re.compile(r"<\s*/?\s*(?:%s)\s*/?\s*>" % KNOWN_TAGS, re.IGNORECASE)
LEFTOVER_TAG_ATTR_RE = re.compile(
    r"<\s*/?\s*(?:%s)\s+[^<>]{0,300}?/?\s*>" % KNOWN_TAGS, re.IGNORECASE
)


@dataclass
class ConvertOptions:
    bullet: str = "• "               # znak przed <li>
    number_ol: bool = True           # numeruj <ol> jako 1. 2. 3.
    blank_between_paragraphs: bool = True
    max_newlines: int = 2            # max kolejnych enterow
    cell_sep: str = " | "            # separator komorek tabeli
    indent: str = "  "               # wciecie list zagniezdzonych
    auto_unescape: bool = True       # obsluga &lt;p&gt;
    keep_alt: bool = False           # dopisz alt z <img>
    strip_leftovers: bool = True     # wytnij resztki typu </strong > z tresci


class _HtmlToText(HTMLParser):
    """Zamienia HTML na tekst, przenoszac strukture na entery."""

    def __init__(self, opts: ConvertOptions):
        super().__init__(convert_charrefs=True)
        self.opts = opts
        self.parts: List[str] = []
        self.pending_nl = 0
        self.at_start = True
        self.skip_depth = 0
        self.pre_depth = 0
        self.list_stack: List[List] = []  # [tag, licznik]
        self.fresh_li = False             # zaraz po znaku listy, tekstu jeszcze nie ma

    # -- pomocnicze ---------------------------------------------------------
    def _write(self, text: str) -> None:
        if not text:
            return
        if self.pending_nl and not self.at_start:
            self.parts.append("\n" * self.pending_nl)
        self.pending_nl = 0
        self.parts.append(text)
        self.at_start = False
        self.fresh_li = False

    def _newline(self, n: int = 1, additive: bool = False) -> None:
        # po samym znaku listy nie lamiemy linii - tekst punktu ma zostac przy nim
        if self.at_start or self.fresh_li:
            return
        limit = max(1, self.opts.max_newlines)
        if additive:
            self.pending_nl = min(limit, self.pending_nl + n)
        else:
            self.pending_nl = min(limit, max(self.pending_nl, n))

    def _block_end_newlines(self, tag: str) -> int:
        # wewnatrz listy akapity nie robia pustych linii miedzy punktami
        if self.list_stack:
            return 1
        if tag in PARAGRAPH_TAGS and self.opts.blank_between_paragraphs:
            return 2
        return 1

    def _list_prefix(self) -> str:
        depth = max(0, len(self.list_stack) - 1)
        prefix = self.opts.indent * depth
        if not self.list_stack:
            return prefix + self.opts.bullet
        tag, counter = self.list_stack[-1]
        if tag == "ol" and self.opts.number_ol:
            self.list_stack[-1][1] = counter + 1
            return f"{prefix}{counter}. "
        return prefix + self.opts.bullet

    # -- HTMLParser ---------------------------------------------------------
    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag in SKIP_TAGS:
            self.skip_depth += 1
            return
        if self.skip_depth:
            return

        if tag == "br":
            self._newline(1, additive=True)
            return
        if tag == "img":
            if self.opts.keep_alt:
                alt = dict(attrs).get("alt") or ""
                if alt.strip():
                    self._write(alt.strip())
            return
        if tag == "pre":
            self.pre_depth += 1
            self._newline(self._block_end_newlines(tag))
            return
        if tag in ("ul", "ol"):
            self._newline(1)
            self.list_stack.append([tag, 1])
            return
        if tag == "li":
            self._newline(1)
            if self.pending_nl and not self.at_start:
                self.parts.append("\n" * self.pending_nl)
                self.pending_nl = 0
            prefix = self._list_prefix()
            if prefix:
                self.parts.append(prefix)
                self.at_start = False
            self.fresh_li = True
            return
        if tag in CELL_TAGS:
            if not self.at_start and not self.pending_nl and not self.fresh_li:
                self._write(self.opts.cell_sep)
            return
        if tag == "hr":
            self._newline(1)
            self._write("---")
            self._newline(1)
            return
        if tag in PARAGRAPH_TAGS or tag in LINE_TAGS:
            self._newline(self._block_end_newlines(tag))

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in SKIP_TAGS:
            self.skip_depth = max(0, self.skip_depth - 1)
            return
        if self.skip_depth:
            return
        if tag == "pre":
            self.pre_depth = max(0, self.pre_depth - 1)
            self._newline(self._block_end_newlines(tag))
            return
        if tag in ("ul", "ol"):
            if self.list_stack:
                self.list_stack.pop()
            # lista zagniezdzona w <li> nie konczy akapitu
            self._newline(1 if self.list_stack else self._block_end_newlines(tag))
            return
        if tag in CELL_TAGS:
            return
        if tag == "li":
            self.fresh_li = False
            self._newline(1)
            return
        if tag in PARAGRAPH_TAGS or tag in LINE_TAGS:
            self._newline(self._block_end_newlines(tag))

    def handle_data(self, data: str) -> None:
        if self.skip_depth or not data:
            return
        if self.pre_depth:
            self._write(data.replace("\r\n", "\n").replace("\r", "\n"))
            return
        text = data.replace(" ", " ")
        text = re.sub(r"\s+", " ", text)
        if not text.strip():
            # spacja tylko miedzy slowami w tej samej linii
            if self.parts and not self.pending_nl and not self.at_start:
                if not self.parts[-1].endswith((" ", "\n")):
                    self.parts.append(" ")
            return
        self._write(text)

    def get_text(self) -> str:
        return "".join(self.parts)


def _clean_line(line: str) -> str:
    """Czysci linie, zachowujac wciecie list zagniezdzonych."""
    m = re.match(r"^([ \t]*)(.*)$", line)
    indent, body = m.group(1), m.group(2)
    body = re.sub(r"[ \t]{2,}", " ", body).strip()
    return (indent + body).rstrip() if body else ""


def _tidy(text: str, max_newlines: int) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace(" ", " ")
    text = "\n".join(_clean_line(ln) for ln in text.split("\n"))
    limit = max(1, max_newlines)
    text = re.sub(r"\n{%d,}" % (limit + 1), "\n" * limit, text)
    return text.strip()


def cell_to_text(value: Any) -> str:
    """Wartosc komorki -> tekst (bez '.0' na koncu liczb calkowitych)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def html_to_text(value: Any, opts: ConvertOptions) -> str:
    """HTML (lub czysty tekst) -> tekst, w ktorym formatowanie niosa entery."""
    text = cell_to_text(value)
    if not text.strip():
        return ""

    has_tags = bool(TAG_RE.search(text))
    if not has_tags and opts.auto_unescape and ENCODED_TAG_RE.search(text):
        text = html.unescape(text)
        has_tags = bool(TAG_RE.search(text))

    if not has_tags:
        # brak HTML - tylko encje i porzadek, istniejace entery zostaja
        return _tidy(_strip_leftovers(html.unescape(text), opts), opts.max_newlines)

    parser = _HtmlToText(opts)
    try:
        parser.feed(text)
        parser.close()
    except Exception:
        pass
    out = parser.get_text()
    if not out.strip():
        out = html.unescape(TAG_RE.sub(" ", text))
    return _tidy(_strip_leftovers(out, opts), opts.max_newlines)


def _strip_leftovers(text: str, opts: ConvertOptions) -> str:
    """Usuwa tagi, ktore byly w tresci zaescape'owane (&lt;/strong&gt;)."""
    if not opts.strip_leftovers or "<" not in text:
        return text
    text = LEFTOVER_TAG_RE.sub("", text)
    return LEFTOVER_TAG_ATTR_RE.sub("", text)


def looks_like_html(value: Any) -> bool:
    s = cell_to_text(value)
    if not s:
        return False
    return bool(TAG_RE.search(s)) or bool(ENCODED_TAG_RE.search(s)) or "&nbsp;" in s


# ----------------------------------------------------------------------------
# Wczytywanie danych (XLSX / CSV)
# ----------------------------------------------------------------------------

ENCODINGS = ("utf-8-sig", "utf-8", "cp1250", "iso-8859-2", "latin-1")


@dataclass
class LoadedTable:
    path: Path
    headers: List[str]
    rows: List[Dict[str, Any]]
    kind: str = "csv"                      # csv | xlsx
    encoding: str = ""
    delimiter: str = ","
    sheet: str = ""
    sheets: List[str] = field(default_factory=list)
    header_row: int = 1
    html_counts: Dict[str, int] = field(default_factory=dict)


def _count_html(headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, int]:
    counts = {h: 0 for h in headers}
    for row in rows[:50000]:
        for h in headers:
            if looks_like_html(row.get(h)):
                counts[h] += 1
    return counts


def list_sheets(path: Path) -> List[str]:
    if not HAS_OPENPYXL:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_xlsx(path: Path, sheet: Optional[str] = None, header_row: int = 1) -> LoadedTable:
    if not HAS_OPENPYXL:
        raise RuntimeError("Brak openpyxl. Zainstaluj: pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = list(wb.sheetnames)
        name = sheet if sheet in sheets else sheets[0]
        ws = wb[name]
        headers: List[str] = []
        rows: List[Dict[str, Any]] = []
        for idx, raw in enumerate(ws.iter_rows(values_only=True), 1):
            if idx < header_row:
                continue
            if idx == header_row:
                seen: Dict[str, int] = {}
                for j, v in enumerate(raw, 1):
                    name_h = cell_to_text(v).strip() or f"kolumna_{get_column_letter(j)}"
                    if name_h in seen:
                        seen[name_h] += 1
                        name_h = f"{name_h}_{seen[name_h]}"
                    else:
                        seen[name_h] = 1
                    headers.append(name_h)
                continue
            if raw is None or all(v is None for v in raw):
                continue
            row: Dict[str, Any] = {}
            for j, h in enumerate(headers):
                v = raw[j] if j < len(raw) else None
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                row[h] = v
            rows.append(row)
    finally:
        wb.close()

    if not headers:
        raise ValueError(f"Arkusz nie ma naglowka w wierszu {header_row}.")
    return LoadedTable(
        path=path, headers=headers, rows=rows, kind="xlsx", sheet=name,
        sheets=sheets, header_row=header_row, html_counts=_count_html(headers, rows),
    )


def _sniff_delimiter(raw: str) -> str:
    """Separator liczony na wierszu naglowka, potem sprawdzony na danych.

    csv.Sniffer myli sie na opisach produktow - w tresci jest wiecej przecinkow
    niz srednikow w naglowku, wiec wybieral przecinek dla pliku ';'.
    """
    header = raw.split("\n", 1)[0]
    candidates = [d for d in (";", ",", "\t", "|") if header.count(d) > 0]
    if not candidates:
        return ","
    candidates.sort(key=lambda d: -header.count(d))
    lines = raw.splitlines(True)[:500]
    best, best_score = candidates[0], -1.0
    for d in candidates:
        try:
            rows = [r for r in csv.reader(lines, delimiter=d)][:200]
        except csv.Error:
            continue
        if len(rows) < 2 or len(rows[0]) < 2:
            continue
        expected = len(rows[0])
        score = sum(1 for r in rows if len(r) == expected) / len(rows)
        if score > best_score:
            best, best_score = d, score
    return best


def read_csv(path: Path) -> LoadedTable:
    raw: Optional[str] = None
    used_enc = "utf-8"
    for enc in ENCODINGS:
        try:
            with open(path, "r", encoding=enc, newline="") as fh:
                raw = fh.read()
            used_enc = enc
            break
        except UnicodeDecodeError:
            continue
    if raw is None:
        raise IOError("Nie udalo sie odczytac pliku w zadnym kodowaniu.")

    delimiter = _sniff_delimiter(raw)
    reader = csv.DictReader(raw.splitlines(True), delimiter=delimiter)
    headers = [h for h in (reader.fieldnames or []) if h is not None]
    rows: List[Dict[str, Any]] = [{h: (r.get(h) or "") for h in headers} for r in reader]
    return LoadedTable(
        path=path, headers=headers, rows=rows, kind="csv", encoding=used_enc,
        delimiter=delimiter, html_counts=_count_html(headers, rows),
    )


def load_table(path: Path, sheet: Optional[str] = None, header_row: int = 1) -> LoadedTable:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return read_xlsx(path, sheet, header_row)
    return read_csv(path)


# ----------------------------------------------------------------------------
# Worker
# ----------------------------------------------------------------------------

@dataclass
class OutputOptions:
    out_path: Path = Path()
    columns: List[str] = field(default_factory=list)
    in_place: bool = True            # nadpisz kolumny (False -> dodaj *_plain)
    suffix: str = "_plain"
    delimiter: str = ","
    encoding: str = "utf-8-sig"
    newline_mode: str = "real"       # real | \n | space
    quote_all: bool = False
    skip_empty_rows: bool = False


class ConvertWorker(QThread):
    progress = pyqtSignal(int, int)
    log = pyqtSignal(str)
    done = pyqtSignal(dict)
    failed = pyqtSignal(str)

    def __init__(self, data: LoadedTable, copts: ConvertOptions, oopts: OutputOptions):
        super().__init__()
        self.data = data
        self.copts = copts
        self.oopts = oopts
        self._stop = False

    def stop(self) -> None:
        self._stop = True

    def _encode_newlines(self, text: str) -> str:
        mode = self.oopts.newline_mode
        if mode == "\\n":
            return text.replace("\n", "\\n")
        if mode == "space":
            return re.sub(r"\n+", " ", text)
        return text

    def run(self) -> None:
        try:
            cols = [c for c in self.oopts.columns if c in self.data.headers]
            if not cols:
                self.failed.emit("Nie wybrano zadnej kolumny do konwersji.")
                return

            headers = list(self.data.headers)
            if not self.oopts.in_place:
                for c in cols:
                    new_name = c + self.oopts.suffix
                    if new_name not in headers:
                        headers.insert(headers.index(c) + 1, new_name)

            targets = cols if self.oopts.in_place else [c + self.oopts.suffix for c in cols]
            total = len(self.data.rows)
            out_rows: List[Dict[str, Any]] = []
            changed = emptied = skipped = 0
            chars_before = chars_after = 0

            for i, row in enumerate(self.data.rows, 1):
                if self._stop:
                    self.failed.emit("Przerwano przez uzytkownika.")
                    return
                new_row = dict(row)
                row_changed = False
                row_all_empty = True
                for c in cols:
                    original = cell_to_text(row.get(c))
                    converted = html_to_text(original, self.copts)
                    chars_before += len(original)
                    chars_after += len(converted)
                    if converted.strip():
                        row_all_empty = False
                    elif original.strip():
                        emptied += 1
                    stored = self._encode_newlines(converted)
                    new_row[c if self.oopts.in_place else c + self.oopts.suffix] = stored
                    if stored != original:
                        row_changed = True
                if self.oopts.skip_empty_rows and row_all_empty:
                    skipped += 1
                else:
                    out_rows.append(new_row)
                if row_changed:
                    changed += 1
                if i % 100 == 0 or i == total:
                    self.progress.emit(i, total)

            self._write(headers, out_rows, targets)
            self.done.emit({
                "rows": len(out_rows),
                "changed": changed,
                "emptied": emptied,
                "skipped": skipped,
                "chars_before": chars_before,
                "chars_after": chars_after,
                "path": str(self.oopts.out_path),
            })
        except Exception as exc:  # pragma: no cover
            self.failed.emit(f"{type(exc).__name__}: {exc}")

    # -- zapis -------------------------------------------------------------
    def _write(self, headers: List[str], rows: List[Dict[str, Any]], targets: List[str]) -> None:
        out = self.oopts.out_path
        out.parent.mkdir(parents=True, exist_ok=True)
        if out.suffix.lower() in (".xlsx", ".xlsm"):
            self._write_xlsx(headers, rows, targets)
        else:
            self._write_csv(headers, rows)

    def _write_xlsx(self, headers: List[str], rows: List[Dict[str, Any]], targets: List[str]) -> None:
        if not HAS_OPENPYXL:
            raise RuntimeError("Brak openpyxl - zapisz jako .csv albo: pip install openpyxl")
        out = self.oopts.out_path
        wb = Workbook()
        ws = wb.active
        ws.title = (self.data.sheet or "opisy")[:31]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([r.get(h) for h in headers])

        wrap = Alignment(wrap_text=True, vertical="top")
        top = Alignment(vertical="top")
        target_idx = {headers.index(t) + 1 for t in targets if t in headers}
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = wrap if cell.column in target_idx else top
        for j, h in enumerate(headers, 1):
            letter = get_column_letter(j)
            ws.column_dimensions[letter].width = 70 if j in target_idx else min(38, max(12, len(h) + 4))
        ws.freeze_panes = "A2"
        wb.save(out)
        self.log.emit(f"Zapisano XLSX: {out}")

    def _write_csv(self, headers: List[str], rows: List[Dict[str, Any]]) -> None:
        out = self.oopts.out_path
        quoting = csv.QUOTE_ALL if self.oopts.quote_all else csv.QUOTE_MINIMAL
        with open(out, "w", encoding=self.oopts.encoding, newline="") as fh:
            writer = csv.DictWriter(
                fh, fieldnames=headers, delimiter=self.oopts.delimiter,
                quoting=quoting, lineterminator="\r\n", extrasaction="ignore",
            )
            writer.writeheader()
            for r in rows:
                writer.writerow({h: cell_to_text(r.get(h)) for h in headers})
        self.log.emit(f"Zapisano CSV: {out}")


# ----------------------------------------------------------------------------
# GUI
# ----------------------------------------------------------------------------

# Skala calego interfejsu (0.85 = 85%). Przelicza czcionki, odstepy i wymiary.
SCALE = 0.85


def px(value: float) -> int:
    """Wymiar w pikselach przeskalowany wspolczynnikiem SCALE."""
    return max(1, round(value * SCALE))


STYLE_TMPL = """
QMainWindow, QWidget { background: #16121a; color: #f2e9f2; }
QGroupBox {
    border: 1px solid #4a3352; border-radius: %(r8)dpx; margin-top: %(m14)dpx;
    padding: %(p10)dpx %(p8)dpx %(p8)dpx %(p8)dpx; font-weight: 600;
}
QGroupBox::title {
    subcontrol-origin: margin; left: %(p12)dpx; padding: 0 %(p6)dpx; color: #ff6fb5;
}
QPushButton {
    background: #6d2f5e; border: 1px solid #a04a86; border-radius: %(r6)dpx;
    padding: %(p7)dpx %(p14)dpx; color: #ffffff; font-weight: 600;
}
QPushButton:hover { background: #8c3c78; }
QPushButton:disabled { background: #33293a; color: #7b6f82; border-color: #3d3145; }
QPushButton#primary {
    background: #c2185b; border-color: #ff4f90; padding: %(p9)dpx %(p18)dpx;
}
QPushButton#primary:hover { background: #e0246c; }
QLineEdit, QComboBox, QSpinBox, QTextEdit, QTableWidget {
    background: #1f1926; border: 1px solid #43354d; border-radius: %(r5)dpx;
    padding: %(p5)dpx; color: #f2e9f2; selection-background-color: #c2185b;
}
QTableWidget { gridline-color: #33293a; }
QHeaderView::section {
    background: #2a2033; color: #ffb3d9; border: 0; border-bottom: 1px solid #43354d;
    padding: %(p5)dpx; font-weight: 600;
}
QProgressBar {
    border: 1px solid #43354d; border-radius: %(r6)dpx; text-align: center;
    background: #1f1926; height: %(h20)dpx;
}
QProgressBar::chunk { background: #c2185b; border-radius: %(r5)dpx; }
QCheckBox::indicator { width: %(cb)dpx; height: %(cb)dpx; }
QLabel#hint { color: #b39bbd; }
QSplitter::handle { background: #43354d; }
QScrollArea { border: 0; }
QScrollBar:vertical { background: #1f1926; width: %(sb)dpx; margin: 0; }
QScrollBar:horizontal { background: #1f1926; height: %(sb)dpx; margin: 0; }
QScrollBar::handle {
    background: #5c3b57; border-radius: %(r5)dpx; min-height: %(sbmin)dpx;
    min-width: %(sbmin)dpx;
}
QScrollBar::handle:hover { background: #8c3c78; }
QScrollBar::add-line, QScrollBar::sub-line { height: 0; width: 0; }
QScrollBar::add-page, QScrollBar::sub-page { background: transparent; }
"""

STYLE = STYLE_TMPL % {
    "r5": px(5), "r6": px(6), "r8": px(8), "m14": px(14),
    "p5": px(5), "p6": px(6), "p7": px(7), "p8": px(8), "p9": px(9),
    "p10": px(10), "p12": px(12), "p14": px(14), "p18": px(18),
    "h20": px(20), "cb": px(15), "sb": px(13), "sbmin": px(30),
}

INPUT_FILTER = "Arkusze i CSV (*.xlsx *.xlsm *.csv);;Excel (*.xlsx *.xlsm);;CSV (*.csv);;Wszystkie (*.*)"


def _free_width(label: QLabel) -> None:
    """Etykieta z dluga trescia nie moze rozpychac layoutu ani go zawijac."""
    label.setWordWrap(False)
    label.setSizePolicy(QSizePolicy.Policy.Ignored, QSizePolicy.Policy.Preferred)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Opisy: HTML → czysty tekst z enterami")
        self.setMinimumSize(px(560), px(360))
        self.setAcceptDrops(True)
        self.data: Optional[LoadedTable] = None
        self.worker: Optional[ConvertWorker] = None
        self._loading = False
        self._build()
        self._size_to_content()

    def _size_to_content(self) -> None:
        """Startowy rozmiar: naturalna wysokosc zawartosci, ale w granicach ekranu."""
        hint = self.centralWidget().widget().sizeHint()
        width, height = px(1240), hint.height() + px(24)
        screen = QApplication.primaryScreen()
        avail = screen.availableGeometry() if screen else None
        if avail:
            # availableGeometry juz pomija pasek zadan - rezerwa tylko na ramke okna
            width = min(width, avail.width() - px(40))
            height = min(height, avail.height() - px(44))
        self.resize(max(width, self.minimumWidth()), max(height, self.minimumHeight()))
        if avail:
            geo = self.frameGeometry()
            geo.moveCenter(avail.center())
            self.move(max(avail.left(), geo.left()), max(avail.top(), geo.top()))

    # -- budowa UI ----------------------------------------------------------
    def _build(self) -> None:
        # cala zawartosc siedzi w obszarze przewijania - przy malym oknie
        # pojawiaja sie suwaki zamiast sciskania kontrolek
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.setCentralWidget(scroll)

        root = QWidget()
        root.setMinimumWidth(px(880))
        scroll.setWidget(root)
        layout = QVBoxLayout(root)
        layout.setSpacing(px(8))
        layout.setContentsMargins(px(10), px(10), px(10), px(10))

        files = QGroupBox("Pliki")
        fl = QGridLayout(files)
        self.in_edit = QLineEdit()
        self.in_edit.setPlaceholderText("XLSX lub CSV z opisami (mozesz przeciagnac plik na okno)")
        btn_in = QPushButton("Wybierz plik...")
        btn_in.clicked.connect(self.pick_input)
        fl.addWidget(QLabel("Wejscie:"), 0, 0)
        fl.addWidget(self.in_edit, 0, 1)
        fl.addWidget(btn_in, 0, 2)

        sheet_row = QHBoxLayout()
        sheet_row.addWidget(QLabel("Arkusz:"))
        self.sheet_cmb = QComboBox()
        self.sheet_cmb.setMinimumWidth(px(180))
        self.sheet_cmb.currentIndexChanged.connect(self._reload_same_file)
        sheet_row.addWidget(self.sheet_cmb)
        sheet_row.addSpacing(px(16))
        sheet_row.addWidget(QLabel("Wiersz naglowka:"))
        self.header_spin = QSpinBox()
        self.header_spin.setRange(1, 50)
        self.header_spin.setValue(1)
        self.header_spin.valueChanged.connect(self._reload_same_file)
        sheet_row.addWidget(self.header_spin)
        sheet_row.addStretch(1)
        fl.addLayout(sheet_row, 1, 1, 1, 2)

        self.out_edit = QLineEdit()
        self.out_edit.setPlaceholderText("Plik wyjsciowy (.xlsx lub .csv)")
        btn_out = QPushButton("Zapisz jako...")
        btn_out.clicked.connect(self.pick_output)
        fl.addWidget(QLabel("Wyjscie:"), 2, 0)
        fl.addWidget(self.out_edit, 2, 1)
        fl.addWidget(btn_out, 2, 2)
        self.info_lbl = QLabel("Brak pliku.")
        self.info_lbl.setObjectName("hint")
        # bez zawijania i bez wplywu na szerokosc: zawijana etykieta wlacza w layoucie
        # height-for-width, a wtedy QScrollArea trzyma wysokosc preferowana zamiast
        # minimalnej i panele nie moglyby sie skurczyc
        _free_width(self.info_lbl)
        fl.addWidget(self.info_lbl, 3, 0, 1, 3)
        layout.addWidget(files)

        mid = QSplitter(Qt.Orientation.Horizontal)

        cols_box = QGroupBox("Kolumny do konwersji")
        cl = QVBoxLayout(cols_box)
        self.cols_table = QTableWidget(0, 3)
        self.cols_table.setHorizontalHeaderLabels(["Konwertuj", "Kolumna", "Wierszy z HTML"])
        self.cols_table.verticalHeader().setVisible(False)
        self.cols_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.cols_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.cols_table.setMinimumHeight(px(190))
        hh = self.cols_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.cols_table.itemSelectionChanged.connect(self.refresh_preview)
        cl.addWidget(self.cols_table)
        brow = QHBoxLayout()
        for text, fn in (
            ("Zaznacz wszystkie", lambda: self._set_all_cols(True)),
            ("Odznacz wszystkie", lambda: self._set_all_cols(False)),
            ("Auto (tylko z HTML)", self._auto_select_cols),
        ):
            b = QPushButton(text)
            b.clicked.connect(fn)
            brow.addWidget(b)
        cl.addLayout(brow)
        mid.addWidget(cols_box)

        opts_box = QGroupBox("Opcje konwersji")
        ol = QGridLayout(opts_box)
        ol.setVerticalSpacing(px(6))
        r = 0

        self.bullet_cmb = QComboBox()
        self.bullet_cmb.setEditable(True)
        self.bullet_cmb.addItems(["• ", "- ", "* ", "– ", "(bez znaku)"])
        self.bullet_cmb.currentTextChanged.connect(self.refresh_preview)
        self.max_nl_spin = QSpinBox()
        self.max_nl_spin.setRange(1, 5)
        self.max_nl_spin.setValue(2)
        self.max_nl_spin.valueChanged.connect(self.refresh_preview)
        self.cell_sep_edit = QLineEdit(" | ")
        self.cell_sep_edit.textChanged.connect(self.refresh_preview)
        self.nl_mode_cmb = QComboBox()
        self.nl_mode_cmb.addItems([
            "prawdziwe entery (Alt+Enter w Excelu)",
            "literalne \\n",
            "bez enterow (spacje)",
        ])
        self.out_delim_cmb = QComboBox()
        self.out_delim_cmb.addItems([", (przecinek)", "; (srednik)", "TAB"])
        self.out_enc_cmb = QComboBox()
        self.out_enc_cmb.addItems(["utf-8-sig (Excel)", "utf-8", "cp1250"])
        self.csv_lbl = QLabel("Separator CSV wyjscia:")
        self.enc_lbl = QLabel("Kodowanie CSV:")

        for label, widget in (
            (QLabel("Znak listy (<li>):"), self.bullet_cmb),
            (QLabel("Max kolejnych enterow:"), self.max_nl_spin),
            (QLabel("Separator komorek tabeli:"), self.cell_sep_edit),
            (QLabel("Zapis enterow:"), self.nl_mode_cmb),
            (self.csv_lbl, self.out_delim_cmb),
            (self.enc_lbl, self.out_enc_cmb),
        ):
            ol.addWidget(label, r, 0)
            ol.addWidget(widget, r, 1)
            r += 1

        self.number_ol_chk = QCheckBox("Numeruj listy <ol> (1. 2. 3.)")
        self.number_ol_chk.setChecked(True)
        self.blank_chk = QCheckBox("Pusta linia miedzy akapitami")
        self.blank_chk.setChecked(True)
        # bez znaku & w tekscie - QCheckBox robi z niego skrot klawiszowy
        self.unescape_chk = QCheckBox("Odkoduj HTML zapisany encjami")
        self.unescape_chk.setChecked(True)
        self.leftover_chk = QCheckBox("Wytnij resztki tagow (</strong >)")
        self.leftover_chk.setChecked(True)
        self.alt_chk = QCheckBox("Dopisz tekst alt z <img>")
        self.in_place_chk = QCheckBox("Nadpisz kolumny (odznacz = *_plain)")
        self.in_place_chk.setChecked(True)
        self.quote_all_chk = QCheckBox("Cytuj wszystkie pola (CSV)")
        self.skip_empty_chk = QCheckBox("Pomin wiersze z pustym wynikiem")

        # przelaczniki po dwa w wierszu - panel jest przez to o ~100 px nizszy
        toggles = (
            self.number_ol_chk, self.blank_chk,
            self.unescape_chk, self.leftover_chk,
            self.alt_chk, self.in_place_chk,
            self.quote_all_chk, self.skip_empty_chk,
        )
        for i, chk in enumerate(toggles):
            if chk is not self.in_place_chk:
                chk.stateChanged.connect(self.refresh_preview)
            ol.addWidget(chk, r + i // 2, i % 2)
        r += (len(toggles) + 1) // 2
        ol.setRowStretch(r, 1)
        mid.addWidget(opts_box)
        mid.setSizes([px(700), px(520)])
        layout.addWidget(mid, 1)

        prev_box = QGroupBox("Podglad")
        pl = QVBoxLayout(prev_box)
        prow = QHBoxLayout()
        prow.addWidget(QLabel("Wiersz:"))
        self.row_spin = QSpinBox()
        self.row_spin.setRange(1, 1)
        self.row_spin.valueChanged.connect(self.refresh_preview)
        prow.addWidget(self.row_spin)
        b_next = QPushButton("Nastepny z HTML")
        b_next.clicked.connect(self.next_html_row)
        prow.addWidget(b_next)
        self.prev_info = QLabel("")
        self.prev_info.setObjectName("hint")
        _free_width(self.prev_info)
        prow.addWidget(self.prev_info)
        prow.addStretch(1)
        pl.addLayout(prow)
        split = QSplitter(Qt.Orientation.Horizontal)
        mono = QFont("Consolas")
        mono.setPointSizeF(9 * SCALE)
        wrap_b, self.before_txt = self._make_view("PRZED (HTML)", mono)
        wrap_a, self.after_txt = self._make_view("PO (tekst z enterami)", mono)
        split.addWidget(wrap_b)
        split.addWidget(wrap_a)
        split.setSizes([px(600), px(600)])
        pl.addWidget(split)
        layout.addWidget(prev_box, 1)

        bottom = QHBoxLayout()
        self.progress = QProgressBar()
        bottom.addWidget(self.progress, 1)
        self.run_btn = QPushButton("Konwertuj i zapisz")
        self.run_btn.setObjectName("primary")
        self.run_btn.clicked.connect(self.start_convert)
        self.run_btn.setEnabled(False)
        bottom.addWidget(self.run_btn)
        layout.addLayout(bottom)

        self.status_lbl = QLabel("")
        self.status_lbl.setObjectName("hint")
        _free_width(self.status_lbl)
        layout.addWidget(self.status_lbl)

        # etykiety pokazuja nazwy tagow (<li>, <img>) - bez tego Qt bierze je
        # za rich text i zjada
        for lbl in root.findChildren(QLabel):
            lbl.setTextFormat(Qt.TextFormat.PlainText)

    def _set_status(self, text: str) -> None:
        """Status w pasku; pelna tresc w podpowiedzi, bo etykieta moze byc przycieta."""
        self.status_lbl.setText(text)
        self.status_lbl.setToolTip(text)

    def _set_info(self, text: str) -> None:
        self.info_lbl.setText(text)
        self.info_lbl.setToolTip(text)

    @staticmethod
    def _make_view(title: str, font: QFont):
        wrap = QWidget()
        v = QVBoxLayout(wrap)
        v.setContentsMargins(0, 0, 0, 0)
        lbl = QLabel(title)
        lbl.setObjectName("hint")
        txt = QTextEdit()
        txt.setReadOnly(True)
        txt.setFont(font)
        txt.setMinimumHeight(px(92))
        v.addWidget(lbl)
        v.addWidget(txt)
        return wrap, txt

    # -- drag & drop --------------------------------------------------------
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            p = Path(url.toLocalFile())
            if p.suffix.lower() in (".xlsx", ".xlsm", ".csv"):
                self.load_file(p)
                break

    # -- opcje -------------------------------------------------------------
    def convert_opts(self) -> ConvertOptions:
        bullet = self.bullet_cmb.currentText()
        if bullet == "(bez znaku)":
            bullet = ""
        return ConvertOptions(
            bullet=bullet,
            number_ol=self.number_ol_chk.isChecked(),
            blank_between_paragraphs=self.blank_chk.isChecked(),
            max_newlines=self.max_nl_spin.value(),
            cell_sep=self.cell_sep_edit.text() or " | ",
            auto_unescape=self.unescape_chk.isChecked(),
            keep_alt=self.alt_chk.isChecked(),
            strip_leftovers=self.leftover_chk.isChecked(),
        )

    def output_opts(self) -> OutputOptions:
        delim = {0: ",", 1: ";", 2: "\t"}[self.out_delim_cmb.currentIndex()]
        enc = {0: "utf-8-sig", 1: "utf-8", 2: "cp1250"}[self.out_enc_cmb.currentIndex()]
        nl = {0: "real", 1: "\\n", 2: "space"}[self.nl_mode_cmb.currentIndex()]
        return OutputOptions(
            out_path=Path(self.out_edit.text().strip()),
            columns=self.selected_columns(),
            in_place=self.in_place_chk.isChecked(),
            delimiter=delim,
            encoding=enc,
            newline_mode=nl,
            quote_all=self.quote_all_chk.isChecked(),
            skip_empty_rows=self.skip_empty_chk.isChecked(),
        )

    def selected_columns(self) -> List[str]:
        cols = []
        for i in range(self.cols_table.rowCount()):
            chk = self.cols_table.cellWidget(i, 0)
            if isinstance(chk, QCheckBox) and chk.isChecked():
                cols.append(self.cols_table.item(i, 1).text())
        return cols

    def _set_all_cols(self, state: bool) -> None:
        for i in range(self.cols_table.rowCount()):
            chk = self.cols_table.cellWidget(i, 0)
            if isinstance(chk, QCheckBox):
                chk.setChecked(state)

    def _auto_select_cols(self) -> None:
        if not self.data:
            return
        for i in range(self.cols_table.rowCount()):
            name = self.cols_table.item(i, 1).text()
            chk = self.cols_table.cellWidget(i, 0)
            if isinstance(chk, QCheckBox):
                chk.setChecked(self.data.html_counts.get(name, 0) > 0)

    # -- pliki -------------------------------------------------------------
    def pick_input(self) -> None:
        start = str(Path(self.in_edit.text()).parent) if self.in_edit.text() else str(Path.home() / "Downloads")
        path, _ = QFileDialog.getOpenFileName(self, "Wybierz plik z opisami", start, INPUT_FILTER)
        if path:
            self.load_file(Path(path))

    def pick_output(self) -> None:
        default = self.out_edit.text() or (str(self._default_out(self.data.path)) if self.data else "")
        filt = "Excel (*.xlsx);;CSV (*.csv)" if HAS_OPENPYXL else "CSV (*.csv)"
        path, _ = QFileDialog.getSaveFileName(self, "Zapisz jako", default, filt)
        if path:
            self.out_edit.setText(path)

    @staticmethod
    def _default_out(path: Path) -> Path:
        suffix = path.suffix.lower() if path.suffix.lower() in (".xlsx", ".xlsm", ".csv") else ".csv"
        if suffix in (".xlsx", ".xlsm") and not HAS_OPENPYXL:
            suffix = ".csv"
        return path.with_name(f"{path.stem} bez html{suffix}")

    def _reload_same_file(self) -> None:
        if self._loading or not self.data:
            return
        self.load_file(self.data.path, keep_output=True)

    def load_file(self, path: Path, keep_output: bool = False) -> None:
        self._loading = True
        try:
            sheet = self.sheet_cmb.currentText() or None
            if self.data is None or self.data.path != path:
                sheet = None
            self._set_status("Wczytuje...")
            QApplication.processEvents()
            try:
                data = load_table(path, sheet, self.header_spin.value())
            except Exception as exc:
                QMessageBox.critical(self, "Blad", f"Nie udalo sie wczytac pliku:\n{exc}")
                self._set_status("")
                return

            self.data = data
            self.in_edit.setText(str(path))
            if not keep_output or not self.out_edit.text():
                self.out_edit.setText(str(self._default_out(path)))

            is_xlsx = data.kind == "xlsx"
            self.sheet_cmb.setEnabled(is_xlsx and len(data.sheets) > 1)
            self.sheet_cmb.blockSignals(True)
            self.sheet_cmb.clear()
            self.sheet_cmb.addItems(data.sheets or ["(CSV)"])
            if data.sheet:
                self.sheet_cmb.setCurrentText(data.sheet)
            self.sheet_cmb.blockSignals(False)
            self.header_spin.setEnabled(is_xlsx)
            for w in (self.csv_lbl, self.out_delim_cmb, self.enc_lbl, self.out_enc_cmb, self.quote_all_chk):
                w.setEnabled(True)

            html_cols = [h for h in data.headers if data.html_counts.get(h, 0) > 0]
            src = f"XLSX / arkusz {data.sheet}" if is_xlsx else f"CSV / {data.encoding} / sep {data.delimiter!r}"
            self._set_info(
                f"{src} | wiersze: {len(data.rows)} | kolumny: {len(data.headers)} | "
                f"z HTML: {', '.join(html_cols) if html_cols else 'brak'}"
            )

            self.cols_table.setRowCount(0)
            for h in data.headers:
                i = self.cols_table.rowCount()
                self.cols_table.insertRow(i)
                cnt = data.html_counts.get(h, 0)
                chk = QCheckBox()
                chk.setChecked(cnt > 0)
                chk.stateChanged.connect(self.refresh_preview)
                self.cols_table.setCellWidget(i, 0, chk)
                self.cols_table.setItem(i, 1, QTableWidgetItem(h))
                self.cols_table.setItem(i, 2, QTableWidgetItem(str(cnt) if cnt else "-"))

            if not html_cols:
                best, best_len = None, 0
                for h in data.headers:
                    total = sum(len(cell_to_text(r.get(h))) for r in data.rows[:200])
                    if total > best_len:
                        best, best_len = h, total
                for i in range(self.cols_table.rowCount()):
                    if self.cols_table.item(i, 1).text() == best:
                        self.cols_table.cellWidget(i, 0).setChecked(True)

            self.row_spin.setRange(1, max(1, len(data.rows)))
            self.row_spin.setValue(1)
            self.run_btn.setEnabled(bool(data.rows))
            self._set_status("Gotowe do konwersji.")
        finally:
            self._loading = False
        self.refresh_preview()

    # -- podglad -----------------------------------------------------------
    def _preview_column(self) -> Optional[str]:
        selected = self.selected_columns()
        sel = self.cols_table.selectedItems()
        if sel:
            item = self.cols_table.item(sel[0].row(), 1)
            if item and item.text() in selected:
                return item.text()
        return selected[0] if selected else None

    def refresh_preview(self) -> None:
        if self._loading or not self.data or not self.data.rows:
            return
        col = self._preview_column()
        if not col:
            self.before_txt.setPlainText("")
            self.after_txt.setPlainText("")
            self.prev_info.setText("")
            return
        idx = min(self.row_spin.value(), len(self.data.rows)) - 1
        original = cell_to_text(self.data.rows[idx].get(col))
        converted = html_to_text(original, self.convert_opts())
        self.before_txt.setPlainText(original)
        self.after_txt.setPlainText(converted)
        self.prev_info.setText(
            f"kolumna: {col} | {len(original)} → {len(converted)} znakow | "
            f"linii: {converted.count(chr(10)) + 1 if converted else 0}"
        )

    def next_html_row(self) -> None:
        if not self.data:
            return
        col = self._preview_column()
        if not col:
            return
        start = self.row_spin.value()
        rows = self.data.rows
        for offset in range(1, len(rows) + 1):
            i = (start - 1 + offset) % len(rows)
            if looks_like_html(rows[i].get(col)):
                self.row_spin.setValue(i + 1)
                return
        self._set_status(f"Brak HTML w kolumnie {col}.")

    # -- konwersja ---------------------------------------------------------
    def start_convert(self) -> None:
        if not self.data:
            return
        if not self.selected_columns():
            QMessageBox.warning(self, "Brak kolumn", "Zaznacz przynajmniej jedna kolumne.")
            return
        out = self.out_edit.text().strip()
        if not out:
            QMessageBox.warning(self, "Brak pliku", "Podaj plik wyjsciowy.")
            return
        out_path = Path(out)
        if out_path.suffix.lower() not in (".xlsx", ".xlsm", ".csv"):
            QMessageBox.warning(self, "Rozszerzenie", "Plik wyjsciowy musi byc .xlsx albo .csv")
            return
        try:
            same = out_path.resolve() == self.data.path.resolve()
        except OSError:
            same = False
        if same:
            QMessageBox.warning(self, "Ten sam plik", "Plik wyjsciowy nie moze byc plikiem wejsciowym.")
            return
        if out_path.exists():
            ans = QMessageBox.question(self, "Nadpisac?", f"Plik istnieje:\n{out_path}\n\nNadpisac?")
            if ans != QMessageBox.StandardButton.Yes:
                return

        self.run_btn.setEnabled(False)
        self.progress.setValue(0)
        self._set_status("Konwertuje...")
        self.worker = ConvertWorker(self.data, self.convert_opts(), self.output_opts())
        self.worker.progress.connect(self._on_progress)
        self.worker.log.connect(self._set_status)
        self.worker.done.connect(self._on_done)
        self.worker.failed.connect(self._on_failed)
        self.worker.start()

    def _on_progress(self, cur: int, total: int) -> None:
        self.progress.setMaximum(max(1, total))
        self.progress.setValue(cur)

    def _on_done(self, stats: dict) -> None:
        self.run_btn.setEnabled(True)
        before, after = stats["chars_before"], stats["chars_after"]
        pct = f" ({100 * after // before}% objetosci)" if before else ""
        msg = (
            f"Zapisano {stats['rows']} wierszy\n"
            f"Zmienionych: {stats['changed']}\n"
            f"Znakow: {before} → {after}{pct}\n"
            f"Pol wyczyszczonych do pustych: {stats['emptied']}\n"
        )
        if stats.get("skipped"):
            msg += f"Pominietych wierszy: {stats['skipped']}\n"
        msg += f"\nPlik: {stats['path']}"
        self._set_status(f"Gotowe: {stats['path']}")
        box = QMessageBox(self)
        box.setWindowTitle("Gotowe")
        box.setText(msg)
        open_btn = box.addButton("Otworz folder", QMessageBox.ButtonRole.ActionRole)
        box.addButton("OK", QMessageBox.ButtonRole.AcceptRole)
        box.exec()
        if box.clickedButton() is open_btn:
            os.startfile(str(Path(stats["path"]).parent))

    def _on_failed(self, err: str) -> None:
        self.run_btn.setEnabled(True)
        self._set_status("Blad.")
        QMessageBox.critical(self, "Blad", err)


def apply_scale(app: QApplication) -> None:
    """Skaluje czcionke i arkusz stylow wspolczynnikiem SCALE."""
    font = app.font()
    base = font.pointSizeF() if font.pointSizeF() > 0 else 9.0
    font.setPointSizeF(base * SCALE)
    app.setFont(font)
    app.setStyleSheet(STYLE)


def main() -> None:
    app = QApplication(sys.argv)
    apply_scale(app)
    win = MainWindow()
    win.show()
    if len(sys.argv) > 1 and Path(sys.argv[1]).is_file():
        win.load_file(Path(sys.argv[1]))
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
