from pathlib import Path
"""
Generator wsadu CDON z feedów XML sm-prods - PySide6

Produkuje plik .xlsx z KOMPLETEM kolumn, które potrafią wczytać:
  - CDON_API_MULTI.py        (import, POST /v2/articles/bulk)
  - CDON_API_update_MULTI.py (aktualizacja, PUT /v2/articles/bulk)

Plik .xlsx wrzuca się do tamtych narzędzi bezpośrednio - bez konwersji do CSV.

Wygenerowany skoroszyt ma trzy arkusze:
  Dane              - wsad; kolumny wariantów mają listy wyboru i limity długości,
                      a po wpisaniu parent_sku podświetlają się w danym wierszu
  Właściwości CDON  - ściąga: co wolno wpisać w którą kolumnę property_*
  Listy             - ukryty; słowniki wartości dla właściwości preset-*

UWAGA: ten skrypt oraz oba powyższe tworzą jedną rodzinę - każdą zmianę
formatu wsadu trzeba nanieść na wszystkie trzy.

Wdrożenie do Centrum Zarządzania (wymaga uprawnień administratora):
  copy /Y "cdon_for_Magda.py" str((Path(__file__).parent / "scripts").resolve())

Kolumny wariantów (docs.cdon.com -> VARIATIONS):
  parent_sku              - wspólny identyfikator grupy; ten sam we wszystkich
                            wierszach jednego produktu (dowolny string, 1-64 znaki,
                            nie musi odpowiadać żadnemu istniejącemu SKU)
  property_<nazwa>        - wartość właściwości, np. property_color, property_size,
                            property_material (nazwy wg listy CDON)
  property_<nazwa>_language - język właściwości free-text (domyślnie en-US)
  property_language       - domyślny język free-text dla całego wiersza
  property_name_N / property_value_N / property_language_N
                          - wolne sloty na właściwości spoza listy CDON
  variational_properties  - nazwy właściwości różnicujących warianty, np. "color;size".
                            Zostaw pustą, a skrypt wyliczy ją sam z różnic między
                            wierszami o tym samym parent_sku.

Stara kolumna 'weight' nadal działa - jedzie jako property 'weight_kg'.
"""

import os
import re
import sys
import threading
import tempfile
import traceback
import xml.etree.ElementTree as ET

import csv
import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import requests

from PySide6.QtCore import Qt, QObject, Signal, QPoint, QEvent
from PySide6.QtGui import QColor, QImage, QPainter, QPolygon
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox,
    QLabel, QPushButton, QLineEdit, QCheckBox, QComboBox, QSpinBox,
    QProgressBar, QPlainTextEdit, QFileDialog, QMessageBox, QScrollArea, QFrame, QListView, QStyledItemDelegate
)

URL_GROUP_SUFFIXES = {
    "Normalna": {"se": "se", "dk": "dk", "fi": "fi"},
    "4%": {"se": "se_4", "dk": "dk_4", "fi": "fi_4"},
    "5,6%": {"se": "se_56", "dk": "dk_56", "fi": "fi_56"},
    "6,4%": {"se": "se_64", "dk": "dk_64", "fi": "fi_64"},
    "8%": {"se": "se_8", "dk": "dk_8", "fi": "fi_8"},
    "10%": {"se": "se_10", "dk": "dk_10", "fi": "fi_10"},
    "12%": {"se": "se_12", "dk": "dk_12", "fi": "fi_12"},
}

GROUP_ORDER = ["Normalna", "4%", "5,6%", "6,4%", "8%", "10%", "12%"]

# --- API sm-prods: lista zakazanych EAN-ów (forbidden_eans_paginated) ---
SMPRODS_TOKEN = "y7SeKeGSfVZtH9dCxwVULWTbcfWBrVq2WKcJssq8Pz8o5t3DFDpQ12BGRGc1S3fOJ2UC3tRMi29ChrseLAsl4GhHKR3Y9ALr9Zfq8pyeYtlExRas7rOfvRTrqKdEOJ8y"
FORBIDDEN_EANS_PAGINATED_URL = "https://api-sm-prods.sm-prods.com/forbidden_eans_paginated"

# Kraje UE (GPSR): jeśli producent jest spoza tej listy, CDON wymaga responsible_person.
EU_COUNTRY_CODES = {
    "AT", "BE", "BG", "HR", "CY", "CZ", "DK", "EE", "FI", "FR", "DE", "GR",
    "HU", "IE", "IT", "LV", "LT", "LU", "MT", "NL", "PL", "PT", "RO", "SK",
    "SI", "ES", "SE"
}

SPEC_MARKETS = ["SE", "DK", "FI"]

# --- RAPORT WARIANTÓW ---
# CSV zapisywany na Pulpicie przez CDON_API_MULTI.py / CDON_API_update_MULTI.py.
# Wskazany tutaj, uzupełnia w generowanym wsadzie parent_sku i kolumny property_*
# dla SKU, które już raz przez te narzędzia przeszły.

VARIANT_REPORT_FILENAME = "CDON_warianty_raport.csv"
VARIANT_REPORT_DELIMITER = ";"

# Kolumny raportu, które NIE są wartościami właściwości.
VARIANT_REPORT_META_COLUMNS = {"sku", "status", "source", "updated_at"}

# Kolumny, których NIE bierzemy z raportu - feed ma świeższe dane.
# (waga może się zmienić bez związku z wariantem; kategoria jest w build_row)
VARIANT_REPORT_IGNORED_COLUMNS = {"property_weight_kg", "property_weight_g"}


def resolve_desktop_dir():
    """
    Pulpit bieżącego użytkownika - ta sama logika co w skryptach API.
    Uwzględnia przekierowanie na OneDrive i polską nazwę 'Pulpit'.
    """
    try:
        import winreg
        key = winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
        try:
            value, _ = winreg.QueryValueEx(key, "Desktop")
        finally:
            winreg.CloseKey(key)
        value = os.path.expandvars(value or "")
        if value and os.path.isdir(value):
            return value
    except Exception:
        pass

    home = os.path.expanduser("~")
    for name in ("Desktop", "Pulpit",
                 os.path.join("OneDrive", "Desktop"),
                 os.path.join("OneDrive", "Pulpit")):
        candidate = os.path.join(home, name)
        if os.path.isdir(candidate):
            return candidate
    return home


def default_variant_report_path():
    """Domyślna ścieżka raportu na Pulpicie (może nie istnieć)."""
    return os.path.join(resolve_desktop_dir(), VARIANT_REPORT_FILENAME)


def load_variant_report(path):
    """
    Wczytuje raport wariantów -> {sku: {kolumna: wartość}}.
    Rzuca wyjątkiem, gdy pliku nie da się odczytać - błąd ma dotrzeć do GUI.
    """
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle, delimiter=VARIANT_REPORT_DELIMITER)
                if not reader.fieldnames or "sku" not in [
                        (name or "").strip() for name in reader.fieldnames]:
                    raise ValueError(
                        "Plik nie wygląda na raport wariantów - brak kolumny 'sku'.")
                entries = {}
                for raw in reader:
                    row = {
                        (key or "").strip(): (value or "").strip()
                        for key, value in raw.items() if key
                    }
                    sku = row.get("sku", "")
                    if sku:
                        entries[sku] = row
                return entries
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    raise ValueError(f"Nie rozpoznano kodowania pliku raportu ({last_error}).")


def report_property_names(report):
    """Nazwy właściwości występujące w raporcie (bez przedrostka property_)."""
    names = set()
    for row in report.values():
        for column, value in row.items():
            if (column.startswith("property_") and column != "property_language"
                    and column not in VARIANT_REPORT_IGNORED_COLUMNS
                    and str(value).strip()):
                names.add(column[len("property_"):])
    return sorted(names)

# --- WARIANTY (docs.cdon.com -> VARIATIONS) ---
# Warianty grupuje wspólny 'parent_sku'; różnicują je właściwości wymienione
# w 'variational_properties', a ich wartości siedzą w kolumnach property_<nazwa>.

FREE_TEXT_PROPERTIES = {
    "color", "connection_type", "flavor", "material", "pattern",
    "phone_case_type", "size"
}

PRESET_PROPERTIES = {
    "preset-color", "preset-connection_type", "preset-hair_type",
    "preset-media_format", "preset-pattern", "preset-refurb_grading",
    "preset-refurb_warranty", "preset-size_SML", "preset-skin_type"
}

NUMERIC_PROPERTIES = {
    "shoe_size_eu", "shoe_size_uk_men", "shoe_size_uk_women",
    "shoe_size_us_men", "shoe_size_us_women",
    "size_cm", "size_gb", "size_m", "size_mm", "size_unit",
    "volume_l", "volume_ml", "weight_g", "weight_kg"
}

KNOWN_PROPERTIES = FREE_TEXT_PROPERTIES | PRESET_PROPERTIES | NUMERIC_PROPERTIES

# Limity CDON - te same wartości egzekwuje importer.
PARENT_SKU_MAX_LENGTH = 64
FREE_TEXT_VALUE_MAX_LENGTH = 36
NUMERIC_VALUE_MAX_LENGTH = 50

# Zamknięte listy wartości dla właściwości preset-* (wprost z dokumentacji CDON).
# W Excelu trafiają do list wyboru, więc nie da się wpisać wartości spoza listy.
PRESET_PROPERTY_VALUES = {
    "preset-color": [
        "red", "blue", "green", "orange", "yellow", "purple", "pink", "gold",
        "silver", "multicolor", "white", "gray", "black", "turquoise", "brown",
        "beige", "transparent"],
    "preset-connection_type": ["bluetooth", "usb-c", "micro-usb", "lightning", "3.5mm"],
    "preset-hair_type": [
        "dry hair", "normal hair", "curly hair", "damaged hair", "frizzy hair",
        "fine hair", "dry scalp", "dyed hair"],
    "preset-media_format": ["dvd", "blu-ray", "4k ultra hd", "cd", "vinyl"],
    "preset-pattern": [
        "abstract and geometry", "animals and animal patterns",
        "nature and environment", "plants and fruit", "space and scifi",
        "text and quotes", "camouflage", "glitter", "marble and stone",
        "vehicles", "flags and symbols", "maps", "characters and celebrities",
        "fantasy", "retro"],
    "preset-refurb_grading": ["a+", "a", "b", "c"],
    "preset-refurb_warranty": ["2", "6", "12", "24"],
    "preset-size_SML": ["one size", "xxs", "xs", "s", "m", "l", "xl", "xxl"],
    "preset-skin_type": [
        "dry skin", "normal skin", "mature skin", "oily skin", "mixed skin", "acne"],
}

# Komplet właściwości CDON w kolejności: free-text, preset, numeryczne.
ALL_PROPERTY_NAMES = (
    sorted(FREE_TEXT_PROPERTIES) + sorted(PRESET_PROPERTIES) + sorted(NUMERIC_PROPERTIES)
)

# Najczęstsze osie wariantów - używane, gdy nie generujemy kompletu właściwości.
DEFAULT_VARIATION_PROPERTIES = ["color", "size"]

# Opis typu właściwości na arkusz-ściągę.
def property_kind(name):
    if name in PRESET_PROPERTIES:
        return "lista wyboru"
    if name in NUMERIC_PROPERTIES:
        return "liczba (max 50 znaków)"
    if name in FREE_TEXT_PROPERTIES:
        return "dowolny tekst (max 36 znaków)"
    return "własna"

STYLESHEET_TEMPLATE = """
QWidget { background-color: #ffffff; color: #1a1a1a; font-family: 'Segoe UI', Arial; font-size: 13px; }
QLabel, QCheckBox, QSlider { background: transparent; }
QGroupBox { background-color: #fff5fa; border: 1px solid #ff69b4; border-radius: 8px;
            margin-top: 9px; padding: 9px 8px 7px 8px; font-weight: bold; }
QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; color: #c2188b; }
QPushButton { background-color: #ff69b4; color: white; border: none; border-radius: 6px;
              padding: 8px 14px; font-weight: bold; }
QPushButton:hover { background-color: #e754a6; }
QPushButton:disabled { background-color: #f2cde1; color: #ffffff; }
QPushButton#stopButton { background-color: #c2188b; }
QPushButton#stopButton:hover { background-color: #a31574; }
QLineEdit, QComboBox, QSpinBox, QPlainTextEdit { background-color: #ffffff; border: 1px solid #ff69b4;
              border-radius: 6px; padding: 5px; color: #1a1a1a; }
QComboBox:hover { border-color: #e754a6; background-color: #fff5fa; }
QComboBox QAbstractItemView { background-color: #ffffff; border: 1px solid #ff69b4;
              border-radius: 6px; padding: 4px; outline: none; color: #1a1a1a;
              selection-background-color: #ff69b4; selection-color: #ffffff; }
QComboBox QAbstractItemView::item { min-height: 26px; padding: 4px 8px;
              border-radius: 4px; border: none; }
QComboBox QAbstractItemView::item:hover { background-color: #ffd6ea; color: #1a1a1a; }
QComboBox QAbstractItemView::item:selected { background-color: #ff69b4; color: #ffffff; }
QLineEdit:hover, QSpinBox:hover { border-color: #e754a6; }
QSpinBox::up-button, QSpinBox::down-button { width: 18px; border: none;
              background: transparent; }
QSpinBox::up-arrow { image: none; width: 0; height: 0;
              border-left: 4px solid transparent; border-right: 4px solid transparent;
              border-bottom: 5px solid #c2188b; }
QSpinBox::down-arrow { image: none; width: 0; height: 0;
              border-left: 4px solid transparent; border-right: 4px solid transparent;
              border-top: 5px solid #c2188b; }
QSpinBox::up-arrow:disabled, QSpinBox::down-arrow:disabled { border-bottom-color: #d9b6c8;
              border-top-color: #d9b6c8; }
QProgressBar { border: 1px solid #ff69b4; border-radius: 6px; text-align: center;
               background-color: #fff5fa; height: 18px; }
QProgressBar::chunk { background-color: #ff69b4; border-radius: 5px; }
QCheckBox::indicator { width: 16px; height: 16px; border: 1px solid #ff69b4;
                       border-radius: 4px; background: #ffffff; }
QCheckBox::indicator:checked { background: #ff69b4; }
"""


# --- Pobieranie i parsowanie XML ---

def download_xml(url):
    """Downloads and parses an XML file from a URL."""
    try:
        response = requests.get(url, timeout=120)
        response.raise_for_status()
        root = ET.fromstring(response.content)
        return root
    except requests.exceptions.RequestException as e:
        raise ConnectionError(f"Błąd podczas pobierania pliku z URL {url}: {e}")
    except ET.ParseError as e:
        raise ValueError(f"Błąd podczas parsowania XML z URL {url}: {e}")
    except Exception as e:
        raise Exception(f"Nieoczekiwany błąd podczas przetwarzania URL {url}: {e}")


def pobierz_forbidden_eans():
    """
    Pobiera zbiór zakazanych EAN-ów z API sm-prods przez paginowany endpoint
    /forbidden_eans_paginated (?page=&page_size=, odpowiedź {page, page_size,
    total, total_pages, data:[{id, ean}]}).
    Zwraca (zbior_ean, blad).
    """
    headers = {"Authorization": f"Bearer {SMPRODS_TOKEN}"}
    try:
        eans = set()
        page = 1
        while True:
            r = requests.get(
                FORBIDDEN_EANS_PAGINATED_URL,
                params={"page": page, "page_size": 1000},
                timeout=60, headers=headers)
            r.raise_for_status()
            payload = r.json()
            for x in payload.get("data", []):
                if x.get("ean"):
                    eans.add(str(x["ean"]).strip())
            if page >= payload.get("total_pages", page):
                break
            page += 1
        return eans, None
    except Exception as e:
        return set(), f"forbidden_eans_paginated: {e}"


def build_products_dict(root):
    """Builds a dictionary of products from the XML root for quick lookup."""
    products = {}
    if root is None:
        return products
    for offer in root.findall('o'):
        product_id = offer.get('id')
        if product_id:
            products[product_id] = offer
    return products


def get_attr(offer, attr_name, default=""):
    """Gets a specific attribute value from an offer's 'attrs' section."""
    attrs = offer.find('attrs')
    if attrs is not None:
        for a in attrs.findall('a'):
            if a.get('name') == attr_name:
                return (a.text or "").strip()
    return default


def get_category(offer):
    cat = offer.find('cat')
    return (cat.text or "").strip() if cat is not None else ""


def get_name(offer):
    name = offer.find('name')
    return (name.text or "").strip() if name is not None else ""


def get_desc(offer):
    desc = offer.find('desc')
    return (desc.text or "").strip() if desc is not None else ""


def get_main_image(offer):
    imgs = offer.find('imgs')
    if imgs is not None:
        main = imgs.find('main')
        if main is not None:
            return main.get('url', '')
    return ""


def get_extra_images(offer):
    """Gets up to 9 extra image URLs from an offer."""
    imgs = offer.find('imgs')
    urls = []
    if imgs is not None:
        all_i = imgs.findall('i')
        for i in all_i[:9]:
            url = i.get('url', '')
            if url:
                urls.append(url)
    return ";".join(urls)


def short(text, length):
    return text if len(text) <= length else text[:length]


def get_price(offer):
    return offer.get('price', "")


def get_stock(offer):
    return offer.get('stock', "")


def get_weight(offer):
    return offer.get('weight', "")


def strip_html_tags(text):
    if not isinstance(text, str):
        return text
    clean = re.compile('<.*?>')
    return re.sub(clean, '', text)


def get_brand(offer):
    """Determines the brand from attributes or product ID."""
    producent = get_attr(offer, "Producent")
    if producent:
        return producent.strip()
    prod_id = offer.get('id', '')
    if '_' in prod_id:
        return prod_id.split('_')[0]
    return prod_id


# --- Definicja kolumn wsadu ---

def split_list_cell(raw):
    """Rozbija 'color;size' albo 'color, size' na listę nazw właściwości."""
    if not raw:
        return []
    return [part.strip() for part in re.split(r"[;,]", str(raw)) if part.strip()]


def build_headers(spec_slots, variation_properties, property_slots):
    """
    Pełna lista kolumn czytanych przez CDON_API_MULTI.py i CDON_API_update_MULTI.py.

    spec_slots          = liczba par name/value na rynek dla sekcji 'specifications'
    variation_properties = nazwy właściwości -> kolumny property_<nazwa>
    property_slots      = liczba wolnych trójek property_name_N/value_N/language_N
    """
    headers = [
        "sku", "parent_sku", "weight", "brand", "gtin", "stock", "mainImage", "extraImages",
        "titleSe", "descriptionSe", "titleDk", "descriptionDk", "titleFi", "descriptionFi",
        "category",
        "originalPriceSe", "originalPriceDk", "originalPriceFi",
        "shippingCostSe", "shippingCostDk", "shippingCostFi",
        "deliveryTimeMinSe", "deliveryTimeMinDk", "deliveryTimeMinFi",
        "deliveryTimeMaxSe", "deliveryTimeMaxDk", "deliveryTimeMaxFi",
        "vatSe", "vatDk", "vatFi",
        "deliverySe", "deliveryDk", "deliveryFi",
        "shipped_from",
        # --- warianty: lista właściwości różnicujących ---
        # Zostaw pustą, a importer wyliczy ją sam z różnic między wierszami
        # o tym samym parent_sku.
        "variational_properties",
        # --- GPSR: producent i osoba odpowiedzialna w UE ---
        "manufacturer_name", "manufacturer_street_address", "manufacturer_city",
        "manufacturer_postal_code", "manufacturer_country", "manufacturer_website",
        "manufacturer_email",
        "responsible_person_name", "responsible_person_phone", "responsible_person_email",
    ]

    # --- właściwości wariantów: po jednej kolumnie na właściwość ---
    # Puste komórki są ignorowane, więc nadmiar kolumn niczemu nie szkodzi.
    if any(name in FREE_TEXT_PROPERTIES for name in variation_properties):
        # jeden język dla wszystkich free-textów w wierszu; w razie potrzeby można
        # dopisać ręcznie kolumnę property_<nazwa>_language dla pojedynczej właściwości
        headers.append("property_language")
    for name in variation_properties:
        headers.append(f"property_{name}")

    # --- wolne sloty na właściwości spoza listy CDON ---
    for idx in range(1, property_slots + 1):
        headers.append(f"property_name_{idx}")
        headers.append(f"property_value_{idx}")
        headers.append(f"property_language_{idx}")

    # --- specyfikacje techniczne per rynek ---
    for market in SPEC_MARKETS:
        headers.append(f"specification_{market}_group")
        for idx in range(1, spec_slots + 1):
            headers.append(f"specification_{market}_name_{idx}")
            headers.append(f"specification_{market}_value_{idx}")

    # awaryjna kolumna: gotowy JSON specyfikacji (używana, gdy kolumny wyżej są puste)
    headers.append("specifications_json")
    return headers


# --- Ubranie arkusza: listy wyboru, podświetlenie wariantów, ściąga ---

LIST_SHEET = "Listy"
REFERENCE_SHEET = "Właściwości CDON"

VARIANT_HEADER_FILL = PatternFill("solid", fgColor="FFD6EC")   # kolumny wariantów
VARIANT_ROW_FILL = PatternFill("solid", fgColor="FFF3F9")      # wiersz z parent_sku
HEADER_FILL = PatternFill("solid", fgColor="EDEDED")

# Ile pustych wierszy poniżej danych ma nadal mieć listy wyboru i podświetlenie -
# żeby dopisywanie wariantów ręcznie w Excelu działało tak samo jak wiersze z feeda.
EXTRA_EDITABLE_ROWS = 500


def _column_index(headers, name):
    """1-based numer kolumny albo None."""
    return headers.index(name) + 1 if name in headers else None


def _build_list_sheet(wb, property_names):
    """
    Ukryty arkusz ze słownikami wartości dla właściwości preset-*.
    Zwraca {nazwa_właściwości: 'Listy!$A$2:$A$18'}.
    """
    preset_names = [name for name in property_names if name in PRESET_PROPERTY_VALUES]
    if not preset_names:
        return {}

    ws = wb.create_sheet(LIST_SHEET)
    ranges = {}
    for col_idx, name in enumerate(preset_names, start=1):
        letter = get_column_letter(col_idx)
        values = PRESET_PROPERTY_VALUES[name]
        ws.cell(row=1, column=col_idx, value=name).font = Font(bold=True)
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.column_dimensions[letter].width = 28
        ranges[name] = f"{LIST_SHEET}!${letter}$2:${letter}${len(values) + 1}"

    ws.sheet_state = "hidden"
    return ranges


def _build_reference_sheet(wb, property_names):
    """Widoczna ściąga: co wolno wpisać w którą kolumnę property_*."""
    ws = wb.create_sheet(REFERENCE_SHEET)
    ws.append(["Kolumna w arkuszu Dane", "Typ wartości", "Dozwolone wartości"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    ws.append([
        "parent_sku", "dowolny tekst (max 64 znaki)",
        "Ta sama wartość we wszystkich wariantach jednego produktu. "
        "Nie musi istnieć jako osobny produkt. Puste = produkt bez wariantów."])
    ws.append([
        "variational_properties", "lista nazw rozdzielona średnikiem",
        "Zostaw puste - importer sam wykryje, czym różnią się wiersze o tym samym parent_sku. "
        "Wypełnij tylko, gdy chcesz nadpisać to wykrycie, np. 'color;size'."])
    ws.append([
        "property_language", "kod języka (np. sv-SE)",
        "Język właściwości tekstowych w tym wierszu. Puste = en-US."])

    for name in property_names:
        allowed = PRESET_PROPERTY_VALUES.get(name)
        ws.append([
            f"property_{name}",
            property_kind(name),
            ", ".join(allowed) if allowed else "dowolna wartość"])

    ws.append([
        "property_name_N / property_value_N", "własna właściwość",
        "Na właściwości spoza powyższej listy. UWAGA: CDON zna tylko nazwy wypisane wyżej, "
        "własne nazwy API zwykle odrzuca."])

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 110
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    return ws


def decorate_sheet(wb, ws, headers, property_names):
    """
    Ubiera arkusz 'Dane':
      - listy wyboru na kolumnach preset-* (pusta komórka nadal dozwolona),
      - podświetlenie kolumn wariantów w wierszach, gdzie wpisano parent_sku,
      - zamrożenie nagłówka i pierwszych kolumn,
      - ukryty arkusz ze słownikami + widoczna ściąga z dozwolonymi wartościami.
    """
    last_row = max(ws.max_row, 1) + EXTRA_EDITABLE_ROWS

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    def column_range(header):
        col = _column_index(headers, header)
        if col is None:
            return None
        letter = get_column_letter(col)
        return f"{letter}2:{letter}{last_row}"

    # --- 1. preset-*: lista wyboru, wartości spoza słownika zablokowane ---
    list_ranges = _build_list_sheet(wb, property_names)
    for name, source in list_ranges.items():
        target = column_range(f"property_{name}")
        if target is None:
            continue
        validation = DataValidation(
            type="list", formula1=source,
            allow_blank=True,          # puste = właściwość pominięta przy imporcie
            showDropDown=False,        # False = strzałka listy JEST widoczna (odwrotna logika OOXML)
            showErrorMessage=True,     # bez tego Excel wpuściłby dowolną wartość
            showInputMessage=True,
            errorStyle="stop",
            errorTitle="Wartość spoza listy CDON",
            error=f"Kolumna property_{name} przyjmuje tylko wartości ze słownika CDON. "
                  f"Zostaw pustą, jeśli ta właściwość nie dotyczy produktu.",
            promptTitle=f"property_{name}",
            prompt="Wybierz z listy albo zostaw puste.")
        ws.add_data_validation(validation)
        validation.add(target)

    # --- 2. free-text: twardy limit 36 znaków (powyżej CDON odrzuca artykuł) ---
    free_text_ranges = [column_range(f"property_{name}") for name in property_names
                        if name in FREE_TEXT_PROPERTIES]
    free_text_ranges = [r for r in free_text_ranges if r]
    if free_text_ranges:
        text_validation = DataValidation(
            type="textLength", operator="lessThanOrEqual",
            formula1=str(FREE_TEXT_VALUE_MAX_LENGTH),
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            errorStyle="stop",
            errorTitle="Za długa wartość",
            error=f"Właściwość tekstowa CDON może mieć najwyżej "
                  f"{FREE_TEXT_VALUE_MAX_LENGTH} znaków.",
            promptTitle="Właściwość tekstowa",
            prompt=f"Dowolny tekst, max {FREE_TEXT_VALUE_MAX_LENGTH} znaków. "
                   f"Zostaw puste, jeśli nie dotyczy.")
        ws.add_data_validation(text_validation)
        for target in free_text_ranges:
            text_validation.add(target)

    # --- 3. numeryczne: ostrzeżenie, gdy wpisane nie jest liczbą ---
    # Ostrzeżenie, a nie blokada, bo dokumentacja dopuszcza tu też wartości
    # formatowane (np. rozmiar buta "42/43") - użytkownik może świadomie potwierdzić.
    numeric_ranges = [column_range(f"property_{name}") for name in property_names
                      if name in NUMERIC_PROPERTIES]
    numeric_ranges = [r for r in numeric_ranges if r]
    if numeric_ranges:
        number_validation = DataValidation(
            type="decimal", operator="between",
            formula1="-1000000000", formula2="1000000000",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            errorStyle="warning",
            errorTitle="To nie wygląda na liczbę",
            error="Ta właściwość CDON oczekuje liczby (np. 0.5 albo 250). "
                  "Wartości formatowane typu '42/43' są dozwolone - wtedy kliknij Tak.",
            promptTitle="Właściwość liczbowa",
            prompt="Liczba, kropka jako separator dziesiętny. Zostaw puste, jeśli nie dotyczy.")
        ws.add_data_validation(number_validation)
        for target in numeric_ranges:
            number_validation.add(target)

    # --- 4. parent_sku: limit 64 znaków ---
    parent_range = column_range("parent_sku")
    if parent_range:
        parent_validation = DataValidation(
            type="textLength", operator="lessThanOrEqual",
            formula1=str(PARENT_SKU_MAX_LENGTH),
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            errorStyle="stop",
            errorTitle="Za długi parent_sku",
            error=f"parent_sku może mieć najwyżej {PARENT_SKU_MAX_LENGTH} znaków.",
            promptTitle="parent_sku",
            prompt="Ta sama wartość we wszystkich wariantach jednego produktu. "
                   "Puste = produkt bez wariantów.")
        ws.add_data_validation(parent_validation)
        parent_validation.add(parent_range)

    # --- podświetlenie kolumn wariantów, gdy wpisano parent_sku ---
    parent_col = _column_index(headers, "parent_sku")
    property_cols = [i + 1 for i, h in enumerate(headers)
                     if h.startswith("property_") or h == "variational_properties"]

    if parent_col:
        parent_letter = get_column_letter(parent_col)
        ws.cell(row=1, column=parent_col).fill = VARIANT_HEADER_FILL
        for col in property_cols:
            ws.cell(row=1, column=col).fill = VARIANT_HEADER_FILL

        if property_cols:
            rule = FormulaRule(formula=[f'${parent_letter}2<>""'], fill=VARIANT_ROW_FILL)
            first = get_column_letter(min(property_cols))
            last = get_column_letter(max(property_cols))
            ws.conditional_formatting.add(f"{first}2:{last}{last_row}", rule)
            ws.conditional_formatting.add(
                f"{parent_letter}2:{parent_letter}{last_row}",
                FormulaRule(formula=[f'${parent_letter}2<>""'], fill=VARIANT_ROW_FILL))

    # --- czytelność przy stu kolumnach ---
    ws.freeze_panes = "C2"
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        if header in ("sku", "parent_sku", "variational_properties"):
            ws.column_dimensions[letter].width = 24
        elif header.startswith("property_"):
            ws.column_dimensions[letter].width = 20

    _build_reference_sheet(wb, property_names)


class GeneratorSettings:
    """Wartości wpisywane do każdego wiersza - z GUI."""

    def __init__(self):
        self.spec_slots = 3
        self.all_properties = True
        self.variant_report_path = ""
        self.variant_report = {}
        self.variation_properties = list(DEFAULT_VARIATION_PROPERTIES)
        self.property_slots = 2
        self.shipped_from = "EU"
        self.vat = {"Se": 25, "Dk": 25, "Fi": 25.5}
        self.delivery_time_min = 4
        self.delivery_time_max = 6
        self.delivery_type = "HomeDelivery"
        self.manufacturer = {
            "manufacturer_name": "",
            "manufacturer_street_address": "",
            "manufacturer_city": "",
            "manufacturer_postal_code": "",
            "manufacturer_country": "",
            "manufacturer_website": "",
            "manufacturer_email": "",
            "responsible_person_name": "",
            "responsible_person_phone": "",
            "responsible_person_email": "",
        }

    def variation_warnings(self):
        """Ostrzeżenia o nazwach właściwości, których CDON nie zna."""
        if self.all_properties:
            return []
        unknown = [name for name in self.variation_properties if name not in KNOWN_PROPERTIES]
        if not unknown:
            return []
        return [
            "Właściwości spoza listy CDON: " + ", ".join(unknown) +
            ". Kolumny powstaną, ale API może je odrzucić. "
            "Znane nazwy to m.in.: color, size, material, pattern, flavor, "
            "connection_type, preset-color, preset-size_SML, size_cm, volume_ml."
        ]

    def property_names(self):
        """
        Właściwości, dla których powstaną kolumny property_<nazwa>.
        Do wybranych dokładamy te z raportu - inaczej wczytane wartości nie
        miałyby gdzie trafić (dotyczy też właściwości spoza listy CDON).
        """
        names = list(ALL_PROPERTY_NAMES) if self.all_properties else list(self.variation_properties)
        for name in report_property_names(self.variant_report):
            if name not in names:
                names.append(name)
        return names

    def manufacturer_warnings(self):
        """Zwraca listę ostrzeżeń o niekompletnych danych producenta (te same reguły co importer)."""
        m = self.manufacturer
        warnings = []
        filled = [v for v in m.values() if v.strip()]
        if not filled:
            return warnings

        if not m["manufacturer_name"].strip():
            warnings.append("Brak 'manufacturer_name' - importer pominie cały blok producenta.")
            return warnings

        missing = [key for key in (
            "manufacturer_street_address", "manufacturer_city",
            "manufacturer_postal_code", "manufacturer_country"
        ) if not m[key].strip()]
        if missing:
            warnings.append(
                "Niekompletny adres producenta (" + ", ".join(missing) +
                ") - importer pominie cały blok producenta.")
            return warnings

        country = m["manufacturer_country"].strip().upper()
        if len(country) != 2:
            warnings.append(f"'manufacturer_country' = '{country}' musi być 2-literowym kodem ISO.")
            return warnings

        if not m["responsible_person_name"].strip():
            if m["responsible_person_phone"].strip() or m["responsible_person_email"].strip():
                warnings.append("Podano telefon/e-mail osoby odpowiedzialnej, ale brak 'responsible_person_name'.")
            if country not in EU_COUNTRY_CODES:
                warnings.append(
                    f"Producent z kraju '{country}' (spoza UE) - CDON wymaga danych osoby odpowiedzialnej w UE.")
        return warnings


def build_row(headers, se_offer, dk_offer, fi_offer, settings):
    """Buduje jeden wiersz wsadu w kolejności zgodnej z 'headers'."""
    name = get_name(se_offer)
    desc = get_desc(se_offer)

    short_name = short(name, 135)
    processed_desc = strip_html_tags(desc) if len(desc) > 9500 else desc
    short_desc = short(processed_desc, 9500)

    values = {
        "sku": se_offer.get('id', ''),
        "weight": get_weight(se_offer),
        "brand": get_brand(se_offer),
        "gtin": get_attr(se_offer, "EAN"),
        "stock": get_stock(se_offer),
        "mainImage": get_main_image(se_offer),
        "extraImages": get_extra_images(se_offer),
        "titleSe": short_name, "descriptionSe": short_desc,
        "titleDk": short_name, "descriptionDk": short_desc,
        "titleFi": short_name, "descriptionFi": short_desc,
        "category": get_category(se_offer),
        "originalPriceSe": get_price(se_offer),
        "originalPriceDk": get_price(dk_offer) if dk_offer is not None else "",
        "originalPriceFi": get_price(fi_offer) if fi_offer is not None else "",
        "shippingCostSe": "0", "shippingCostDk": "0", "shippingCostFi": "0",
        "deliveryTimeMinSe": settings.delivery_time_min,
        "deliveryTimeMinDk": settings.delivery_time_min,
        "deliveryTimeMinFi": settings.delivery_time_min,
        "deliveryTimeMaxSe": settings.delivery_time_max,
        "deliveryTimeMaxDk": settings.delivery_time_max,
        "deliveryTimeMaxFi": settings.delivery_time_max,
        "vatSe": settings.vat["Se"], "vatDk": settings.vat["Dk"], "vatFi": settings.vat["Fi"],
        "deliverySe": settings.delivery_type,
        "deliveryDk": settings.delivery_type,
        "deliveryFi": settings.delivery_type,
        "shipped_from": settings.shipped_from,
    }
    values.update(settings.manufacturer)

    # Warianty z raportu: jeśli to SKU przeszło już przez importer/updater,
    # wpisujemy zapamiętane parent_sku i property_* zamiast zostawiać puste.
    reported = settings.variant_report.get(values["sku"])
    if reported:
        for column, value in reported.items():
            if column in VARIANT_REPORT_META_COLUMNS or not str(value).strip():
                continue
            if column in VARIANT_REPORT_IGNORED_COLUMNS:
                continue  # waga z feeda jest świeższa niż ta z raportu
            if column == "category" and values.get("category"):
                continue  # kategoria z feeda jest świeższa niż ta z raportu
            values[column] = value

    # pozostałe kolumny wariantów, specyfikacji i specifications_json zostają
    # puste - do uzupełnienia w Excelu
    return [values.get(header, "") for header in headers]


# --- GUI (PySide6) ---

class _ComboPopupRelayout(QObject):
    """
    Przelicza układ rozwiniętej listy w chwili jej pokazania.

    Reguła ::item z arkusza stylów zmienia wysokość pozycji (34 px zamiast 18),
    ale QListView układa wiersze według metryk sprzed jej zastosowania - bez
    tego pozycje nachodzą na siebie i podświetlenie zasłania sąsiednie wpisy.
    """

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Show:
            obj.doItemsLayout()
        return False


def polish_combo(combo):
    """
    Ładniejsza lista rozwijana.

    Styl Fusion rysuje ramkę kontenera popupu z palety, a nie z arkusza stylów -
    stąd czarna obwódka wokół rozwiniętej listy. Nadanie kontenerowi własnego
    arkusza przełącza go na renderowanie stylesheetem i obwódka znika.

    Wywołuj zaraz po utworzeniu combo, PRZED addItems(): setView() zeruje
    bieżącą pozycję widoku, więc później podświetlenie wybranego wpisu znika.
    """
    view = QListView()
    view.setFrameShape(QFrame.NoFrame)
    combo.setView(view)

    # QComboBox używa własnego delegata, który ignoruje wysokość z reguły
    # ::item; zwykły QStyledItemDelegate liczy ją poprawnie.
    combo.setItemDelegate(QStyledItemDelegate(combo))

    # filtr trzymamy przy widoku, żeby nie zniknął po wyjściu z funkcji
    view._popup_relayout = _ComboPopupRelayout(view)
    view.installEventFilter(view._popup_relayout)

    container = view.parentWidget()
    if container is None:
        return
    container.setObjectName("comboPopup")
    container.setFrameShape(QFrame.NoFrame)
    container.setStyleSheet("QFrame#comboPopup { border: none; background: transparent; }")
    container.setWindowFlags(
        container.windowFlags() | Qt.FramelessWindowHint | Qt.NoDropShadowWindowHint)
    container.setAttribute(Qt.WA_TranslucentBackground, True)


ARROW_COLOR = "#c2188b"
ARROW_COLOR_DISABLED = "#d9b6c8"

# Doklejane do arkusza tylko wtedy, gdy udało się zapisać ikony strzałek.
ARROW_STYLES = """
QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: center right;
                       width: 24px; border: none; background: transparent; }
QComboBox::down-arrow { image: url(__ARROW__); width: 10px; height: 6px; margin-right: 8px; }
QComboBox::down-arrow:disabled { image: url(__ARROW_OFF__); }
"""


def _draw_arrow_icon(path, color, width=10, height=6):
    """Rysuje trójkąt skierowany w dół do pliku PNG. Zwraca True przy sukcesie."""
    image = QImage(width, height, QImage.Format_ARGB32)
    image.fill(Qt.transparent)
    painter = QPainter(image)
    try:
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor(color))
        painter.drawPolygon(QPolygon([
            QPoint(0, 0), QPoint(width, 0), QPoint(width // 2, height)
        ]))
    finally:
        painter.end()
    return image.save(path, "PNG")


def _ensure_arrow_icons():
    """
    Przygotowuje strzałki list rozwijanych jako pliki PNG w katalogu tymczasowym.

    Qt nie umie narysować trójkąta z samych reguł arkusza stylów (cssowy trick
    z zerowym boksem i ramkami daje kwadratową plamę), a ostylowanie
    ::drop-down wyłącza natywny wskaźnik - dlatego dostarczamy własny obrazek.
    Katalog tymczasowy, a nie folder skryptu, bo narzędzia bywają uruchamiane
    z lokalizacji tylko do odczytu (np. str((Path(__file__).parent / "Program Files").resolve())).

    Zwraca (ścieżka_normalna, ścieżka_nieaktywna) albo (None, None).
    """
    try:
        directory = os.path.join(tempfile.gettempdir(), "cdon_gui_icons")
        os.makedirs(directory, exist_ok=True)
        normal = os.path.join(directory, f"arrow_{ARROW_COLOR.lstrip('#')}.png")
        disabled = os.path.join(directory, f"arrow_{ARROW_COLOR_DISABLED.lstrip('#')}.png")
        if not _draw_arrow_icon(normal, ARROW_COLOR):
            return None, None
        if not _draw_arrow_icon(disabled, ARROW_COLOR_DISABLED):
            return None, None
        # arkusz stylów wymaga ukośników w przód także na Windowsie
        return normal.replace("\\", "/"), disabled.replace("\\", "/")
    except Exception:
        return None, None


def build_stylesheet():
    """
    Arkusz stylów z podstawionymi ścieżkami do strzałek.
    Gdy ikon nie da się zapisać, pomijamy stylowanie ::drop-down - wtedy Qt
    rysuje własny wskaźnik i lista nadal działa, tylko wygląda skromniej.
    """
    normal, disabled = _ensure_arrow_icons()
    if not normal:
        return STYLESHEET_TEMPLATE
    return (STYLESHEET_TEMPLATE + ARROW_STYLES
            .replace("__ARROW__", normal)
            .replace("__ARROW_OFF__", disabled))


def _use_height_for_width(label):
    """
    Każe layoutowi liczyć wysokość etykiety z jej rzeczywistej szerokości.
    Bez tego QLabel z wordWrap rezerwuje miejsce na kilka linii nawet wtedy,
    gdy przy szerokim oknie tekst mieści się w jednej.
    """
    policy = label.sizePolicy()
    policy.setHeightForWidth(True)
    label.setSizePolicy(policy)


class WorkerSignals(QObject):
    log = Signal(str)
    progress = Signal(int, int)
    status = Signal(str)
    warning = Signal(str, str)
    finished = Signal(bool, str)


class FeedGeneratorWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Generator wsadu CDON - Qt")
        self.resize(960, 820)
        self.setMinimumSize(700, 480)
        self.setStyleSheet(build_stylesheet())

        self.is_running = False
        self.stop_event = threading.Event()

        self.signals = WorkerSignals()
        self.signals.log.connect(self._append_log)
        self.signals.progress.connect(self._set_progress)
        self.signals.status.connect(self.status_label_set)
        self.signals.warning.connect(self._show_warning)
        self.signals.finished.connect(self._on_finished)

        self._create_widgets()

    # --- Budowa interfejsu ---

    def _create_widgets(self):
        # Ustawienia trafiają do obszaru przewijanego, a akcje i log zostają
        # przypięte na dole - inaczej na ekranie 1080 px okno nie mieści się
        # w całości i Qt ściska widgety poniżej ich minimalnej wysokości.
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        outer.addWidget(scroll, 1)

        content = QWidget()
        scroll.setWidget(content)

        layout = QVBoxLayout(content)
        layout.setContentsMargins(12, 10, 12, 6)
        layout.setSpacing(6)

        header = QLabel("GENERATOR WSADU CDON")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("font-size: 18px; font-weight: bold; color: #ff69b4; padding: 4px;")
        layout.addWidget(header)

        # --- Źródło feedów ---
        source_box = QGroupBox("Źródło feedów")
        source_grid = QGridLayout(source_box)
        source_grid.setColumnStretch(1, 1)

        source_grid.addWidget(QLabel("Przedrostek pliku XML:"), 0, 0)
        self.prefix_edit = QLineEdit()
        self.prefix_edit.setPlaceholderText("np. armaservis")
        source_grid.addWidget(self.prefix_edit, 0, 1, 1, 3)

        source_grid.addWidget(QLabel("Grupy końcówek URL:"), 1, 0)
        groups_row = QGridLayout()
        self.group_checks = {}
        for index, group_name in enumerate(GROUP_ORDER):
            check = QCheckBox(group_name)
            check.setChecked(group_name == "Normalna")
            groups_row.addWidget(check, index // 4, index % 4)
            self.group_checks[group_name] = check
        source_grid.addLayout(groups_row, 1, 1, 1, 3)

        layout.addWidget(source_box)

        # --- Wartości domyślne wiersza ---
        defaults_box = QGroupBox("Wartości wpisywane do każdego wiersza")
        defaults_grid = QGridLayout(defaults_box)

        defaults_grid.addWidget(QLabel("shipped_from:"), 0, 0)
        self.shipped_from_combo = QComboBox()
        polish_combo(self.shipped_from_combo)
        self.shipped_from_combo.addItems(["EU", "NON_EU"])
        defaults_grid.addWidget(self.shipped_from_combo, 0, 1)

        defaults_grid.addWidget(QLabel("Typ dostawy:"), 0, 2)
        self.delivery_combo = QComboBox()
        polish_combo(self.delivery_combo)
        self.delivery_combo.addItems(["HomeDelivery", "ServicePoint", "Mailbox"])
        defaults_grid.addWidget(self.delivery_combo, 0, 3)

        defaults_grid.addWidget(QLabel("Czas dostawy min:"), 1, 0)
        self.delivery_min_spin = QSpinBox()
        self.delivery_min_spin.setRange(1, 9)
        self.delivery_min_spin.setValue(4)
        defaults_grid.addWidget(self.delivery_min_spin, 1, 1)

        defaults_grid.addWidget(QLabel("Czas dostawy max:"), 1, 2)
        self.delivery_max_spin = QSpinBox()
        self.delivery_max_spin.setRange(1, 9)
        self.delivery_max_spin.setValue(6)
        defaults_grid.addWidget(self.delivery_max_spin, 1, 3)

        defaults_grid.addWidget(QLabel("VAT SE / DK / FI (%):"), 2, 0)
        vat_row = QHBoxLayout()
        self.vat_se_edit = QLineEdit("25")
        self.vat_dk_edit = QLineEdit("25")
        self.vat_fi_edit = QLineEdit("25.5")
        for widget in (self.vat_se_edit, self.vat_dk_edit, self.vat_fi_edit):
            widget.setMaximumWidth(80)
            vat_row.addWidget(widget)
        vat_row.addStretch(1)
        defaults_grid.addLayout(vat_row, 2, 1, 1, 3)

        defaults_grid.addWidget(QLabel("Pary specyfikacji na rynek:"), 3, 0)
        self.spec_slots_spin = QSpinBox()
        self.spec_slots_spin.setRange(0, 20)
        self.spec_slots_spin.setValue(3)
        self.spec_slots_spin.setToolTip(
            "Ile pustych par kolumn specification_XX_name_N / specification_XX_value_N\n"
            "wygenerować dla każdego rynku (SE, DK, FI) do ręcznego uzupełnienia.")
        defaults_grid.addWidget(self.spec_slots_spin, 3, 1)

        defaults_grid.addWidget(QLabel("Wolne właściwości (sloty):"), 3, 2)
        self.property_slots_spin = QSpinBox()
        self.property_slots_spin.setRange(0, 20)
        self.property_slots_spin.setValue(2)
        self.property_slots_spin.setToolTip(
            "Ile pustych trójek property_name_N / property_value_N / property_language_N\n"
            "wygenerować na właściwości spoza listy poniżej.")
        defaults_grid.addWidget(self.property_slots_spin, 3, 3)

        layout.addWidget(defaults_box)

        # --- Warianty ---
        variants_box = QGroupBox("Warianty produktów (parent_sku)")
        variants_grid = QGridLayout(variants_box)
        variants_grid.setColumnStretch(1, 1)

        self.all_properties_check = QCheckBox(
            f"Wszystkie właściwości CDON ({len(ALL_PROPERTY_NAMES)} kolumn) z listami wyboru")
        self.all_properties_check.setChecked(True)
        self.all_properties_check.setToolTip(
            "Wsad dostaje kolumnę property_<nazwa> dla każdej właściwości znanej CDON.\n"
            "Puste kolumny są przy imporcie ignorowane, więc nadmiar nie szkodzi.")
        self.all_properties_check.toggled.connect(
            lambda checked: self.variation_edit.setEnabled(not checked))
        variants_grid.addWidget(self.all_properties_check, 0, 0, 1, 2)

        variants_grid.addWidget(QLabel("Tylko wybrane właściwości:"), 1, 0)
        self.variation_edit = QLineEdit(";".join(DEFAULT_VARIATION_PROPERTIES))
        self.variation_edit.setEnabled(False)
        self.variation_edit.setPlaceholderText("np. color;size")
        self.variation_edit.setToolTip(
            "Nazwy rozdzielone średnikiem. Dla każdej powstanie pusta kolumna\n"
            "property_<nazwa> do wypełnienia w Excelu.")
        variants_grid.addWidget(self.variation_edit, 1, 1)

        variants_hint = QLabel(
            "Ten sam 'parent_sku' we wszystkich wierszach jednego produktu; w kolumnach property_ "
            "wartości odróżniające warianty. Puste kolumny import ignoruje, a 'variational_properties' "
            "zostaw pustą - importer wyliczy ją sam. Ściąga: arkusz 'Właściwości CDON'.")
        variants_hint.setWordWrap(True)
        _use_height_for_width(variants_hint)
        variants_hint.setStyleSheet("color: #8a6b7c; font-weight: normal;")
        variants_grid.addWidget(QLabel("Raport wariantów:"), 2, 0)
        report_row = QHBoxLayout()
        report_row.setSpacing(6)
        self.report_edit = QLineEdit()
        self.report_edit.setPlaceholderText(
            "opcjonalnie - CSV zapisany na Pulpicie przez importer/updater")
        report_row.addWidget(self.report_edit, 1)
        self.report_browse_btn = QPushButton("Wybierz...")
        self.report_browse_btn.setMaximumWidth(110)
        self.report_browse_btn.clicked.connect(self.browse_variant_report)
        report_row.addWidget(self.report_browse_btn)
        self.report_clear_btn = QPushButton("Wyczyść")
        self.report_clear_btn.setMaximumWidth(90)
        self.report_clear_btn.clicked.connect(lambda: self.report_edit.setText(""))
        report_row.addWidget(self.report_clear_btn)
        variants_grid.addLayout(report_row, 2, 1)

        self.report_status = QLabel("")
        self.report_status.setStyleSheet("color: #8a6b7c; font-weight: normal;")
        variants_grid.addWidget(self.report_status, 3, 0, 1, 2)
        self.report_edit.textChanged.connect(self._refresh_report_status)

        variants_grid.addWidget(variants_hint, 4, 0, 1, 2)

        # raport z Pulpitu podpowiada się sam, jeśli już istnieje
        default_report = default_variant_report_path()
        if os.path.exists(default_report):
            self.report_edit.setText(default_report)
        else:
            self._refresh_report_status()

        layout.addWidget(variants_box)

        # --- Producent (GPSR) ---
        manufacturer_box = QGroupBox("Producent i osoba odpowiedzialna w UE (GPSR) - opcjonalne")
        manufacturer_grid = QGridLayout(manufacturer_box)
        manufacturer_grid.setColumnStretch(1, 1)
        manufacturer_grid.setColumnStretch(3, 1)
        manufacturer_grid.setColumnStretch(5, 1)

        self.manufacturer_edits = {}
        manufacturer_fields = [
            ("manufacturer_name", "Nazwa producenta:", 0, 0),
            ("manufacturer_email", "E-mail producenta:", 0, 2),
            ("manufacturer_website", "Strona www:", 0, 4),
            ("manufacturer_street_address", "Ulica i numer:", 1, 0),
            ("manufacturer_city", "Miasto:", 1, 2),
            ("manufacturer_postal_code", "Kod pocztowy:", 1, 4),
            ("manufacturer_country", "Kraj (ISO-2, np. CN):", 2, 0),
            ("responsible_person_name", "Osoba odpowiedzialna - nazwa:", 2, 2),
            ("responsible_person_email", "Osoba odpowiedzialna - e-mail:", 2, 4),
            ("responsible_person_phone", "Osoba odpowiedzialna - telefon:", 3, 0),
        ]
        for key, label, row, col in manufacturer_fields:
            manufacturer_grid.addWidget(QLabel(label), row, col)
            edit = QLineEdit()
            manufacturer_grid.addWidget(edit, row, col + 1)
            self.manufacturer_edits[key] = edit

        manufacturer_box.setToolTip(
            "Zostaw puste, jeśli dane producenta uzupełnisz później w Excelu.\n"
            "Producent spoza UE wymaga wypełnienia osoby odpowiedzialnej.")

        layout.addWidget(manufacturer_box)

        # --- Plik wyjściowy ---
        output_box = QGroupBox("Plik wyjściowy")
        output_grid = QGridLayout(output_box)
        output_grid.setColumnStretch(1, 1)

        output_grid.addWidget(QLabel("Folder docelowy:"), 0, 0)
        self.output_edit = QLineEdit(os.path.join(os.path.expanduser('~'), 'Desktop'))
        output_grid.addWidget(self.output_edit, 0, 1)
        self.browse_btn = QPushButton("Przeglądaj...")
        self.browse_btn.clicked.connect(self.browse_output_directory)
        output_grid.addWidget(self.browse_btn, 0, 2)

        layout.addWidget(output_box)
        layout.addStretch(1)

        # --- Akcje i log: zawsze widoczne, poza obszarem przewijanym ---
        bottom = QWidget()
        bottom_layout = QVBoxLayout(bottom)
        bottom_layout.setContentsMargins(14, 4, 14, 8)
        bottom_layout.setSpacing(4)
        outer.addWidget(bottom)

        button_row = QHBoxLayout()
        button_row.setSpacing(8)

        self.generate_btn = QPushButton("Generuj plik Excel")
        self.generate_btn.setMinimumHeight(38)
        self.generate_btn.clicked.connect(self.run_processing)
        button_row.addWidget(self.generate_btn, 3)

        self.stop_btn = QPushButton("ZATRZYMAJ")
        self.stop_btn.setObjectName("stopButton")
        self.stop_btn.setMinimumHeight(38)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self.stop_process)
        button_row.addWidget(self.stop_btn, 1)

        bottom_layout.addLayout(button_row)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        bottom_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("Gotowy")
        self.status_label.setAlignment(Qt.AlignCenter)
        bottom_layout.addWidget(self.status_label)

        log_box = QGroupBox("Postęp")
        log_layout = QVBoxLayout(log_box)
        self.log_view = QPlainTextEdit()
        self.log_view.setReadOnly(True)
        self.log_view.setMaximumBlockCount(5000)
        self.log_view.setMinimumHeight(60)
        self.log_view.setMaximumHeight(76)
        log_layout.addWidget(self.log_view)
        bottom_layout.addWidget(log_box)

    # --- Sloty GUI ---

    def browse_variant_report(self):
        start_dir = os.path.dirname(self.report_edit.text().strip()) or resolve_desktop_dir()
        filename, _ = QFileDialog.getOpenFileName(
            self, "Wybierz raport wariantów", start_dir,
            "Raport wariantów (*.csv);;Wszystkie pliki (*)")
        if filename:
            self.report_edit.setText(filename)

    def _refresh_report_status(self):
        """Pokazuje, ile produktów wczyta się z raportu - od razu przy wyborze pliku."""
        path = self.report_edit.text().strip()
        if not path:
            self.report_status.setText(
                "Bez raportu warianty trzeba wpisać ręcznie w Excelu.")
            return
        if not os.path.exists(path):
            self.report_status.setText(f"Nie znaleziono pliku: {path}")
            return
        try:
            report = load_variant_report(path)
        except Exception as exc:
            self.report_status.setText(f"Nie udało się wczytać raportu: {exc}")
            return
        names = report_property_names(report)
        self.report_status.setText(
            f"Wczytam {len(report)} produktów wariantowych"
            + (f"; właściwości: {', '.join(names)}" if names else ""))

    def browse_output_directory(self):
        directory = QFileDialog.getExistingDirectory(self, "Wybierz folder docelowy", self.output_edit.text())
        if directory:
            self.output_edit.setText(directory)

    def _append_log(self, message):
        self.log_view.appendPlainText(message)

    def _set_progress(self, current, total):
        self.progress_bar.setRange(0, max(total, 1))
        self.progress_bar.setValue(current)

    def status_label_set(self, text):
        self.status_label.setText(text)

    def _show_warning(self, title, message):
        QMessageBox.warning(self, title, message)

    def log(self, message):
        self.signals.log.emit(message)

    def collect_settings(self):
        """Zbiera ustawienia z GUI; zwraca (settings, blad)."""
        settings = GeneratorSettings()
        settings.spec_slots = self.spec_slots_spin.value()
        settings.property_slots = self.property_slots_spin.value()
        report_path = self.report_edit.text().strip()
        if report_path:
            if not os.path.exists(report_path):
                return None, f"Nie znaleziono raportu wariantów:\n{report_path}"
            try:
                settings.variant_report = load_variant_report(report_path)
            except Exception as exc:
                return None, f"Nie udało się wczytać raportu wariantów:\n{exc}"
            settings.variant_report_path = report_path

        settings.all_properties = self.all_properties_check.isChecked()
        settings.variation_properties = split_list_cell(self.variation_edit.text())
        if not settings.all_properties and not settings.variation_properties:
            return None, "Podaj przynajmniej jedną właściwość wariantu albo zaznacz 'Wszystkie właściwości CDON'."
        settings.shipped_from = self.shipped_from_combo.currentText()
        settings.delivery_type = self.delivery_combo.currentText()
        settings.delivery_time_min = self.delivery_min_spin.value()
        settings.delivery_time_max = self.delivery_max_spin.value()

        if settings.delivery_time_min > settings.delivery_time_max:
            return None, "Czas dostawy 'min' nie może być większy niż 'max'."

        for market, widget in (("Se", self.vat_se_edit), ("Dk", self.vat_dk_edit), ("Fi", self.vat_fi_edit)):
            raw = widget.text().strip().replace(',', '.')
            if not raw:
                return None, f"Podaj stawkę VAT dla rynku {market.upper()}."
            try:
                settings.vat[market] = float(raw)
            except ValueError:
                return None, f"Nieprawidłowa stawka VAT dla rynku {market.upper()}: '{widget.text()}'."

        for key, widget in self.manufacturer_edits.items():
            value = widget.text().strip()
            if key == "manufacturer_country":
                value = value.upper()
            settings.manufacturer[key] = value

        return settings, None

    def run_processing(self):
        prefix = self.prefix_edit.text().strip()
        output_dir = self.output_edit.text().strip()
        selected_groups = [name for name, check in self.group_checks.items() if check.isChecked()]

        if not prefix:
            QMessageBox.warning(self, "Brakujące dane", "Proszę wprowadzić przedrostek pliku.")
            return
        if not output_dir:
            QMessageBox.warning(self, "Brakujące dane", "Proszę podać folder docelowy.")
            return
        if not os.path.isdir(output_dir):
            QMessageBox.critical(self, "Błąd", f"Podany folder docelowy nie istnieje:\n{output_dir}")
            return
        if not selected_groups:
            QMessageBox.warning(self, "Brakujące dane", "Wybierz co najmniej jedną grupę końcówek URL.")
            return

        settings, error = self.collect_settings()
        if error:
            QMessageBox.warning(self, "Nieprawidłowe ustawienia", error)
            return

        variation_warnings = settings.variation_warnings()
        if variation_warnings:
            answer = QMessageBox.question(
                self, "Nieznane właściwości wariantów",
                "\n\n".join(variation_warnings) + "\n\nGenerować mimo to?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if answer != QMessageBox.Yes:
                return

        warnings = settings.manufacturer_warnings()
        if warnings:
            answer = QMessageBox.question(
                self, "Dane producenta niekompletne",
                "\n\n".join(warnings) + "\n\nGenerować mimo to?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if answer != QMessageBox.Yes:
                return

        filename = f"{prefix}_output_feeds.xlsx"
        output_file_path = os.path.join(output_dir, filename)

        if os.path.exists(output_file_path):
            answer = QMessageBox.question(
                self, "Potwierdzenie",
                f"Plik '{filename}' już istnieje w wybranej lokalizacji.\n\nCzy chcesz go nadpisać?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if answer != QMessageBox.Yes:
                return

        # Kolejność grup zgodna z GROUP_ORDER, żeby wynik był powtarzalny
        ordered_groups = [name for name in GROUP_ORDER if name in selected_groups]

        self.is_running = True
        self.stop_event.clear()
        self.generate_btn.setEnabled(False)
        self.generate_btn.setText("Przetwarzanie...")
        self.stop_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        self.log_view.clear()

        threading.Thread(
            target=self.process_feeds,
            args=(ordered_groups, prefix, output_file_path, settings),
            daemon=True
        ).start()

    def stop_process(self):
        if self.is_running:
            self.stop_event.set()
            self.stop_btn.setEnabled(False)
            self.stop_btn.setText("ZATRZYMYWANIE...")
            self.log("[INFO] Wysłano STOP - przerwę po bieżącym feedzie.")

    # --- Wątek roboczy ---

    def process_feeds(self, selected_groups, prefix, output_file_path, settings):
        """Pobiera feedy i buduje plik Excel. Działa w osobnym wątku."""
        try:
            self.signals.status.emit("Pobieram listę zakazanych EAN-ów z sm-prods...")
            self.log("Pobieram listę zakazanych EAN-ów z sm-prods...")
            forbidden_eans, blad_forbidden = pobierz_forbidden_eans()
            if blad_forbidden:
                self.log(f"[WARN] {blad_forbidden}")
                self.signals.warning.emit(
                    "Brak listy forbidden EAN",
                    "Nie udało się pobrać listy zakazanych EAN-ów - "
                    f"generowanie przebiegnie BEZ tego filtra.\n\n{blad_forbidden}")
            else:
                self.log(f"Zakazanych EAN-ów na liście: {len(forbidden_eans)}")

            property_names = settings.property_names()
            headers = build_headers(settings.spec_slots, property_names, settings.property_slots)
            self.log(f"Kolumn w wsadzie: {len(headers)} (w tym {settings.spec_slots} par specyfikacji na rynek)")
            if settings.variant_report:
                self.log(f"Raport wariantów: {settings.variant_report_path}")
                self.log(f"  wczytano {len(settings.variant_report)} produktów - "
                         f"parent_sku i właściwości wypełnią się automatycznie")
            else:
                self.log("Raport wariantów: nie wskazano - kolumny wariantów będą puste.")

            preset_count = sum(1 for name in property_names if name in PRESET_PROPERTY_VALUES)
            self.log(f"Kolumny wariantów: parent_sku, variational_properties oraz "
                     f"{len(property_names)} kolumn property_ ({preset_count} z listą wyboru) "
                     f"+ {settings.property_slots} wolnych slotów.")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dane"
            ws.append(headers)

            base_url = "https://sm-prods.com/feeds/"
            total_groups = len(selected_groups)
            skipped_forbidden = 0
            written_rows = 0
            report_hits = 0

            for group_index, group_name in enumerate(selected_groups, start=1):
                if self.stop_event.is_set():
                    self.log(f"[INFO] Przerwano przed grupą '{group_name}'.")
                    break

                group_suffixes = URL_GROUP_SUFFIXES[group_name]
                se_url = f"{base_url}{prefix}_cdon_{group_suffixes['se']}.xml"
                dk_url = f"{base_url}{prefix}_cdon_{group_suffixes['dk']}.xml"
                fi_url = f"{base_url}{prefix}_cdon_{group_suffixes['fi']}.xml"

                self.signals.status.emit(f"[{group_index}/{total_groups}] Pobieranie feeda SE ({group_name})...")
                root_se = download_xml(se_url)
                if self.stop_event.is_set():
                    break
                self.signals.status.emit(f"[{group_index}/{total_groups}] Pobieranie feeda DK ({group_name})...")
                root_dk = download_xml(dk_url)
                if self.stop_event.is_set():
                    break
                self.signals.status.emit(f"[{group_index}/{total_groups}] Pobieranie feeda FI ({group_name})...")
                root_fi = download_xml(fi_url)

                self.signals.status.emit(f"[{group_index}/{total_groups}] Budowanie słowników produktów ({group_name})...")
                se_offers = build_products_dict(root_se)
                dk_offers = build_products_dict(root_dk)
                fi_offers = build_products_dict(root_fi)

                total = len(se_offers)
                count = 0
                self.log(f"[{group_index}/{total_groups}] Grupa '{group_name}': {total} produktów w feedzie SE")

                for pid, se_offer in se_offers.items():
                    if self.stop_event.is_set():
                        self.log(f"[INFO] Przerwano w trakcie grupy '{group_name}'.")
                        break

                    count += 1
                    if count % 200 == 0:
                        self.signals.status.emit(f"[{group_index}/{total_groups}] {group_name}: produkt {count}/{total}...")
                        self.signals.progress.emit(count, total)

                    gtin = get_attr(se_offer, "EAN")

                    # FILTR ZAKAZANYCH EAN-ów - pomiń produkt, jeśli EAN na liście.
                    if forbidden_eans and gtin and gtin.strip() in forbidden_eans:
                        skipped_forbidden += 1
                        continue

                    ws.append(build_row(
                        headers, se_offer, dk_offers.get(pid), fi_offers.get(pid), settings
                    ))
                    written_rows += 1
                    if se_offer.get('id', '') in settings.variant_report:
                        report_hits += 1

                self.signals.progress.emit(total, total)

            if settings.variant_report:
                missed = len(settings.variant_report) - report_hits
                self.log(f"Warianty z raportu wypełnione w {report_hits} wierszach"
                         + (f" ({missed} SKU z raportu nie ma w tych feedach)" if missed > 0 else ""))

            self.signals.status.emit("Dodaję listy wyboru i podświetlenie wariantów...")
            decorate_sheet(wb, ws, headers, property_names)

            self.signals.status.emit("Zapisywanie pliku Excel...")
            self.log(f"Zapisywanie pliku: {output_file_path}")
            wb.save(output_file_path)

            summary = f"Plik '{os.path.basename(output_file_path)}' został zapisany."
            summary += f"\n\nWierszy: {written_rows}\nKolumn: {len(headers)}"
            if forbidden_eans:
                summary += f"\nOdsiano produktów z zakazanym EAN: {skipped_forbidden}"
            if settings.variant_report:
                summary += f"\nWarianty wypełnione z raportu: {report_hits} wierszy"
            if self.stop_event.is_set():
                summary += "\n\n(Przerwano przez użytkownika - plik zawiera dane częściowe)"

            self.log("--- GOTOWE ---")
            self.signals.finished.emit(True, summary)

        except Exception as e:
            self.log(f"[BŁĄD] {e}")
            traceback.print_exc()
            self.signals.finished.emit(False, str(e))

    def _on_finished(self, ok, message):
        self.is_running = False
        self.generate_btn.setEnabled(True)
        self.generate_btn.setText("Generuj plik Excel")
        self.stop_btn.setEnabled(False)
        self.stop_btn.setText("ZATRZYMAJ")
        self.status_label.setText("Gotowy")
        if ok:
            QMessageBox.information(self, "Sukces", message)
        else:
            QMessageBox.critical(self, "Błąd", message)

    def closeEvent(self, event):
        if self.is_running:
            answer = QMessageBox.question(
                self, "Generowanie w toku",
                "Generowanie wciąż trwa. Zatrzymać i zamknąć?",
                QMessageBox.Yes | QMessageBox.No
            )
            if answer != QMessageBox.Yes:
                event.ignore()
                return
            self.stop_event.set()
        event.accept()


def main():
    app = QApplication.instance() or QApplication(sys.argv)
    app.setStyle("Fusion")
    window = FeedGeneratorWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()