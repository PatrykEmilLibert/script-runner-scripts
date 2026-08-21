"""
XML -> CSV z wyczyszczonymi opisami.

Kopia xmlcsv2new.py z dwiema dodatkowymi kolumnami, wypelnianymi zawsze:

  desc_plain      - opis bez zadnego HTML-a, ale z zachowanymi enterami
                    (akapity, punktory, numeracja), przyciety do 2000 znakow
                    po kropce / koncu pozycji listy
  desc_marketing  - opis w HTML zawezonym do bialej listy Cdiscount
                    (b, strong, i, em, u, span, p, div, br, hr, ul, ol, li,
                    h1-h6), przyciety do 5000 znakow na granicy bloku,
                    z domknieciem tagow

Reszta - linki, filtry, raport bledow, uklad CSV - dziala identycznie jak
w xmlcsv2new.py. Skrypt jest samowystarczalny: nie wymaga zadnych plikow obok.
"""

import concurrent.futures
import csv
import ctypes
import os
import sys
import tempfile
import urllib.request
import xml.etree.ElementTree as ET
from ctypes import wintypes
from datetime import datetime
from urllib.parse import urlparse

import openpyxl
from openpyxl.utils import get_column_letter

try:
    from PySide6.QtCore import QThread, Qt, Signal
    from PySide6.QtGui import QColor, QFont, QLinearGradient, QPainter
    from PySide6.QtWidgets import (
        QApplication,
        QCheckBox,
        QComboBox,
        QFileDialog,
        QFrame,
        QGridLayout,
        QHBoxLayout,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPlainTextEdit,
        QProgressBar,
        QPushButton,
        QScrollArea,
        QSizePolicy,
        QVBoxLayout,
        QWidget,
    )
except ImportError as error:
    raise ImportError(
        "Brak PySide6. Zainstaluj: pip install PySide6 openpyxl"
    ) from error

# >>> WBUDOWANY-BLOK-START: czyszczenie-opisow
# Kod poniżej jest WKLEJONY automatycznie z desc_cleaner.py przez
# gen_skrypty_opisow.py. Nie edytuj go tutaj — zmiany nanoś w desc_cleaner.py
# i uruchom generator ponownie. Skrypt jest samowystarczalny: nie potrzebuje
# żadnych plików obok siebie.
import html
import re
from dataclasses import dataclass
from html.parser import HTMLParser
from typing import List, Optional, Tuple

# ---------------------------------------------------------------------------
# 1. KOREKTA ZNAKOW  (z "tlumaczenia v2.py")
# ---------------------------------------------------------------------------

EMOJI_PATTERN = re.compile(
    "["
    "\U0001F600-\U0001F64F"
    "\U0001F300-\U0001F5FF"
    "\U0001F680-\U0001F6FF"
    "\U0001F700-\U0001F77F"
    "\U0001F780-\U0001F7FF"
    "\U0001F800-\U0001F8FF"
    "\U0001F900-\U0001F9FF"
    "\U0001FA00-\U0001FA6F"
    "\U0001FA70-\U0001FAFF"
    "\U00002702-\U000027B0"
    "\U000024C2-\U0001F251"
    "\U0001f926-\U0001f937"
    "\U00010000-\U0010ffff"
    "♀-♂"
    "☀-⭕"
    "‍"
    "⏏"
    "⏩"
    "⌚"
    "️"
    "〰"
    "]+",
    flags=re.UNICODE,
)

POLISH_CHAR_MAP = {
    '&#378;ó&#322;ty': 'żółty',

    '&Aacute;': 'Ą', '&Cacute;': 'Ć', '&Eacute;': 'Ę',
    '&Lacute;': 'Ł', '&Nacute;': 'Ń', '&Oacute;': 'Ó',
    '&Sacute;': 'Ś', '&Zacute;': 'Ź', '&Zdot;': 'Ż',

    '&aacute;': 'ą', '&cacute;': 'ć', '&eacute;': 'ę',
    '&lacute;': 'ł', '&nacute;': 'ń', '&oacute;': 'ó',
    '&sacute;': 'ś', '&zacute;': 'ź', '&zdot;': 'ż',

    '&#260;': 'Ą', '&#262;': 'Ć', '&#280;': 'Ę', '&#321;': 'Ł',
    '&#323;': 'Ń', '&#211;': 'Ó', '&#346;': 'Ś', '&#377;': 'Ź',
    '&#379;': 'Ż',

    '&#261;': 'ą', '&#263;': 'ć', '&#281;': 'ę', '&#322;': 'ł',
    '&#324;': 'ń', '&#243;': 'ó', '&#347;': 'ś', '&#378;': 'ź',
    '&#380;': 'ż',

    '&deg;': '°', '&bull;': '•', '&ndash;': '–', '&rsquo;': '’',
    '&bdquo;': '„', '&rdquo;': '”',
    '&#10036;&#65039;': '', '&#10035;&#65039;': '', '&#9851;&#65039;': '',
    '&#128209;': '', '&#8222;': '„', '&#8221;': '”',
    '&#8216;': '‘', '&#8217;': '’', '&#8211;': '–', '&#8203;': '',
    '&#9989;': '', '&#9749;': '', '&#11088;': '', '&#10003;': '',
    '&#34;': '"', '&#39;': "'", '&#x2013;': '–', '&#2013;': '–',
    '&#2019;': '’', '&nbsp;': ' ', '&#178;': '²',
    '&#8220;': '“', '&#8230;': '…', '&#9679;': '•',

    '✔': '', '✅': '', '❓': '', '▶️': '',
    '⭐': '', '⚡': '', '➡': '',
}
# Najdluzsze klucze najpierw - zeby "&#10036;&#65039;" poszlo przed "&#10036;".
POLISH_CHAR_MAP_SORTED = dict(
    sorted(POLISH_CHAR_MAP.items(), key=lambda item: len(item[0]), reverse=True)
)

# Encje strukturalne HTML - rozwijane tylko wtedy, gdy tekst nie idzie juz
# przez parser HTML (inaczej zakodowany tekst stalby sie prawdziwym tagiem).
_MARKUP_ENTITIES = (
    ('&amp;', '&'), ('&lt;', '<'), ('&gt;', '>'),
    ('&quot;', '"'), ('&apos;', "'"),
)


def correct_text(text, keep_markup: bool = False):
    """
    Poprawia znaki: encje HTML -> polskie litery, usuwa emoji i smieci.

    keep_markup=True zostawia &lt; &gt; &amp; nietkniete.
    """
    if not isinstance(text, str) or not text:
        return text

    corrected = text
    for wrong, good in POLISH_CHAR_MAP_SORTED.items():
        if wrong in corrected:
            corrected = corrected.replace(wrong, good)

    if not keep_markup:
        for wrong, good in _MARKUP_ENTITIES:
            corrected = corrected.replace(wrong, good)

    corrected = EMOJI_PATTERN.sub('', corrected)
    return corrected.lstrip()


PLAIN_LIMIT = 2000
MARKETING_LIMIT = 5000

# ---------------------------------------------------------------------------
# 2. HTML -> CZYSTY TEKST  (z "csv_html_to_text_gui.py")
# ---------------------------------------------------------------------------

TAG_RE = re.compile(r"<\s*/?\s*[a-zA-Z][^>]*>")
ENCODED_TAG_RE = re.compile(r"&lt;\s*/?\s*[a-zA-Z]")

SKIP_TAGS = {
    "script", "style", "head", "noscript", "template", "svg", "iframe",
    # elementy interaktywne i media - ich tresc ("Kup", "Dodaj do koszyka")
    # nie ma nic wspolnego z opisem produktu
    "form", "button", "select", "option", "optgroup", "textarea", "datalist",
    "video", "audio", "object", "canvas", "dialog", "marquee",
}
# Zakazane, ale bez tagu zamykajacego - samo pominiecie tagu, bez licznika.
SKIP_VOID_TAGS = {"input", "link", "meta", "base", "param", "source", "track", "area", "col"}

PARAGRAPH_TAGS = {
    "p", "div", "h1", "h2", "h3", "h4", "h5", "h6", "blockquote", "table",
    "ul", "ol", "dl", "section", "article", "header", "footer", "main",
    "aside", "nav", "figure", "figcaption", "address", "fieldset", "form",
    "pre", "center", "hr", "caption",
}
LINE_TAGS = {"li", "tr", "dt", "dd", "thead", "tbody", "tfoot", "legend"}
CELL_TAGS = {"td", "th"}


@dataclass
class ConvertOptions:
    bullet: str = "• "
    number_ol: bool = True
    blank_between_paragraphs: bool = True
    max_newlines: int = 2
    cell_sep: str = " – "   # myslnik, nie "|" - "|" bywa separatorem CSV
    indent: str = "  "
    auto_unescape: bool = True
    keep_alt: bool = False
    bullet_char_for_ul_nested: bool = True
    strip_urls: bool = False    # marketplace'y zwykle nie chca adresow w opisie
    strip_emails: bool = False


class _HtmlToText(HTMLParser):
    """Zamienia HTML na tekst, zachowujac strukture jako entery."""

    def __init__(self, opts: ConvertOptions):
        super().__init__(convert_charrefs=True)
        self.opts = opts
        self.parts: List[str] = []
        self.pending_nl = 0
        self.at_start = True
        self.skip_depth = 0
        self.pre_depth = 0
        self.list_stack: List[List] = []

    def _write(self, text: str) -> None:
        if not text:
            return
        if self.pending_nl and not self.at_start:
            self.parts.append("\n" * self.pending_nl)
        self.pending_nl = 0
        self.parts.append(text)
        self.at_start = False

    def _newline(self, n: int = 1, additive: bool = False) -> None:
        if self.at_start:
            return
        limit = max(1, self.opts.max_newlines)
        if additive:
            self.pending_nl = min(limit, self.pending_nl + n)
        else:
            self.pending_nl = min(limit, max(self.pending_nl, n))

    def _block_end_newlines(self, tag: str) -> int:
        if tag in PARAGRAPH_TAGS and self.opts.blank_between_paragraphs:
            return 2
        return 1

    def _list_prefix(self) -> str:
        depth = max(0, len(self.list_stack) - 1)
        prefix = self.opts.indent * depth
        if not self.list_stack:
            return prefix + self.opts.bullet
        tag, counter = self.list_stack[-1]
        if tag == "ol" and self.opts.number_ol:
            self.list_stack[-1][1] = counter + 1
            return f"{prefix}{counter}. "
        return prefix + self.opts.bullet

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag in SKIP_VOID_TAGS:
            return
        if tag in SKIP_TAGS:
            self.skip_depth += 1
            return
        if self.skip_depth:
            return

        if tag == "br":
            self._newline(1, additive=True)
            return
        if tag == "img":
            if self.opts.keep_alt:
                alt = dict(attrs).get("alt") or ""
                if alt.strip():
                    self._write(alt.strip())
            return
        if tag == "pre":
            self.pre_depth += 1
            self._newline(self._block_end_newlines(tag))
            return
        if tag in ("ul", "ol"):
            self._newline(1)
            self.list_stack.append([tag, 1])
            return
        if tag == "li":
            self._newline(1)
            if self.pending_nl and not self.at_start:
                self.parts.append("\n" * self.pending_nl)
                self.pending_nl = 0
            self.parts.append(self._list_prefix())
            self.at_start = False
            return
        if tag in CELL_TAGS:
            if not self.at_start and not self.pending_nl:
                self._write(self.opts.cell_sep)
            return
        if tag == "hr":
            self._newline(1)
            self._write("---")
            self._newline(1)
            return
        if tag in PARAGRAPH_TAGS or tag in LINE_TAGS:
            self._newline(self._block_end_newlines(tag) if tag in PARAGRAPH_TAGS else 1)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in SKIP_VOID_TAGS:
            return
        if tag in SKIP_TAGS:
            self.skip_depth = max(0, self.skip_depth - 1)
            return
        if self.skip_depth:
            return
        if tag == "pre":
            self.pre_depth = max(0, self.pre_depth - 1)
            self._newline(self._block_end_newlines(tag))
            return
        if tag in ("ul", "ol"):
            if self.list_stack:
                self.list_stack.pop()
            self._newline(1 if self.list_stack else self._block_end_newlines(tag))
            return
        if tag in PARAGRAPH_TAGS or tag in LINE_TAGS or tag in CELL_TAGS:
            if tag in CELL_TAGS:
                return
            self._newline(self._block_end_newlines(tag) if tag in PARAGRAPH_TAGS else 1)

    def handle_data(self, data: str) -> None:
        if self.skip_depth or not data:
            return
        if self.pre_depth:
            self._write(data.replace("\r\n", "\n").replace("\r", "\n"))
            return
        text = data.replace("\xa0", " ")
        text = re.sub(r"\s+", " ", text)
        if not text.strip():
            if self.parts and not self.pending_nl and not self.at_start:
                if not self.parts[-1].endswith((" ", "\n")):
                    self.parts.append(" ")
            return
        self._write(text)

    def get_text(self) -> str:
        return "".join(self.parts)


def _clean_line(line: str) -> str:
    """Czysci linie zachowujac wciecie (listy zagniezdzone)."""
    m = re.match(r"^([ \t]*)(.*)$", line)
    indent, body = m.group(1), m.group(2)
    body = re.sub(r"[ \t]{2,}", " ", body).strip()
    return (indent + body).rstrip() if body else ""


def _tidy(text: str, max_newlines: int) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\xa0", " ")
    text = "\n".join(_clean_line(ln) for ln in text.split("\n"))
    limit = max(1, max_newlines)
    text = re.sub(r"\n{%d,}" % (limit + 1), "\n" * limit, text)
    return text.strip()


def html_to_plain(value, opts: Optional[ConvertOptions] = None) -> str:
    """HTML (lub czysty tekst) -> tekst z enterami. Zero tagow na wyjsciu."""
    opts = opts or ConvertOptions()
    if value is None:
        return ""
    text = str(value)
    if not text.strip():
        return ""

    has_tags = bool(TAG_RE.search(text))
    if not has_tags and opts.auto_unescape and ENCODED_TAG_RE.search(text):
        text = html.unescape(text)
        has_tags = bool(TAG_RE.search(text))

    if not has_tags:
        return _tidy(_strip_links(html.unescape(text), opts), opts.max_newlines)

    parser = _HtmlToText(opts)
    try:
        parser.feed(text)
        parser.close()
    except Exception:
        pass
    out = parser.get_text()
    if not out.strip():
        out = html.unescape(TAG_RE.sub(" ", text))
    return _tidy(_strip_links(out, opts), opts.max_newlines)


def _strip_links(text: str, opts: ConvertOptions) -> str:
    """Opcjonalnie wycina adresy URL i e-maile z czystego tekstu."""
    original = text
    # e-mail pierwszy: inaczej URL_RE wyjmie "sklep.pl" ze srodka adresu
    # i zostanie sierotka "kontakt@"
    if opts.strip_emails:
        text = EMAIL_RE.sub("", text)
    if opts.strip_urls:
        text = URL_RE.sub("", text)
    if text == original:
        return text
    return tidy_after_link_removal(text)


def looks_like_html(value) -> bool:
    if not value:
        return False
    s = str(value)
    return bool(TAG_RE.search(s)) or bool(ENCODED_TAG_RE.search(s)) or "&nbsp;" in s


# ---------------------------------------------------------------------------
# 3. HTML -> HTML DOZWOLONY PRZEZ CDISCOUNT
# ---------------------------------------------------------------------------

# Tagi wprost dozwolone przez Cdiscount.
ALLOWED_TAGS = {
    "b", "strong", "i", "em", "u", "span",
    "p", "div", "br", "hr",
    "ul", "ol", "li",
    "h1", "h2", "h3", "h4", "h5", "h6",
}
VOID_TAGS = {"br", "hr"}

# Tagi wyrzucane RAZEM Z TRESCIA (skrypty, media, formularze, kod strony).
DROP_TREE_TAGS = {
    "script", "style", "head", "noscript", "template", "svg", "math",
    "iframe", "video", "audio", "object", "canvas", "form", "button",
    "select", "option", "optgroup", "textarea", "datalist", "output",
    "progress", "meter", "applet", "frame", "frameset", "dialog",
    "picture", "map", "marquee", "blink",
}
# Zakazane, ale bez tagu zamykajacego - samo pominiecie, inaczej "polkna"
# caly dalszy opis (np. <input value="x"> bez </input>).
DROP_VOID_TAGS = {
    "input", "link", "meta", "base", "param", "source", "track",
    "area", "col", "embed", "wbr",
}

# Tagi blokowe, ktorych sama tresc zostaje - zamieniamy je na <div>.
BLOCK_TO_DIV = {
    "html", "body", "header", "footer", "main", "section", "article",
    "aside", "nav", "figure", "figcaption", "address", "fieldset",
    "blockquote", "pre", "center", "dl", "dt", "dd", "details", "summary",
    "legend", "caption",
}

TABLE_TAGS = {"table", "thead", "tbody", "tfoot", "tr", "td", "th", "colgroup", "col"}

# Tagi blokowe uzywane jako "bezpieczne miejsce ciecia" przy przycinaniu.
BLOCK_CLOSERS = {"p", "div", "li", "ul", "ol", "h1", "h2", "h3", "h4", "h5", "h6"}

URL_RE = re.compile(
    # adres ze schematem / z "www." - koncowa kropka lub przecinek nalezy do
    # zdania, nie do adresu, wiec ostatni znak dopasowania nie moze byc
    # interpunkcja ("Zajrzyj na www.sklep.pl." -> zostaje kropka)
    r"(?i)\b(?:https?://|ftp://|www\.)[^\s<>\"']*[^\s<>\"'.,;:!?)\]]"
    # goly adres domenowy, np. "sklep.pl/oferta"
    r"|\b[a-z0-9][a-z0-9\-]*\.(?:pl|com|de|fr|net|eu|nl|cz|it|es)"
    r"(?:/[^\s<>\"']*[^\s<>\"'.,;:!?)\]])?\b"
)
EMAIL_RE = re.compile(r"(?i)\b[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}\b")

# Po wycieciu adresu zostaje dziura: "Sprawdz na  !" albo "Zobacz (  )".
# Sprzatamy TYLKO wtedy, gdy faktycznie cos usunelismy - inaczej zniszczylibysmy
# francuska typografie, w ktorej spacja przed "!" i "?" jest poprawna.
_SPACE_BEFORE_PUNCT_RE = re.compile(r"[ \t]+(?=[.,;:!?])")
_EMPTY_BRACKETS_RE = re.compile(r"[(\[]\s*[)\]]")


def tidy_after_link_removal(text: str) -> str:
    """Zasklepia dziure po usunietym adresie URL / e-mailu."""
    text = _EMPTY_BRACKETS_RE.sub("", text)
    text = _SPACE_BEFORE_PUNCT_RE.sub("", text)
    return re.sub(r"[ \t]{2,}", " ", text)


@dataclass
class HtmlOptions:
    strip_urls: bool = True          # Cdiscount zabrania linkow i adresow URL
    strip_emails: bool = True
    tables_to_lists: bool = True     # <table> -> <ul><li>Nazwa: Wartosc</li>
    keep_img_alt: bool = False
    max_consecutive_br: int = 2
    normalize_equals: bool = True    # "==" blokuje walidator Cdiscount
    headers_to_bold: bool = False    # h1..h6 -> <p><strong>...</strong></p>


class _HtmlSanitizer(HTMLParser):
    """Przepuszcza tylko tagi z bialej listy Cdiscount, reszte rozpakowuje."""

    def __init__(self, opts: HtmlOptions):
        super().__init__(convert_charrefs=True)
        self.opts = opts
        self.out: List[str] = []
        self.stack: List[str] = []
        self.drop_depth = 0
        self.table_depth = 0
        self.in_cell = False
        self.cell: List[str] = []
        self.row: List[str] = []
        self.row_is_header = False
        self.removed_links = False
        self._stack_backup: List[str] = []

    # -- wyjscie ------------------------------------------------------------
    def _sink(self) -> List[str]:
        return self.cell if self.in_cell else self.out

    def _emit(self, chunk: str) -> None:
        if chunk:
            self._sink().append(chunk)

    def _open(self, tag: str) -> None:
        self._emit(f"<{tag}>")
        self.stack.append(tag)

    def _close(self, tag: str) -> None:
        """Domyka tag razem ze wszystkim, co zostalo otwarte w srodku."""
        if tag not in self.stack:
            return
        while self.stack:
            top = self.stack.pop()
            self._emit(f"</{top}>")
            if top == tag:
                break

    def _is_header(self, tag: str) -> bool:
        return len(tag) == 2 and tag[0] == "h" and tag[1].isdigit()

    def _last_index(self, tags) -> int:
        """Pozycja ostatniego z podanych tagow na stosie albo -1."""
        for index in range(len(self.stack) - 1, -1, -1):
            if self.stack[index] in tags:
                return index
        return -1

    def _map_tag(self, tag: str) -> Optional[str]:
        if tag in ALLOWED_TAGS:
            if self._is_header(tag) and self.opts.headers_to_bold:
                return None
            return tag
        if tag in BLOCK_TO_DIV:
            return "div"
        return None

    # -- komorki tabeli -----------------------------------------------------
    def _enter_cell(self) -> None:
        self.in_cell = True
        self.cell = []
        self._stack_backup = self.stack
        self.stack = []

    def _exit_cell(self) -> str:
        while self.stack:
            self._emit(f"</{self.stack.pop()}>")
        raw = "".join(self.cell)
        self.in_cell = False
        self.stack = self._stack_backup
        text = re.sub(r"<[^>]+>", " ", raw)
        return re.sub(r"\s{2,}", " ", text).strip()

    # -- HTMLParser ---------------------------------------------------------
    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()

        if tag in DROP_VOID_TAGS:
            return
        if tag in DROP_TREE_TAGS:
            self.drop_depth += 1
            return
        if self.drop_depth:
            return

        if tag == "img":
            if self.opts.keep_img_alt:
                alt = (dict(attrs).get("alt") or "").strip()
                if alt:
                    self.handle_data(alt)
            return

        if tag in TABLE_TAGS:
            if self.opts.tables_to_lists and not self.in_cell:
                self._handle_table_start(tag)
            return

        if tag == "br":
            self._emit("<br>")
            return
        if tag == "hr":
            self._close("p")
            self._emit("<hr>")
            return

        if tag == "li":
            # Zamykamy poprzedni <li> tylko wtedy, gdy nie zdazyla sie w nim
            # otworzyc lista zagniezdzona - inaczej _close("li") zwinelby ja
            # i cala hierarchia splaszczylaby sie do jednego poziomu.
            last_li = self._last_index(("li",))
            last_list = self._last_index(("ul", "ol"))
            if last_li > last_list:
                self._close("li")
            if last_list < 0 and "li" not in self.stack:
                self._open("ul")
            self._open("li")
            return

        if tag in ("ul", "ol"):
            if "p" in self.stack:
                self._close("p")
            self._open(tag)
            return

        mapped = self._map_tag(tag)
        if mapped is None:
            # naglowek zamieniany na pogrubiony akapit
            if self._is_header(tag) and self.opts.headers_to_bold:
                self._close("p")
                self._open("p")
                self._open("strong")
            # reszta (m.in. <a>, <font>, <sup>) - tylko rozpakowanie tresci
            return

        if mapped == "p" and "p" in self.stack:
            self._close("p")
        self._open(mapped)

    def handle_startendtag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag in DROP_TREE_TAGS or tag in DROP_VOID_TAGS:
            return
        if tag in VOID_TAGS or tag == "img":
            self.handle_starttag(tag, attrs)
            return
        self.handle_starttag(tag, attrs)
        self.handle_endtag(tag)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()

        if tag in DROP_VOID_TAGS:
            return
        if tag in DROP_TREE_TAGS:
            self.drop_depth = max(0, self.drop_depth - 1)
            return
        if self.drop_depth:
            return
        if tag in ("img", "br", "hr"):
            return

        if tag in TABLE_TAGS:
            if self.opts.tables_to_lists:
                self._handle_table_end(tag)
            return

        mapped = self._map_tag(tag)
        if mapped is None:
            if self._is_header(tag) and self.opts.headers_to_bold:
                self._close("p")
            return
        self._close(mapped)

    def handle_data(self, data: str) -> None:
        if self.drop_depth or not data:
            return
        if self.table_depth and not self.in_cell:
            return

        text = data.replace("\xa0", " ").replace("​", "")
        text = re.sub(r"\s+", " ", text)
        if not text.strip():
            sink = self._sink()
            if sink and not sink[-1].endswith((" ", ">")):
                self._emit(" ")
            return

        before_links = text
        if self.opts.strip_emails:
            text = EMAIL_RE.sub("", text)
        if self.opts.strip_urls:
            text = URL_RE.sub("", text)
        if text != before_links:
            self.removed_links = True
        if self.opts.normalize_equals:
            text = re.sub(r"={2,}", "=", text)
        text = re.sub(r"\s{2,}", " ", text)
        if not text.strip():
            return

        self._emit(html.escape(text, quote=False))

    def close(self):  # type: ignore[override]
        super().close()
        if self.in_cell:
            self._exit_cell()
        while self.stack:
            self._emit(f"</{self.stack.pop()}>")

    # -- tabele -------------------------------------------------------------
    def _handle_table_start(self, tag: str) -> None:
        if tag == "table":
            self._close("p")
            self.table_depth += 1
            self._open("ul")
            return
        if not self.table_depth:
            return
        if tag == "tr":
            self.row = []
            self.row_is_header = True
            return
        if tag in ("td", "th"):
            if tag == "td":
                self.row_is_header = False
            self._enter_cell()

    def _handle_table_end(self, tag: str) -> None:
        if not self.table_depth:
            return
        if tag in ("td", "th"):
            if self.in_cell:
                self.row.append(self._exit_cell())
            return
        if tag == "tr":
            cells = [c for c in self.row if c]
            header_row = self.row_is_header
            self.row = []
            self.row_is_header = False
            if not cells:
                return
            if header_row:
                # wiersz zlozony wylacznie z <th> to naglowek tabeli - nie para
                # "nazwa: wartosc", wiec sklejamy myslnikiem i pogrubiamy
                line = " – ".join(cells)
                self._open("li")
                self._open("strong")
                self._emit(html.escape(line, quote=False))
                self._close("strong")
                self._close("li")
                return
            line = f"{cells[0]}: {cells[1]}" if len(cells) == 2 else " – ".join(cells)
            self._open("li")
            self._emit(html.escape(line, quote=False))
            self._close("li")
            return
        if tag == "table":
            if self.in_cell:
                self.row.append(self._exit_cell())
            self._close("ul")
            self.table_depth = max(0, self.table_depth - 1)

    def get_html(self) -> str:
        return "".join(self.out)


_EMPTY_EL_RE = re.compile(
    r"<(p|div|span|b|strong|i|em|u|h[1-6]|li|ul|ol)>(?:\s|<br\s*/?>|&nbsp;)*</\1>",
    re.IGNORECASE,
)


def _post_process_html(s: str, opts: HtmlOptions) -> str:
    if not s:
        return ""
    max_br = max(1, opts.max_consecutive_br)
    s = re.sub(
        r"(?:\s*<br\s*/?>\s*){%d,}" % (max_br + 1),
        "<br>" * max_br,
        s,
        flags=re.IGNORECASE,
    )
    # <br> tuz przed/po granicy bloku nic nie wnosi
    s = re.sub(r"(?:\s*<br\s*/?>\s*)+(</(?:p|div|li|h[1-6]|ul|ol)>)", r"\1", s, flags=re.I)
    s = re.sub(r"(<(?:p|div|li|h[1-6]|ul|ol)>)(?:\s*<br\s*/?>\s*)+", r"\1", s, flags=re.I)

    while True:
        new = _EMPTY_EL_RE.sub("", s)
        if new == s:
            break
        s = new

    s = re.sub(r"[ \t]{2,}", " ", s)
    s = re.sub(r"\s+(</(?:p|div|li|h[1-6]|span|b|strong|i|em|u)>)", r"\1", s, flags=re.I)
    s = re.sub(r"^(?:\s|<br\s*/?>)+", "", s, flags=re.I)
    s = re.sub(r"(?:\s|<br\s*/?>)+$", "", s, flags=re.I)
    return s.strip()


def html_to_marketing(value, opts: Optional[HtmlOptions] = None) -> str:
    """
    HTML -> HTML zawezony do bialej listy Cdiscount.

    - zostaja: b, strong, i, em, u, span, p, div, br, hr, ul, ol, li, h1-h6
    - znikaja z trescia: script, style, iframe, video, audio, form, input, button...
    - <table> zamieniane na <ul><li>Nazwa: Wartosc</li></ul>
    - <a> i <img> rozpakowane / usuniete, adresy URL wycinane z tekstu
    - wszystkie atrybuty usuwane (zero style, class, data-*, brak "==" w kodzie)
    """
    opts = opts or HtmlOptions()
    if value is None:
        return ""
    text = str(value)
    if not text.strip():
        return ""

    has_tags = bool(TAG_RE.search(text))
    if not has_tags and ENCODED_TAG_RE.search(text):
        text = html.unescape(text)
        has_tags = bool(TAG_RE.search(text))

    if not has_tags:
        # czysty tekst - odtwarzamy akapity z enterow
        plain = _tidy(html.unescape(text), 2)
        stripped = plain
        if opts.strip_emails:
            stripped = EMAIL_RE.sub("", stripped)
        if opts.strip_urls:
            stripped = URL_RE.sub("", stripped)
        plain = stripped if stripped == plain else tidy_after_link_removal(stripped)
        blocks = [b.strip() for b in re.split(r"\n{2,}", plain) if b.strip()]
        parts = []
        for block in blocks:
            body = "<br>".join(html.escape(ln, quote=False) for ln in block.split("\n"))
            parts.append(f"<p>{body}</p>")
        return _post_process_html("".join(parts), opts)

    parser = _HtmlSanitizer(opts)
    try:
        parser.feed(text)
        parser.close()
    except Exception:
        pass
    result = _post_process_html(parser.get_html(), opts)
    if parser.removed_links:
        # dziura po adresie moze siegac przez granice wezlow tekstowych
        # ("...na <a>sklep.pl</a>!"), wiec czyscimy dopiero zlozony HTML
        result = tidy_after_link_removal(result)
    return result


# ---------------------------------------------------------------------------
# 4. INTELIGENTNE PRZYCINANIE
# ---------------------------------------------------------------------------

# Skroty, po ktorych kropka praktycznie nigdy nie konczy zdania - nawet gdy
# dalej idzie wielka litera ("ul. Kwiatowa", "nr. Katalogowy", "dr. Nowak").
# Jednostek (cm, kg, szt...) tu NIE ma - one czesto koncza zdanie, a przypadek
# "5 cm. dalej" wylapuje regula "po kropce mala litera / cyfra".
ABBREVIATIONS = {
    "np", "tj", "tzn", "ul", "al", "nr", "str", "dr", "mgr", "inż", "inz",
    "prof", "tel", "im", "pn", "ww", "wg", "ds", "poz", "art", "pkt",
    "rys", "fot", "tab", "por", "zob", "cd", "ang", "niem", "franc",
}

_SENTENCE_END_RE = re.compile("[.!?…](?=[\"'”’»)\\]]?(?:\\s|$))")
_CLOSERS = "\"'”’»)]"


def _is_real_sentence_end(text: str, dot_pos: int) -> bool:
    """
    Odrzuca kropki, ktore nie koncza zdania:
    - po skrocie ("np.", "ul.")
    - gdy zaraz po niej idzie mala litera albo cyfra ("ok. 5 kg", "2.5 m")
    - po inicjale ("J. Kowalski")
    - po numerze pozycji listy na poczatku linii ("3. Montaz")
    """
    before = text[:dot_pos]
    after = text[dot_pos + 1:dot_pos + 16].lstrip(_CLOSERS + " \t\r\n")
    next_char = after[:1]
    if next_char and (next_char.islower() or next_char.isdigit()):
        return False

    word = re.search(r"([A-Za-zÀ-ɏ]+)$", before)
    if word:
        token = word.group(1).lower()
        if len(token) == 1:
            return False          # inicjal, np. "J. Kowalski"
        return token not in ABBREVIATIONS

    if re.search(r"\d$", before):
        line_start = before.rfind("\n") + 1
        return not before[line_start:].strip().isdigit()   # "3." = marker listy
    return True


def _find_cut(window: str, floor: int) -> int:
    """Zwraca najlepsza pozycje ciecia w oknie tekstu albo -1."""
    # 1. koniec akapitu
    pos = window.rfind("\n\n")
    if pos >= floor:
        return pos
    # 2. koniec linii / pozycji listy
    pos = window.rfind("\n")
    if pos >= floor:
        return pos
    # 3. koniec zdania
    best = -1
    for m in _SENTENCE_END_RE.finditer(window):
        end = m.end()
        if end < len(window) and window[end] in _CLOSERS:
            end += 1
        if end < floor:
            continue
        if not _is_real_sentence_end(window, m.start()):
            continue
        best = end
    if best >= floor:
        return best
    # 4. srednik / dwukropek
    for sep in ("; ", ": "):
        pos = window.rfind(sep)
        if pos >= floor:
            return pos + 1
    # 5. przecinek
    pos = window.rfind(", ")
    if pos >= floor:
        return pos + 1
    # 6. ostatnia spacja
    pos = window.rfind(" ")
    if pos >= floor:
        return pos
    return -1


_DANGLING_LINE_RE = re.compile(r"^\s*(?:[•\-\*–]|\d+\.)?\s*$")
_TAIL_JUNK_RE = re.compile("[\\s,;:\\-–—/(\\[„“\"']+$")


def _tidy_tail(text: str) -> str:
    """Ucina wiszacy punktor / niedokonczona interpunkcje na koncu."""
    lines = text.split("\n")
    while lines and _DANGLING_LINE_RE.match(lines[-1] or ""):
        lines.pop()
    text = "\n".join(lines).rstrip()
    text = _TAIL_JUNK_RE.sub("", text)
    return text.strip()


def smart_truncate_text(text: str, limit: int, soft_ratio: float = 0.6) -> str:
    """
    Przycina tekst do `limit` znakow tak, zeby nie urwac w polowie zdania.

    Priorytet ciecia: koniec akapitu > koniec linii/pozycji listy > koniec
    zdania > srednik/dwukropek > przecinek > spacja. Ciecie musi wypasc
    powyzej soft_ratio * limit, inaczej schodzimy nizej w priorytecie.
    """
    if not text:
        return ""
    text = text.strip()
    if limit <= 0 or len(text) <= limit:
        return text

    window = text[:limit]
    floor = int(limit * soft_ratio)
    cut = _find_cut(window, floor)
    if cut < 0:
        cut = limit
    result = _tidy_tail(window[:cut])
    return result if result else window[:limit].rstrip()


_TOKEN_RE = re.compile(r"(<[^>]+>)")
_TAG_NAME_RE = re.compile(r"</?\s*([a-zA-Z0-9]+)")


def _reserve(stack: List[str]) -> int:
    """Ile znakow trzeba zostawic na domkniecie otwartych tagow."""
    return sum(len(t) + 3 for t in stack)


def _cut_fragment(text: str, avail: int, min_keep: int = 40) -> str:
    """Przycina wezel tekstowy wewnatrz HTML; pilnuje, by nie przeciac encji."""
    if avail <= 0:
        return ""
    window = text[:avail]
    cut = _find_cut(window, int(avail * 0.4))
    piece = window[:cut] if cut > 0 else window
    piece = re.sub(r"&[a-zA-Z#0-9]*$", "", piece)      # nie tnij w srodku &amp;
    piece = _TAIL_JUNK_RE.sub("", piece).rstrip()
    return piece if len(piece) >= min_keep else ""


def smart_truncate_html(s: str, limit: int, soft_ratio: float = 0.6,
                        opts: Optional[HtmlOptions] = None) -> str:
    """
    Przycina HTML do `limit` znakow (liczonych razem ze znacznikami),
    zawsze zwracajac poprawnie domkniety, sensownie zakonczony fragment.
    """
    opts = opts or HtmlOptions()
    if not s:
        return ""
    if limit <= 0 or len(s) <= limit:
        return s

    tokens = [t for t in _TOKEN_RE.split(s) if t]
    out: List[str] = []
    stack: List[str] = []
    length = 0
    snapshot: Optional[Tuple[int, int, List[str]]] = None
    text_cut: Optional[Tuple[List[str], int, List[str]]] = None

    for tok in tokens:
        if tok.startswith("<") and tok.endswith(">"):
            m = _TAG_NAME_RE.match(tok)
            name = m.group(1).lower() if m else ""
            closing = tok.startswith("</")
            if closing:
                new_stack = stack[:-1] if (stack and stack[-1] == name) else list(stack)
            elif name in VOID_TAGS:
                new_stack = list(stack)
            else:
                new_stack = stack + [name]
            if length + len(tok) + _reserve(new_stack) > limit:
                break
            out.append(tok)
            length += len(tok)
            stack = new_stack
            if closing and name in BLOCK_CLOSERS:
                snapshot = (len(out), length, list(stack))
        else:
            avail = limit - _reserve(stack) - length
            if len(tok) <= avail:
                out.append(tok)
                length += len(tok)
            else:
                piece = _cut_fragment(tok, avail)
                if piece:
                    text_cut = (out + [piece], length + len(piece), list(stack))
                break

    floor = limit * soft_ratio
    candidates = []
    if snapshot and snapshot[1] >= floor:
        candidates.append(("snap", snapshot[1]))
    if text_cut and text_cut[1] >= floor:
        candidates.append(("text", text_cut[1]))

    if candidates:
        choice = max(candidates, key=lambda c: c[1])[0]
    elif text_cut:
        choice = "text"
    elif snapshot:
        choice = "snap"
    else:
        choice = "raw"

    if choice == "snap":
        out = out[:snapshot[0]]
        stack = list(snapshot[2])
    elif choice == "text":
        out = list(text_cut[0])
        stack = list(text_cut[2])

    for tag in reversed(stack):
        out.append(f"</{tag}>")

    return _post_process_html("".join(out), opts)


# ---------------------------------------------------------------------------
# 5. WEJSCIE DLA SKRYPTOW FEEDOWYCH
# ---------------------------------------------------------------------------

def build_descriptions(
    raw_desc,
    plain_limit: int = PLAIN_LIMIT,
    html_limit: int = MARKETING_LIMIT,
    fix_chars: bool = True,
    plain_opts: Optional[ConvertOptions] = None,
    html_opts: Optional[HtmlOptions] = None,
) -> Tuple[str, str]:
    """
    Z jednego surowego opisu robi dwie kolumny:

    1. desc_plain     - czysty tekst z enterami, max `plain_limit` znakow
    2. desc_marketing - HTML z bialej listy Cdiscount, max `html_limit` znakow

    Zwraca krotke (desc_plain, desc_marketing).
    """
    if raw_desc is None:
        return "", ""
    source = str(raw_desc)
    if not source.strip():
        return "", ""

    if fix_chars:
        # keep_markup=True: nie zamieniamy &lt; na <, zeby zakodowany tekst
        # nie stal sie nagle prawdziwym tagiem przed sanityzacja.
        source = correct_text(source, keep_markup=True)

    if plain_opts is None:
        # Domyslnie tak samo jak dla wersji HTML: obie kolumny ida na ten sam
        # marketplace, ktory nie chce w opisie adresow ani e-maili.
        plain_opts = ConvertOptions(strip_urls=True, strip_emails=True)

    plain = html_to_plain(source, plain_opts)
    marketing = html_to_marketing(source, html_opts)

    if fix_chars:
        plain = correct_text(plain) or ""
        marketing = correct_text(marketing, keep_markup=True) or ""

    plain = smart_truncate_text(plain, plain_limit)
    marketing = smart_truncate_html(marketing, html_limit, opts=html_opts)
    return plain, marketing
# <<< WBUDOWANY-BLOK-KONIEC: czyszczenie-opisow

MAX_WORKERS = 10
DEFAULT_OUTPUT_DIR = os.path.join(os.path.expanduser("~"), "Downloads")
BASE_FIELDS = [
    "id",
    "id_bl",
    "url",
    "price",
    "avail",
    "weight",
    "stock",
    "cat",
    "name",
    "desc",
    "desc_plain",
    "desc_marketing",
]
# Limity narzucone przez Cdiscount: 2000 znakow czystego tekstu, 5000 HTML-a.
PLAIN_DESC_LIMIT = PLAIN_LIMIT
MARKETING_DESC_LIMIT = MARKETING_LIMIT
# W feedach id_bl siedzi jako <attrs><a name="id_bl">..., wyciągamy go na stałą kolumnę.
ID_BL_ATTR = "id_bl"
CSV_DELIMITER = "|"
FILTER_MODE_INCLUDE = "include"
FILTER_MODE_EXCLUDE = "exclude"


def clean_text(text):
    """Zastępuje znaki nowej linii i inne białe znaki pojedynczą spacją."""
    if not text:
        return ""
    return " ".join(text.split())


def clean_field(text):
    """
    clean_text + korekta znaków — dla pól tekstowych (cat, name, desc, atrybuty).

    URL-i, cen i identyfikatorów nie ruszamy: podmiana encji mogłaby uszkodzić
    adres, a emoji tam nie występują.
    """
    return correct_text(clean_text(text)) or ""


def inner_markup(element):
    """
    Zwraca surową treść elementu razem z HTML-em.

    Opis w feedzie bywa tekstem/CDATA (wtedy .text wystarcza) albo prawdziwymi
    pod-elementami XML (wtedy trzeba je zserializować z powrotem).
    """
    if element is None:
        return ""
    parts = [element.text or ""]
    for child in element:
        try:
            parts.append(ET.tostring(child, encoding="unicode"))
        except Exception:
            parts.append("".join(child.itertext()))
    return "".join(parts)


def load_filter_ids(file_path):
    """
    Wczytuje listę ID z pliku CSV/tekstowego (jedna wartość w wierszu, brana z
    pierwszej kolumny). Zwraca (zbiór_id, None) lub (set(), powod_bledu).
    """
    try:
        ids = set()
        with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
            for line in handle:
                token = line.strip()
                if not token:
                    continue
                for delimiter in (",", ";", "\t", "|"):
                    if delimiter in token:
                        token = token.split(delimiter, 1)[0].strip()
                        break
                if token:
                    ids.add(token)
        # Usuń typowy nagłówek kolumny, jeśli się pojawił.
        ids.discard("id")
        ids.discard("ID")
        ids.discard("Id")
        return ids, None
    except Exception as error:
        return set(), str(error)


def is_available(row):
    """Sprawdza, czy wiersz ma avail = 1 (produkt dostępny od ręki)."""
    return (row.get("avail") or "").strip() == "1"


def recalculate_columns(rows):
    """
    Przelicza zbiór atrybutów i maksymalną liczbę obrazów na podstawie
    zachowanych wierszy, aby uniknąć pustych kolumn po odfiltrowaniu.
    """
    attributes = set()
    max_images = 0
    for row in rows:
        for key in row:
            if key in BASE_FIELDS:
                continue
            if key.startswith("image") and key[5:].isdigit():
                max_images = max(max_images, int(key[5:]) + 1)
            else:
                attributes.add(key)
    return attributes, max_images


def download_xml(url, target_path):
    """Pobiera plik XML z podanego URL. Zwraca (True, None) lub (False, powód)."""
    try:
        opener = urllib.request.build_opener()
        opener.addheaders = [("User-agent", "Mozilla/5.0")]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(url, target_path)
        return True, None
    except Exception as error:
        return False, str(error)


def parse_xml(file_path):
    """
    Parsuje plik XML i ekstrahuje dane produktowe.
    Zwraca (atrybuty, maks_obrazow, dane, None) lub ([], 0, [], powod_bledu).

    Pola tekstowe przechodzą korektę znaków (encje HTML -> polskie litery,
    usuwanie emoji) — tak samo jak w "tlumaczenia v2.py". Do każdego wiersza
    dochodzą kolumny desc_plain i desc_marketing wyliczone z surowej treści
    <desc> (razem z HTML-em, jeśli feed go tam trzyma).
    """
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        attributes = set()
        max_images = 0
        rows = []

        for element in root.findall("o"):
            cat_elem = element.find("cat")
            name_elem = element.find("name")
            desc_elem = element.find("desc")

            row = {
                "id": element.get("id"),
                "id_bl": "",
                "url": element.get("url"),
                "price": element.get("price"),
                "avail": element.get("avail"),
                "weight": element.get("weight"),
                "stock": element.get("stock"),
                "cat": clean_field(cat_elem.text) if cat_elem is not None else "",
                "name": clean_field(name_elem.text) if name_elem is not None else "",
                "desc": clean_field(desc_elem.text) if desc_elem is not None else "",
            }

            desc_plain, desc_marketing = build_descriptions(
                inner_markup(desc_elem),
                plain_limit=PLAIN_DESC_LIMIT,
                html_limit=MARKETING_DESC_LIMIT,
            )
            row["desc_plain"] = desc_plain
            row["desc_marketing"] = desc_marketing

            attrs_elem = element.find("attrs")
            if attrs_elem is not None:
                for attr in attrs_elem.findall("a"):
                    attr_name = attr.get("name")
                    if not attr_name:
                        continue
                    if attr_name == ID_BL_ATTR:
                        row["id_bl"] = clean_text(attr.text)
                        continue
                    if attr_name in BASE_FIELDS:
                        continue
                    attributes.add(attr_name)
                    row[attr_name] = clean_field(attr.text)

            images_in_row = 0
            imgs_elem = element.find("imgs")
            if imgs_elem is not None:
                main_image = imgs_elem.find("main")
                if main_image is not None and main_image.get("url"):
                    row["image0"] = main_image.get("url")
                    images_in_row = 1

                start_index = 1 if "image0" in row else 0
                for i, img in enumerate(imgs_elem.findall("i"), start=start_index):
                    if img.get("url"):
                        row[f"image{i}"] = img.get("url")
                        images_in_row = max(images_in_row, i + 1)

            max_images = max(max_images, images_in_row)
            rows.append(row)

        return sorted(attributes), max_images, rows, None

    except FileNotFoundError as error:
        return [], 0, [], f"Nie znaleziono pliku: {file_path} ({error})"
    except ET.ParseError as error:
        return [], 0, [], f"Błąd parsowania XML w {os.path.basename(file_path)}: {error}"
    except Exception as error:
        return [], 0, [], f"Nieoczekiwany błąd parsowania: {error}"


def write_csv(rows, attributes, max_images, file_path):
    """Zapisuje połączone dane do jednego pliku CSV."""
    fields = BASE_FIELDS + list(attributes) + [f"image{i}" for i in range(max_images)]
    try:
        with open(file_path, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=fields, delimiter=CSV_DELIMITER, extrasaction="ignore"
            )
            writer.writeheader()
            writer.writerows(rows)
        return True, None
    except Exception as error:
        return False, str(error)


def save_error_report(download_errors, parse_errors, output_dir):
    """Zapisuje raport błędów pobierania/parsowania do pliku XLSX. Zwraca ścieżkę lub None."""
    if not download_errors and not parse_errors:
        return None

    workbook = openpyxl.Workbook()
    sheets = []

    if download_errors:
        sheet = workbook.active
        sheet.title = "Bledy pobierania"
        sheet.append(["Nieudany URL", "Powód błędu", "Plik"])
        for url, reason, file_name in download_errors:
            sheet.append([url, reason, file_name])
        sheets.append(sheet)
    else:
        workbook.remove(workbook.active)

    if parse_errors:
        sheet = workbook.create_sheet("Bledy parsowania")
        sheet.append(["URL", "Powód błędu", "Plik"])
        for url, reason, file_name in parse_errors:
            sheet.append([url, reason, file_name])
        sheets.append(sheet)

    for sheet in sheets:
        for column in sheet.columns:
            max_length = 0
            letter = get_column_letter(column[0].column)
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[letter].width = min(max_length + 2, 80)

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    report_path = os.path.join(output_dir, f"RAPORT_BLEDOW_XMLCSV_{timestamp}.xlsx")
    workbook.save(report_path)
    return report_path


def download_and_parse_url(url):
    """Pobiera i parsuje jeden URL. Przeznaczone do uruchamiania w osobnym wątku."""
    temp_dir = tempfile.gettempdir()
    file_name = os.path.basename(urlparse(url).path) or f"feed_{abs(hash(url))}.xml"
    base_name = os.path.splitext(file_name)[0]
    temp_stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    local_xml_path = os.path.join(temp_dir, f"temp_{base_name}_{temp_stamp}.xml")

    success, error_message = download_xml(url, local_xml_path)
    if not success:
        return "download_error", (url, error_message, file_name)

    attributes, max_images, rows, parse_error = parse_xml(local_xml_path)

    try:
        os.remove(local_xml_path)
    except OSError:
        pass

    if parse_error:
        return "parse_error", (url, parse_error, file_name)

    if not rows:
        return "parse_error", (url, "Brak elementów <o> po parsowaniu", file_name)

    return "success", (rows, attributes, max_images, base_name, file_name)


def enable_windows_acrylic(win_id):
    if os.name != "nt":
        return False

    class ACCENT_POLICY(ctypes.Structure):
        _fields_ = [
            ("AccentState", ctypes.c_int),
            ("AccentFlags", ctypes.c_int),
            ("GradientColor", ctypes.c_uint32),
            ("AnimationId", ctypes.c_int),
        ]

    class WINDOWCOMPOSITIONATTRIBDATA(ctypes.Structure):
        _fields_ = [
            ("Attribute", ctypes.c_int),
            ("Data", ctypes.c_void_p),
            ("SizeOfData", ctypes.c_size_t),
        ]

    ACCENT_ENABLE_ACRYLICBLURBEHIND = 4
    WCA_ACCENT_POLICY = 19

    try:
        user32 = ctypes.windll.user32
        set_window_composition_attribute = user32.SetWindowCompositionAttribute
    except Exception:
        return False

    hwnd = wintypes.HWND(int(win_id))
    accent = ACCENT_POLICY(
        AccentState=ACCENT_ENABLE_ACRYLICBLURBEHIND,
        AccentFlags=2,
        GradientColor=0xEEF5E8FF,
        AnimationId=0,
    )
    data = WINDOWCOMPOSITIONATTRIBDATA(
        Attribute=WCA_ACCENT_POLICY,
        Data=ctypes.cast(ctypes.pointer(accent), ctypes.c_void_p),
        SizeOfData=ctypes.sizeof(accent),
    )

    try:
        return bool(set_window_composition_attribute(hwnd, ctypes.byref(data)))
    except Exception:
        return False


class ProcessorThread(QThread):
    progress_signal = Signal(str, float, str)
    done_signal = Signal(dict)
    error_signal = Signal(str)

    def __init__(
        self,
        urls,
        output_dir,
        filter_ids=None,
        avail_only=False,
        filter_mode=FILTER_MODE_INCLUDE,
    ):
        super().__init__()
        self.urls = urls
        self.output_dir = output_dir
        self.filter_ids = set(filter_ids) if filter_ids else set()
        self.avail_only = bool(avail_only)
        self.filter_mode = filter_mode

    def _emit_progress(self, message, value, tone="normal"):
        self.progress_signal.emit(message, value, tone)

    def run(self):
        try:
            if not self.urls:
                self.error_signal.emit("Podaj co najmniej jeden URL pliku XML.")
                return

            if not os.path.exists(self.output_dir):
                try:
                    os.makedirs(self.output_dir, exist_ok=True)
                except Exception as error:
                    self.error_signal.emit(
                        f"Nie można utworzyć katalogu zapisu:\n{self.output_dir}\n{error}"
                    )
                    return

            all_rows = []
            all_attributes = set()
            global_max_images = 0
            base_names = []
            download_errors = []
            parse_errors = []
            total = len(self.urls)

            self._emit_progress(
                f"Rozpoczynam przetwarzanie {total} linków (max {MAX_WORKERS} wątków)...", 0.02
            )

            with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                futures = [executor.submit(download_and_parse_url, url) for url in self.urls]
                for index, future in enumerate(concurrent.futures.as_completed(futures), start=1):
                    progress = 0.02 + (0.93 * index / total)
                    try:
                        status, data = future.result()
                        if status == "success":
                            rows, attributes, max_images, base_name, file_name = data
                            all_rows.extend(rows)
                            all_attributes.update(attributes)
                            global_max_images = max(global_max_images, max_images)
                            base_names.append(base_name)
                            self._emit_progress(f"Pobrano {index}/{total}: {file_name}", progress)
                        elif status == "download_error":
                            download_errors.append(data)
                            self._emit_progress(
                                f"Błąd pobierania {index}/{total}: {data[2]}", progress, "warn"
                            )
                        else:
                            parse_errors.append(data)
                            self._emit_progress(
                                f"Błąd parsowania {index}/{total}: {data[2]}", progress, "warn"
                            )
                    except Exception as error:
                        parse_errors.append(("?", f"Błąd krytyczny wątku: {error}", "?"))
                        self._emit_progress(f"Błąd krytyczny wątku {index}/{total}", progress, "warn")

            error_count = len(download_errors) + len(parse_errors)

            rows_before_filter = len(all_rows)
            id_filter_applied = bool(self.filter_ids)
            rows_after_avail = rows_before_filter

            if self.avail_only:
                all_rows = [row for row in all_rows if is_available(row)]
                rows_after_avail = len(all_rows)
                self._emit_progress(
                    f"Filtr avail=1: pozostawiono {rows_after_avail}/{rows_before_filter} wierszy",
                    0.95,
                )

            exclude_mode = self.filter_mode == FILTER_MODE_EXCLUDE
            if id_filter_applied:
                if exclude_mode:
                    all_rows = [row for row in all_rows if row.get("id") not in self.filter_ids]
                else:
                    all_rows = [row for row in all_rows if row.get("id") in self.filter_ids]
                mode_name = "wyklucz" if exclude_mode else "dołącz"
                self._emit_progress(
                    f"Filtr ID ({mode_name}): pozostawiono "
                    f"{len(all_rows)}/{rows_after_avail} wierszy",
                    0.96,
                )

            if self.avail_only or id_filter_applied:
                all_attributes, global_max_images = recalculate_columns(all_rows)

            if not all_rows:
                report_path = save_error_report(download_errors, parse_errors, self.output_dir)
                if rows_before_filter:
                    message = (
                        f"Wszystkie {rows_before_filter} wierszy zostało odfiltrowanych.\n\n"
                        f"Filtr avail=1: {'tak' if self.avail_only else 'nie'}\n"
                        f"Filtr ID: "
                        f"{('wyklucz' if exclude_mode else 'dołącz') if id_filter_applied else 'nie'}"
                    )
                else:
                    message = (
                        "Nie udało się pobrać ani sparsować danych z żadnego podanego URL.\n\n"
                        f"Błędy pobierania: {len(download_errors)}\n"
                        f"Błędy parsowania: {len(parse_errors)}"
                    )
                if report_path:
                    message += f"\n\nRaport błędów:\n{os.path.basename(report_path)}"
                self.done_signal.emit(
                    {
                        "ok": False,
                        "title": "Brak danych",
                        "message": message,
                        "tone": "warn",
                        "progress": 0.0,
                    }
                )
                return

            combined_name = "_".join(base_names)
            if len(combined_name) > 100:
                combined_name = f"{base_names[0]}_and_{len(base_names) - 1}_more"

            timestamp = datetime.now().strftime("%d%m%y-%H%M%S")
            csv_path = os.path.join(self.output_dir, f"{combined_name}_{timestamp}.csv")

            self._emit_progress("Zapisywanie połączonych danych...", 0.98)
            save_ok, save_error = write_csv(
                all_rows, sorted(all_attributes), global_max_images, csv_path
            )

            report_path = save_error_report(download_errors, parse_errors, self.output_dir)

            summary_lines = [
                f"Przetworzone pliki XML: {len(base_names)}/{total}",
                f"Wiersze produktów: {len(all_rows)}",
                f"Kolumny atrybutów: {len(all_attributes)}",
                f"Maks. liczba obrazów: {global_max_images}",
                f"Błędy pobierania: {len(download_errors)}",
                f"Błędy parsowania: {len(parse_errors)}",
            ]

            plain_lengths = [len(row.get("desc_plain") or "") for row in all_rows]
            html_lengths = [len(row.get("desc_marketing") or "") for row in all_rows]
            summary_lines.append(
                f"desc_plain: max {max(plain_lengths) if plain_lengths else 0}"
                f"/{PLAIN_DESC_LIMIT} zn., "
                f"desc_marketing: max {max(html_lengths) if html_lengths else 0}"
                f"/{MARKETING_DESC_LIMIT} zn."
            )
            summary_lines.append(
                f"Produkty bez opisu: {sum(1 for v in plain_lengths if v == 0)}"
            )

            extra_lines = []
            if self.avail_only:
                extra_lines.append(
                    f"Filtr avail=1: {rows_after_avail}/{rows_before_filter} wierszy"
                )
            if id_filter_applied:
                extra_lines.append(
                    f"Filtr ID ({'wyklucz' if exclude_mode else 'dołącz'}): "
                    f"{len(all_rows)}/{rows_after_avail} wierszy "
                    f"(lista: {len(self.filter_ids)} ID)"
                )
            for offset, line in enumerate(extra_lines):
                summary_lines.insert(2 + offset, line)

            if save_ok:
                summary_lines.append(f"\nZapisano CSV:\n{os.path.abspath(csv_path)}")
            else:
                summary_lines.append(f"\nBłąd zapisu CSV: {save_error}")

            if report_path:
                summary_lines.append(f"Raport błędów: {os.path.basename(report_path)}")

            ok = save_ok and error_count == 0
            self.done_signal.emit(
                {
                    "ok": ok,
                    "title": "Sukces" if ok else "Zakończono z błędami",
                    "message": "\n".join(summary_lines),
                    "tone": "ok" if ok else "warn",
                    "progress": 1.0,
                }
            )
        except Exception as error:
            self.error_signal.emit(str(error))


class GradientBackgroundWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.setObjectName("Root")

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        gradient = QLinearGradient(0, 0, self.width(), self.height())
        gradient.setColorAt(0.0, QColor("#e8daf5"))
        gradient.setColorAt(0.5, QColor("#e0d0f0"))
        gradient.setColorAt(1.0, QColor("#d8c8eb"))
        painter.fillRect(self.rect(), gradient)

        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor("#f5d5e8"))
        painter.drawEllipse(-150, -100, 450, 350)

        painter.setBrush(QColor("#dfc8e8"))
        painter.drawEllipse(self.width() - 300, -80, 400, 300)

        painter.setBrush(QColor("#f8c2df"))
        painter.drawEllipse(self.width() - 240, self.height() - 190, 380, 260)


class GlassCard(QFrame):
    def __init__(self, object_name="GlassCard"):
        super().__init__()
        self.setObjectName(object_name)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Konwerter XML → CSV z opisami")
        self.resize(760, 640)
        self.setMinimumSize(620, 560)

        self.worker = None
        self.filter_ids = set()
        self.filter_path = None

        self.root = GradientBackgroundWidget()
        self.setCentralWidget(self.root)

        self._build_ui()
        self._apply_styles()

    def showEvent(self, event):
        super().showEvent(event)
        enable_windows_acrylic(self.winId())

    def _build_ui(self):
        root_layout = QVBoxLayout(self.root)
        root_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet("QScrollArea { background: transparent; }")

        scroll_host = QWidget()
        scroll_host.setStyleSheet("background: transparent;")
        outer_layout = QVBoxLayout(scroll_host)
        outer_layout.setContentsMargins(22, 22, 22, 22)
        outer_layout.setSpacing(14)

        header = GlassCard("HeaderCard")
        header.setMinimumHeight(120)
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(28, 20, 28, 20)
        header_layout.setSpacing(6)

        title = QLabel("Konwerter XML → CSV z opisami")
        title.setObjectName("Title")
        title.setWordWrap(True)

        subtitle = QLabel(
            "Pobierz wiele feedów XML, połącz je w jeden plik CSV i dołóż dwie "
            "kolumny opisu: desc_plain (czysty tekst, 2000 zn.) oraz "
            "desc_marketing (HTML dla Cdiscount, 5000 zn.)."
        )
        subtitle.setObjectName("Subtitle")
        subtitle.setWordWrap(True)

        header_layout.addWidget(title)
        header_layout.addWidget(subtitle)
        outer_layout.addWidget(header)

        form_card = GlassCard()
        form_layout = QVBoxLayout(form_card)
        form_layout.setContentsMargins(18, 16, 18, 16)
        form_layout.setSpacing(12)

        form_layout.addWidget(
            self._section_label("Linki XML", "Wklej URL-e plików XML, każdy w nowej linii.")
        )
        self.url_input = QPlainTextEdit()
        self.url_input.setPlaceholderText(
            "https://example.com/feed.xml\nhttps://example.com/feed2.xml"
        )
        self.url_input.setMinimumHeight(180)
        self.url_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        form_layout.addWidget(self.url_input, 1)

        outer_layout.addWidget(form_card, 1)

        settings_card = GlassCard()
        settings_layout = QGridLayout(settings_card)
        settings_layout.setContentsMargins(18, 16, 18, 16)
        settings_layout.setHorizontalSpacing(10)
        settings_layout.setVerticalSpacing(10)

        output_label = QLabel("Folder zapisu CSV:")
        output_label.setObjectName("FieldLabel")
        self.output_input = QLineEdit(DEFAULT_OUTPUT_DIR)
        self.output_btn = QPushButton("Wybierz folder")
        self.output_btn.clicked.connect(self.pick_output_dir)

        settings_layout.addWidget(output_label, 0, 0)
        settings_layout.addWidget(self.output_input, 0, 1)
        settings_layout.addWidget(self.output_btn, 0, 2)

        filter_label = QLabel("Filtr ID (CSV, opcjonalnie):")
        filter_label.setObjectName("FieldLabel")
        self.filter_input = QLineEdit()
        self.filter_input.setReadOnly(True)
        self.filter_input.setPlaceholderText("Brak filtra – w CSV znajdą się wszystkie wiersze")

        self.filter_btn = QPushButton("Wybierz plik CSV")
        self.filter_btn.clicked.connect(self.pick_filter_file)
        self.filter_clear_btn = QPushButton("Wyczyść")
        self.filter_clear_btn.clicked.connect(self.clear_filter_file)

        filter_buttons = QWidget()
        filter_buttons_layout = QHBoxLayout(filter_buttons)
        filter_buttons_layout.setContentsMargins(0, 0, 0, 0)
        filter_buttons_layout.setSpacing(8)
        filter_buttons_layout.addWidget(self.filter_btn)
        filter_buttons_layout.addWidget(self.filter_clear_btn)

        settings_layout.addWidget(filter_label, 1, 0)
        settings_layout.addWidget(self.filter_input, 1, 1)
        settings_layout.addWidget(filter_buttons, 1, 2)

        mode_label = QLabel("Tryb filtra ID:")
        mode_label.setObjectName("FieldLabel")
        self.filter_mode_combo = QComboBox()
        self.filter_mode_combo.addItem("Zostaw tylko wskazane ID", FILTER_MODE_INCLUDE)
        self.filter_mode_combo.addItem("Wyklucz wskazane ID", FILTER_MODE_EXCLUDE)
        self.filter_mode_combo.setToolTip(
            "Wskazane ID trafiają do CSV (dołącz) albo są z niego usuwane (wyklucz)."
        )
        settings_layout.addWidget(mode_label, 2, 0)
        settings_layout.addWidget(self.filter_mode_combo, 2, 1, 1, 2)

        self.avail_only_check = QCheckBox("Tylko dostępne produkty (avail = 1)")
        self.avail_only_check.setObjectName("AvailCheck")
        self.avail_only_check.setToolTip(
            "Do CSV trafią wyłącznie wiersze, w których atrybut avail ma wartość 1."
        )
        settings_layout.addWidget(self.avail_only_check, 3, 1, 1, 2)

        settings_layout.setColumnStretch(1, 1)
        outer_layout.addWidget(settings_card)

        self.run_btn = QPushButton("Przetwórz na JEDEN plik CSV")
        self.run_btn.setObjectName("RunButton")
        self.run_btn.clicked.connect(self.run_processing)
        outer_layout.addWidget(self.run_btn)

        self.progress = QProgressBar()
        self.progress.setRange(0, 1000)
        self.progress.setValue(0)
        outer_layout.addWidget(self.progress)

        self.status = QLabel("Gotowy.")
        self.status.setObjectName("Status")
        self.status.setWordWrap(True)
        outer_layout.addWidget(self.status)

        scroll.setWidget(scroll_host)
        root_layout.addWidget(scroll)

    def _section_label(self, title, subtitle):
        wrapper = QWidget()
        layout = QVBoxLayout(wrapper)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        t = QLabel(title)
        t.setObjectName("SectionTitle")
        s = QLabel(subtitle)
        s.setObjectName("SectionSub")
        layout.addWidget(t)
        layout.addWidget(s)
        return wrapper

    def _apply_styles(self):
        self.setStyleSheet(
            """
            QWidget#Root {
                background: transparent;
                color: #4c1636;
                font-family: 'Segoe UI', 'Tahoma', sans-serif;
                font-size: 13px;
            }
            QFrame#HeaderCard {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 150, 200, 100),
                    stop:1 rgba(180, 120, 200, 110));
                border: 1px solid rgba(255,255,255,100);
                border-radius: 22px;
            }
            QFrame#GlassCard {
                background: rgba(255,255,255,50);
                border: 1px solid rgba(255,255,255,80);
                border-radius: 20px;
            }
            QLabel#Title {
                color: #4c1636;
                font-size: 26px;
                font-weight: 800;
                margin-top: 2px;
                line-height: 1.2;
            }
            QLabel#Subtitle {
                color: #6a2a52;
                font-size: 13px;
                line-height: 1.25;
            }
            QLabel#SectionTitle {
                color: #4c1636;
                font-size: 16px;
                font-weight: 700;
            }
            QLabel#SectionSub {
                color: #7d4165;
                font-size: 12px;
            }
            QLabel#FieldLabel {
                color: #5b2142;
                font-weight: 700;
            }
            QComboBox {
                background: rgba(255,255,255,70);
                border: 1px solid rgba(255,255,255,120);
                border-radius: 12px;
                padding: 8px;
                color: #4c1636;
            }
            QComboBox:focus {
                border: 1px solid rgba(255,200,230,180);
                background: rgba(255,255,255,90);
            }
            QComboBox::drop-down {
                border: none;
                width: 24px;
            }
            QComboBox QAbstractItemView {
                background: #f3e4f2;
                color: #4c1636;
                border: 1px solid rgba(255,255,255,150);
                selection-background-color: rgba(255,120,180,140);
                outline: none;
            }
            QPlainTextEdit, QLineEdit {
                background: rgba(255,255,255,70);
                border: 1px solid rgba(255,255,255,120);
                border-radius: 12px;
                padding: 8px;
                color: #4c1636;
                selection-background-color: rgba(255,150,200,150);
            }
            QPlainTextEdit:focus, QLineEdit:focus {
                border: 1px solid rgba(255,200,230,180);
                background: rgba(255,255,255,90);
            }
            QCheckBox#AvailCheck {
                color: #5b2142;
                font-weight: 700;
                spacing: 8px;
            }
            QCheckBox#AvailCheck::indicator {
                width: 18px;
                height: 18px;
                border-radius: 6px;
                border: 1px solid rgba(255,255,255,150);
                background: rgba(255,255,255,70);
            }
            QCheckBox#AvailCheck::indicator:checked {
                background: rgba(255,120,180,180);
                border: 1px solid rgba(255,255,255,180);
            }
            QPushButton {
                background: rgba(255,150,200,120);
                border: 1px solid rgba(255,255,255,150);
                border-radius: 12px;
                color: #4c1636;
                padding: 9px 14px;
                font-weight: 700;
            }
            QPushButton:hover {
                background: rgba(255,150,200,160);
            }
            QPushButton:disabled {
                background: rgba(200,150,180,80);
                color: rgba(76,22,54,120);
            }
            QPushButton#RunButton {
                min-height: 44px;
                border-radius: 14px;
                font-size: 15px;
                background: rgba(255,120,180,140);
            }
            QPushButton#RunButton:hover {
                background: rgba(255,120,180,180);
            }
            QProgressBar {
                background: rgba(255,255,255,60);
                border-radius: 7px;
                border: 1px solid rgba(255,255,255,100);
                min-height: 12px;
                text-align: center;
                color: transparent;
            }
            QProgressBar::chunk {
                border-radius: 7px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255,120,180,180),
                    stop:1 rgba(180,100,160,180));
            }
            QLabel#Status {
                color: #5b2142;
                font-size: 13px;
                font-weight: 600;
                padding-bottom: 6px;
            }
            """
        )

    def pick_output_dir(self):
        selected = QFileDialog.getExistingDirectory(
            self, "Wybierz folder zapisu", self.output_input.text().strip() or DEFAULT_OUTPUT_DIR
        )
        if selected:
            self.output_input.setText(selected)

    def pick_filter_file(self):
        start_dir = self.output_input.text().strip() or DEFAULT_OUTPUT_DIR
        selected, _ = QFileDialog.getOpenFileName(
            self,
            "Wybierz plik CSV z listą ID",
            start_dir,
            "Pliki CSV (*.csv);;Pliki tekstowe (*.txt);;Wszystkie pliki (*)",
        )
        if not selected:
            return

        ids, error = load_filter_ids(selected)
        if error:
            QMessageBox.warning(self, "Błąd", f"Nie można wczytać pliku:\n{error}")
            return
        if not ids:
            QMessageBox.warning(self, "Uwaga", "Wybrany plik nie zawiera żadnych ID.")
            return

        self.filter_path = selected
        self.filter_ids = ids
        self.filter_input.setText(f"{os.path.basename(selected)} — {len(ids)} ID")

    def clear_filter_file(self):
        self.filter_path = None
        self.filter_ids = set()
        self.filter_input.clear()

    def run_processing(self):
        urls = [line.strip() for line in self.url_input.toPlainText().splitlines() if line.strip()]
        output_dir = self.output_input.text().strip() or DEFAULT_OUTPUT_DIR

        if not urls:
            QMessageBox.warning(self, "Błąd", "Musisz podać co najmniej jeden URL pliku XML.")
            return

        self.output_input.setText(output_dir)
        self.run_btn.setEnabled(False)
        self.run_btn.setText("Przetwarzanie...")
        self.status.setText("Start przetwarzania...")
        self.progress.setValue(10)

        self.worker = ProcessorThread(
            urls,
            output_dir,
            self.filter_ids,
            self.avail_only_check.isChecked(),
            self.filter_mode_combo.currentData(),
        )
        self.worker.progress_signal.connect(self.on_progress)
        self.worker.done_signal.connect(self.on_done)
        self.worker.error_signal.connect(self.on_error)
        self.worker.start()

    def on_progress(self, message, value, tone):
        color = "#a35300" if tone == "warn" else "#8d3b68"
        self.status.setStyleSheet(f"color: {color}; font-weight: 600;")
        self.status.setText(message)
        self.progress.setValue(max(0, min(1000, int(value * 1000))))

    def on_done(self, payload):
        tone = payload.get("tone", "ok")
        color = "#1f7a4c" if tone == "ok" else "#a35300"
        self.status.setStyleSheet(f"color: {color}; font-weight: 700;")
        self.status.setText(payload.get("title", "Zakończono"))
        self.progress.setValue(int(payload.get("progress", 1.0) * 1000))

        if payload.get("ok", False):
            QMessageBox.information(self, payload.get("title", "Sukces"), payload.get("message", ""))
        else:
            QMessageBox.warning(self, payload.get("title", "Uwaga"), payload.get("message", ""))

        self._reset_run_button()

    def on_error(self, message):
        self.status.setStyleSheet("color: #a35300; font-weight: 700;")
        self.status.setText("Błąd krytyczny")
        self.progress.setValue(0)
        self._reset_run_button()
        QMessageBox.critical(self, "Błąd", message)

    def _reset_run_button(self):
        self.run_btn.setEnabled(True)
        self.run_btn.setText("Przetwórz na JEDEN plik CSV")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
